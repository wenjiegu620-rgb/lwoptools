import os
from datetime import datetime
from typing import List

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


_HEADER_FILL = PatternFill("solid", fgColor="2F5496")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_ALT_FILL = PatternFill("solid", fgColor="DCE6F1")


def _header_row(ws, values: list, col_widths: list = None):
    ws.append(values)
    for i, cell in enumerate(ws[ws.max_row], start=1):
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    if col_widths:
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w


def generate_report(
    project_id: str,
    total_available: int,
    results: list,
    output_dir: str,
) -> str:
    """
    生成 Excel 交付报告，返回报告文件路径。

    results: process_case 返回值的列表，每项为
        {"task_name": str, "episode_uuid": str, "files": [str, ...]}
    """
    wb = openpyxl.Workbook()

    # ── Sheet 1: 总览 ──────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "总览"

    unique_tasks = sorted(set(r["task_name"] for r in results))
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    filename_ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    summary_rows = [
        ("项目 ID", project_id),
        ("下载时间", timestamp),
        ("平台打包完成总数", total_available),
        ("本次下载数量", len(results)),
        ("涉及 Task 数", len(unique_tasks)),
        ("Task 列表", "、".join(unique_tasks)),
        ("输出目录", os.path.abspath(output_dir)),
    ]

    ws_summary.column_dimensions["A"].width = 22
    ws_summary.column_dimensions["B"].width = 60

    for i, (key, val) in enumerate(summary_rows, start=1):
        ws_summary.cell(row=i, column=1, value=key).font = Font(bold=True)
        ws_summary.cell(row=i, column=2, value=str(val))

    # ── Sheet 2: 文件清单 ──────────────────────────────────────────
    ws_manifest = wb.create_sheet("文件清单")
    _header_row(
        ws_manifest,
        ["Task Name", "Episode UUID", "文件路径"],
        [30, 40, 70],
    )

    row_num = 2
    for r in results:
        for file_path in r["files"]:
            fill = _ALT_FILL if row_num % 2 == 0 else None
            for col, val in enumerate([r["task_name"], r["episode_uuid"], file_path], start=1):
                cell = ws_manifest.cell(row=row_num, column=col, value=val)
                if fill:
                    cell.fill = fill
            row_num += 1

    report_path = os.path.join(output_dir, f"delivery_report_{filename_ts}.xlsx")
    wb.save(report_path)
    return report_path
