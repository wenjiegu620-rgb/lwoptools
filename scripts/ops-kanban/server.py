#!/usr/bin/env python3
"""
排期看板 本地服务器
运行: python3 server.py
访问: http://localhost:8000
"""
import json, os, re, glob, subprocess, uuid as _uuid, time, threading
from collections import defaultdict
from dotenv import load_dotenv
import openpyxl
load_dotenv(os.path.join(os.path.dirname(__file__), ".env"))
from datetime import date, datetime, timedelta
from flask import Flask, request, jsonify, Response, send_from_directory, session, redirect
from clickhouse_driver import Client
import pymysql
import urllib.request, urllib.error, ssl
from metrics_helpers import dedupe_case_rows, aggregate_collect_rows, calc_online_hours

# ─── UUID 校验 ────────────────────────────────────────────────
UUID_RE = re.compile(r'^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$', re.I)

# ─── 内存缓存 ─────────────────────────────────────────────────
_cache = {}  # key -> (value, expire_ts)

def _cache_get(key):
    entry = _cache.get(key)
    if entry and entry[1] > time.time():
        return entry[0]
    return None

def _cache_set(key, value, ttl):
    _cache[key] = (value, time.time() + ttl)

# ─── 磁盘缓存（重启不丢失，用于累计数据） ──────────────────────
import pickle
_DISK_CACHE_DIR = os.path.join(os.path.dirname(__file__), "vcache")
os.makedirs(_DISK_CACHE_DIR, exist_ok=True)

def _disk_key_path(key):
    safe = key.replace(":", "_").replace("/", "_")
    return os.path.join(_DISK_CACHE_DIR, safe + ".pkl")

def _disk_cache_get(key):
    path = _disk_key_path(key)
    try:
        with open(path, "rb") as f:
            value, expire_ts = pickle.load(f)
        if expire_ts > time.time():
            return value
    except Exception:
        pass
    return None

def _disk_cache_set(key, value, ttl):
    path = _disk_key_path(key)
    try:
        with open(path, "wb") as f:
            pickle.dump((value, time.time() + ttl), f)
    except Exception:
        pass

def safe_uuids(ids):
    return [i for i in ids if UUID_RE.match(i)]

# ─── Clickhouse（项目搜索） ───────────────────────────────────
CH = dict(host=os.environ["CH_HOST"], port=int(os.environ.get("CH_PORT", 9000)),
          user=os.environ["CH_USER"], password=os.environ["CH_PASSWORD"], database="asset")

def ck():
    return Client(**CH)

# ─── MySQL（存量查询，与 delivery tracker 完全一致） ──────────
MYSQL = dict(
    host=os.environ["MYSQL_HOST"], port=int(os.environ.get("MYSQL_PORT", 3306)),
    user=os.environ["MYSQL_USER"], password=os.environ["MYSQL_PASSWORD"],
    database="asset", charset="utf8mb4",
    cursorclass=pymysql.cursors.DictCursor,
    connect_timeout=30,
    read_timeout=300,
    write_timeout=300,
)

def mysql():
    return pymysql.connect(**MYSQL)

# 采集统计中排除的废弃项目（新库 project_uuid）
_EXCLUDED_PROJECT_UUIDS = {'73234c27-bb25-4831-98ac-d2d762ec0175'}
_EXCL_UUID_CLAUSE = "AND hc.project_uuid NOT IN (" + ",".join(f"'{u}'" for u in _EXCLUDED_PROJECT_UUIDS) + ")"

# ─── MySQL 新库（human_case，0403+ 新项目） ───────────────────
MYSQL_NEW = dict(
    host=os.environ.get("MYSQL_NEW_HOST", "rr-uf6y79x928m716yju.mysql.rds.aliyuncs.com"), port=int(os.environ.get("MYSQL_NEW_PORT", 3306)),
    user=os.environ.get("MYSQL_NEW_USER", "human_case_readonly"), password=os.environ["MYSQL_NEW_PASSWORD"],
    database=os.environ.get("MYSQL_NEW_DB", "human_case"), charset="utf8mb4",
    cursorclass=pymysql.cursors.DictCursor,
    connect_timeout=30, read_timeout=300, write_timeout=300,
)

def mysql_new():
    return pymysql.connect(**MYSQL_NEW)


def _query_new_db_daily(date_str, group=None):
      """从新库查当日采集/质检数据。核心修复：所有统计先按 case 去重，再按 producer 聚合。"""
      conn = mysql_new()
      try:
          with conn.cursor() as cur:
              group_clause = "AND tag_info.vendor = %s" if group else ""

              tag_info_sql = """
                  SELECT
                      t.human_case_id,
                      MAX(JSON_UNQUOTE(JSON_EXTRACT(t.value, '$.producer.producer_name'))) AS producer,
                      MAX(JSON_UNQUOTE(JSON_EXTRACT(t.value, '$.producer.producer_group'))) AS vendor,
                      MAX(IFNULL(CAST(JSON_EXTRACT(t.value, '$.data_info.duration') AS DECIMAL(10,2)), 0)) AS vsec
                  FROM human_case_tag t
                  WHERE t.type = 'produce_tags'
                  GROUP BY t.human_case_id
              """

              prod_day_sql = """
                  SELECT
                      hcn.human_case_id,
                      MIN(hcn.node_created_at) AS t_start,
                      MAX(hcn.node_updated_at) AS t_end
                  FROM human_case_node hcn
                  WHERE hcn.node_name = 'human_case_produce_complete'
                    AND hcn.node_status = 3
                    AND DATE(hcn.node_updated_at) = %s
                  GROUP BY hcn.human_case_id
              """

              qc_day_sql = """
                  SELECT
                      raw.human_case_id,
                      MAX(raw.passed) AS passed
                  FROM (
                      SELECT
                          hcn.human_case_id,
                          CASE WHEN hcn.node_status = 3 THEN 1 ELSE 0 END AS passed
                      FROM human_case_node hcn
                      WHERE hcn.node_name = 'human_case_sampling'
                        AND hcn.node_status IN (3, 4)
                        AND DATE(hcn.node_updated_at) = %s

                      UNION ALL

                      SELECT
                          hcn.human_case_id,
                          CASE WHEN hcn.node_status = 3 THEN 1 ELSE 0 END AS passed
                      FROM human_case_node hcn
                      WHERE hcn.node_name = 'human_case_inspect'
                        AND hcn.node_status IN (3, 4)
                        AND DATE(hcn.node_updated_at) = %s
                        AND NOT EXISTS (
                            SELECT 1
                            FROM human_case_node s
                            WHERE s.human_case_id = hcn.human_case_id
                              AND s.node_name = 'human_case_sampling'
                        )
                  ) raw
                  GROUP BY raw.human_case_id
              """

              qc_status_sql = """
                  SELECT
                      hcn_qc.human_case_id,
                      MAX(CASE WHEN hcn_qc.node_name='human_case_sampling' AND hcn_qc.node_status=3 THEN 1 ELSE 0
  END) AS sampling_status3,
                      MAX(CASE WHEN hcn_qc.node_name='human_case_sampling' THEN 1 ELSE 0 END) AS sampling_any,
                      MAX(CASE WHEN hcn_qc.node_name='human_case_inspect' AND hcn_qc.node_status=3 THEN 1 ELSE 0
  END) AS inspect_status3,
                      MAX(CASE WHEN hcn_qc.node_name='human_case_sampling' AND hcn_qc.node_status IN (3,4) THEN 1
  ELSE 0 END) AS sampling_status34,
                      MAX(CASE WHEN hcn_qc.node_name='human_case_inspect' AND hcn_qc.node_status IN (3,4) THEN 1
  ELSE 0 END) AS inspect_status34
                  FROM human_case_node hcn_qc
                  WHERE hcn_qc.node_name IN ('human_case_sampling', 'human_case_inspect')
                  GROUP BY hcn_qc.human_case_id
              """

              args_produce = [date_str] + ([group] if group else [])
              cur.execute(f"""
                  SELECT
                      hc.id AS case_id,
                      'new' AS db_source,
                      tag_info.producer AS producer,
                      tag_info.vendor AS vendor,
                      prod.t_start AS t_start,
                      prod.t_end AS t_end,
                      tag_info.vsec AS vsec
                  FROM human_case hc
                  JOIN ({prod_day_sql}) prod
                    ON prod.human_case_id = hc.id
                  JOIN ({tag_info_sql}) tag_info
                    ON tag_info.human_case_id = hc.id
                  WHERE hc.deleted_at IS NULL
                    {group_clause}
                  ORDER BY tag_info.producer, prod.t_start
              """, args_produce)
              produce_rows = cur.fetchall()

              args_qc = [date_str, date_str] + ([group] if group else [])
              cur.execute(f"""
                  SELECT
                      tag_info.producer AS producer,
                      COUNT(*) AS qc_total_cases,
                      SUM(q.passed) AS qc_passed_cases,
                      SUM(CASE WHEN q.passed = 1 THEN tag_info.vsec ELSE 0 END) / 3600.0 AS qc_h,
                      SUM(tag_info.vsec) / 3600.0 AS qc_total_h
                  FROM human_case hc
                  JOIN ({tag_info_sql}) tag_info
                    ON tag_info.human_case_id = hc.id
                  JOIN ({qc_day_sql}) q
                    ON q.human_case_id = hc.id
                  WHERE hc.deleted_at IS NULL
                    {group_clause}
                  GROUP BY tag_info.producer
              """, args_qc)
              qc_map = {
                  r["producer"]: {
                      "qc_cases": int(r["qc_passed_cases"] or 0),
                      "qc_total": int(r["qc_total_cases"] or 0),
                      "qc_h": float(r["qc_h"] or 0),
                      "qc_total_h": float(r["qc_total_h"] or 0),
                  }
                  for r in cur.fetchall()
              }

              args_cqc = [date_str] + ([group] if group else [])
              cur.execute(f"""
                  SELECT
                      tag_info.producer AS producer,
                      SUM(CASE
                          WHEN COALESCE(ns.sampling_status3, 0) = 1 THEN tag_info.vsec
                          WHEN COALESCE(ns.sampling_any, 0) = 1 THEN 0
                          WHEN COALESCE(ns.inspect_status3, 0) = 1 THEN tag_info.vsec
                          ELSE 0
                      END) / 3600.0 AS collect_qc_h,
                      SUM(CASE
                          WHEN COALESCE(ns.sampling_status34, 0) = 1 THEN tag_info.vsec
                          WHEN COALESCE(ns.sampling_any, 0) = 1 THEN 0
                          WHEN COALESCE(ns.inspect_status34, 0) = 1 THEN tag_info.vsec
                          ELSE 0
                      END) / 3600.0 AS collect_inspected_h
                  FROM human_case hc
                  JOIN ({prod_day_sql}) prod
                    ON prod.human_case_id = hc.id
                  JOIN ({tag_info_sql}) tag_info
                    ON tag_info.human_case_id = hc.id
                  LEFT JOIN ({qc_status_sql}) ns
                    ON ns.human_case_id = hc.id
                  WHERE hc.deleted_at IS NULL
                    {group_clause}
                  GROUP BY tag_info.producer
              """, args_cqc)
              _cqc_rows = cur.fetchall()
              collect_qc_map = {r["producer"]: float(r["collect_qc_h"] or 0) for r in _cqc_rows}
              collect_inspected_map = {r["producer"]: float(r["collect_inspected_h"] or 0) for r in _cqc_rows}

          produce_rows = _dedupe_produce_rows(produce_rows)
          return produce_rows, qc_map, collect_qc_map, collect_inspected_map, {}, set(), {}, {}
      except Exception:
          import traceback
          traceback.print_exc()
          return [], {}, {}, {}, {}, set(), {}, {}
      finally:
          conn.close()


def _query_new_db_cumul(group=None):
      """从新库查累计数据。核心修复：所有累计统计先按 case 去重。"""
      conn = mysql_new()
      try:
          with conn.cursor() as cur:
              group_clause = "AND tag_info.vendor = %s" if group else ""
              args = ([group] if group else [])

              tag_info_sql = """
                  SELECT
                      t.human_case_id,
                      MAX(JSON_UNQUOTE(JSON_EXTRACT(t.value, '$.producer.producer_name'))) AS producer,
                      MAX(JSON_UNQUOTE(JSON_EXTRACT(t.value, '$.producer.producer_group'))) AS vendor,
                      MAX(IFNULL(CAST(JSON_EXTRACT(t.value, '$.data_info.duration') AS DECIMAL(10,2)), 0)) AS vsec
                  FROM human_case_tag t
                  WHERE t.type = 'produce_tags'
                  GROUP BY t.human_case_id
              """

              prod_all_sql = """
                  SELECT
                      hcn.human_case_id,
                      MIN(DATE(hcn.node_updated_at)) AS first_collect_date
                  FROM human_case_node hcn
                  WHERE hcn.node_name = 'human_case_produce_complete'
                    AND hcn.node_status = 3
                  GROUP BY hcn.human_case_id
              """

              qc_all_sql = """
                  SELECT
                      raw.human_case_id,
                      MAX(raw.passed) AS passed
                  FROM (
                      SELECT
                          hcn.human_case_id,
                          CASE WHEN hcn.node_status = 3 THEN 1 ELSE 0 END AS passed
                      FROM human_case_node hcn
                      WHERE hcn.node_name = 'human_case_sampling'
                        AND hcn.node_status IN (3, 4)

                      UNION ALL

                      SELECT
                          hcn.human_case_id,
                          CASE WHEN hcn.node_status = 3 THEN 1 ELSE 0 END AS passed
                      FROM human_case_node hcn
                      WHERE hcn.node_name = 'human_case_inspect'
                        AND hcn.node_status IN (3, 4)
                        AND NOT EXISTS (
                            SELECT 1
                            FROM human_case_node s
                            WHERE s.human_case_id = hcn.human_case_id
                              AND s.node_name = 'human_case_sampling'
                        )
                  ) raw
                  GROUP BY raw.human_case_id
              """

              pending_inspect_sql = """
                  SELECT DISTINCT hcn.human_case_id
                  FROM human_case_node hcn
                  WHERE hcn.node_name = 'human_case_inspect'
                    AND hcn.node_status IN (1, 2)
              """

              pending_sampling_sql = """
                  SELECT DISTINCT hcn.human_case_id
                  FROM human_case_node hcn
                  WHERE hcn.node_name = 'human_case_sampling'
                    AND hcn.node_status IN (1, 2)
              """

              cur.execute(f"""
                  SELECT
                      tag_info.producer AS producer,
                      SUM(tag_info.vsec) / 3600.0 AS total_collect_h,
                      MIN(prod.first_collect_date) AS first_collect_date
                  FROM human_case hc
                  JOIN ({prod_all_sql}) prod
                    ON prod.human_case_id = hc.id
                  JOIN ({tag_info_sql}) tag_info
                    ON tag_info.human_case_id = hc.id
                  WHERE hc.deleted_at IS NULL
                    {group_clause}
                  GROUP BY tag_info.producer
              """, args)
              _vrows = cur.fetchall()
              total_collect_map = {r["producer"]: float(r["total_collect_h"] or 0) for r in _vrows}
              first_collect_map_v = {
                  r["producer"]: str(r["first_collect_date"]) if r["first_collect_date"] else ""
                  for r in _vrows
              }

              cur.execute(f"""
                  SELECT
                      tag_info.producer AS producer,
                      COUNT(*) AS total_qc_total,
                      SUM(q.passed) AS total_qc_passed,
                      SUM(CASE WHEN q.passed = 1 THEN tag_info.vsec ELSE 0 END) / 3600.0 AS total_qc_h,
                      SUM(tag_info.vsec) / 3600.0 AS total_qc_total_h
                  FROM human_case hc
                  JOIN ({tag_info_sql}) tag_info
                    ON tag_info.human_case_id = hc.id
                  JOIN ({qc_all_sql}) q
                    ON q.human_case_id = hc.id
                  WHERE hc.deleted_at IS NULL
                    {group_clause}
                  GROUP BY tag_info.producer
              """, args)
              total_qc_map = {
                  r["producer"]: {
                      "total_qc_h": float(r["total_qc_h"] or 0),
                      "total_qc_passed": int(r["total_qc_passed"] or 0),
                      "total_qc_total": int(r["total_qc_total"] or 0),
                      "total_qc_total_h": float(r["total_qc_total_h"] or 0),
                  }
                  for r in cur.fetchall()
              }

              cur.execute(f"""
                  SELECT
                      tag_info.producer AS producer,
                      SUM(tag_info.vsec) / 3600.0 AS pending_inspect_h
                  FROM human_case hc
                  JOIN ({pending_inspect_sql}) pi
                    ON pi.human_case_id = hc.id
                  JOIN ({tag_info_sql}) tag_info
                    ON tag_info.human_case_id = hc.id
                  WHERE hc.deleted_at IS NULL
                    {group_clause}
                  GROUP BY tag_info.producer
              """, args)
              pending_inspect_map = {
                  r["producer"]: float(r["pending_inspect_h"] or 0)
                  for r in cur.fetchall()
              }

              cur.execute(f"""
                  SELECT
                      tag_info.producer AS producer,
                      SUM(tag_info.vsec) / 3600.0 AS pending_sampling_h
                  FROM human_case hc
                  JOIN ({pending_sampling_sql}) ps
                    ON ps.human_case_id = hc.id
                  JOIN ({tag_info_sql}) tag_info
                    ON tag_info.human_case_id = hc.id
                  WHERE hc.deleted_at IS NULL
                    {group_clause}
                  GROUP BY tag_info.producer
              """, args)
              pending_sampling_map = {
                  r["producer"]: float(r["pending_sampling_h"] or 0)
                  for r in cur.fetchall()
              }

          return total_collect_map, total_qc_map, pending_inspect_map, pending_sampling_map, first_collect_map_v
      except Exception:
          return {}, {}, {}, {}, {}
      finally:
          conn.close()


def _merge_float_map(m1, m2):
    merged = dict(m1)
    for p, v in m2.items():
        merged[p] = merged.get(p, 0.0) + v
    return merged

def _merge_qc_map(m1, m2):
    merged = dict(m1)
    for p, v in m2.items():
        if p in merged:
            all_keys = set(merged[p]) | set(v)
            merged[p] = {k: merged[p].get(k, 0) + v.get(k, 0) for k in all_keys}
        else:
            merged[p] = dict(v)
    return merged

def _merge_date_map(m1, m2):
    merged = dict(m1)
    for p, v in m2.items():
        if p not in merged or (v and (not merged[p] or v < merged[p])):
            merged[p] = v
    return merged

def _dedupe_produce_rows(rows):
      """按 (db_source, case_id) 去重，避免同一 case 被重复累计"""
      merged = {}

      for row in rows or []:
          case_id = str(row.get("case_id") or "")
          db_source = row.get("db_source") or "unknown"

          # 兼容旧缓存里没有 case_id 的情况，退化成弱去重
          if case_id:
              key = (db_source, case_id)
          else:
              key = (
                  db_source,
                  row.get("producer"),
                  row.get("vendor") or row.get("producer_group"),
                  row.get("t_start"),
                  row.get("t_end"),
                  float(row.get("vsec") or 0),
              )

          if key not in merged:
              merged[key] = dict(row)
              continue

          cur = merged[key]

          if row.get("t_start") and (not cur.get("t_start") or row["t_start"] < cur["t_start"]):
              cur["t_start"] = row["t_start"]
          if row.get("t_end") and (not cur.get("t_end") or row["t_end"] > cur["t_end"]):
              cur["t_end"] = row["t_end"]

          cur_vsec = float(cur.get("vsec") or 0)
          row_vsec = float(row.get("vsec") or 0)
          if row_vsec > cur_vsec:
              cur["vsec"] = row["vsec"]

          if not cur.get("producer") and row.get("producer"):
              cur["producer"] = row["producer"]
          if not cur.get("vendor") and row.get("vendor"):
              cur["vendor"] = row["vendor"]

      result = list(merged.values())
      result.sort(key=lambda r: (
          str(r.get("producer") or ""),
          r.get("t_start") or r.get("t_end") or datetime.min
      ))
      return result

def _query_new_db_week_qc(week_start, date_str, producers):
      """从新库查指定采集员在本周（week_start~date_str）的质检通过时长，先按 case 去重。"""
      if not producers:
          return {}

      conn = mysql_new()
      try:
          with conn.cursor() as cur:
              ph = ",".join(["%s"] * len(producers))

              tag_info_sql = """
                  SELECT
                      t.human_case_id,
                      MAX(JSON_UNQUOTE(JSON_EXTRACT(t.value, '$.producer.producer_name'))) AS producer,
                      MAX(IFNULL(CAST(JSON_EXTRACT(t.value, '$.data_info.duration') AS DECIMAL(10,2)), 0)) AS vsec
                  FROM human_case_tag t
                  WHERE t.type = 'produce_tags'
                  GROUP BY t.human_case_id
              """

              qc_week_sql = """
                  SELECT
                      raw.human_case_id,
                      MAX(raw.passed) AS passed
                  FROM (
                      SELECT
                          hcn.human_case_id,
                          CASE WHEN hcn.node_status = 3 THEN 1 ELSE 0 END AS passed
                      FROM human_case_node hcn
                      WHERE hcn.node_name = 'human_case_sampling'
                        AND hcn.node_status IN (3, 4)
                        AND DATE(hcn.node_updated_at) BETWEEN %s AND %s

                      UNION ALL

                      SELECT
                          hcn.human_case_id,
                          CASE WHEN hcn.node_status = 3 THEN 1 ELSE 0 END AS passed
                      FROM human_case_node hcn
                      WHERE hcn.node_name = 'human_case_inspect'
                        AND hcn.node_status IN (3, 4)
                        AND DATE(hcn.node_updated_at) BETWEEN %s AND %s
                        AND NOT EXISTS (
                            SELECT 1
                            FROM human_case_node s
                            WHERE s.human_case_id = hcn.human_case_id
                              AND s.node_name = 'human_case_sampling'
                        )
                  ) raw
                  GROUP BY raw.human_case_id
              """

              cur.execute(f"""
                  SELECT
                      tag_info.producer AS producer,
                      SUM(CASE WHEN q.passed = 1 THEN tag_info.vsec ELSE 0 END) / 3600.0 AS week_qc_h
                  FROM human_case hc
                  JOIN ({tag_info_sql}) tag_info
                    ON tag_info.human_case_id = hc.id
                  JOIN ({qc_week_sql}) q
                    ON q.human_case_id = hc.id
                  WHERE hc.deleted_at IS NULL
                    AND tag_info.producer IN ({ph})
                  GROUP BY tag_info.producer
              """, [week_start, date_str, week_start, date_str] + producers)

              return {r["producer"]: float(r["week_qc_h"] or 0) for r in cur.fetchall()}
      except Exception:
          return {}
      finally:
          conn.close()


def _query_vendor_week_stats(week_start, week_end):
    """查指定周范围内每个供应商组的采集时长、活跃采集员数、质检通过时长/总时长。
    老库（human_cases / human_case_nodes）+ 新库（human_case / human_case_node）都查，按 vendor 合并。
    """
    result = {}

    def _merge(vendor, collect_h, active, qc_pass_h, qc_total_h):
        if not vendor:
            return
        if vendor not in result:
            result[vendor] = {'vendor': vendor, 'collect_h': 0.0,
                              'active_collectors': set(), 'qc_pass_h': 0.0, 'qc_total_h': 0.0}
        result[vendor]['collect_h']    += collect_h
        result[vendor]['active_collectors'] |= active   # set of producer names
        result[vendor]['qc_pass_h']    += qc_pass_h
        result[vendor]['qc_total_h']   += qc_total_h

    # ── 老库 ──────────────────────────────────────────────────────
    try:
        conn = mysql()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT hc.produced_by_group AS vendor,
                       SUM(IFNULL(hc.video_seconds, 0)) / 3600.0 AS collect_h,
                       GROUP_CONCAT(DISTINCT hc.producer) AS producers
                FROM human_cases hc
                JOIN human_case_nodes hcn ON hcn.human_case_id = hc.id
                WHERE hcn.node_name = 'human_case_upload'
                  AND hcn.node_status = 3
                  AND DATE(hcn.node_updated_at) BETWEEN %s AND %s
                  AND hc.deleted_at IS NULL
                  AND hc.produced_by_group IS NOT NULL AND hc.produced_by_group != ''
                GROUP BY vendor
            """, [week_start, week_end])
            for r in cur.fetchall():
                producers = set((r['producers'] or '').split(',')) if r['producers'] else set()
                _merge(r['vendor'], float(r['collect_h'] or 0), producers, 0.0, 0.0)

            cur.execute("""
                SELECT hc.produced_by_group AS vendor,
                       SUM(CASE WHEN q.passed = 1 THEN IFNULL(hc.video_seconds, 0) ELSE 0 END) / 3600.0 AS qc_pass_h,
                       SUM(IFNULL(hc.video_seconds, 0)) / 3600.0 AS qc_total_h
                FROM human_cases hc
                JOIN (
                    SELECT hcn.human_case_id, CASE WHEN hcn.node_status = 3 THEN 1 ELSE 0 END AS passed
                    FROM human_case_nodes hcn
                    WHERE hcn.node_name = 'human_case_sampling' AND hcn.node_status IN (3, 4)
                      AND DATE(hcn.node_updated_at) BETWEEN %s AND %s
                    UNION ALL
                    SELECT hcn.human_case_id, CASE WHEN hcn.node_status = 3 THEN 1 ELSE 0 END AS passed
                    FROM human_case_nodes hcn
                    WHERE hcn.node_name = 'human_case_inspect' AND hcn.node_status IN (3, 4)
                      AND DATE(hcn.node_updated_at) BETWEEN %s AND %s
                      AND hcn.human_case_id NOT IN (
                          SELECT DISTINCT human_case_id FROM human_case_nodes WHERE node_name = 'human_case_sampling'
                      )
                ) q ON q.human_case_id = hc.id
                WHERE hc.deleted_at IS NULL
                  AND hc.produced_by_group IS NOT NULL AND hc.produced_by_group != ''
                GROUP BY vendor
            """, [week_start, week_end, week_start, week_end])
            for r in cur.fetchall():
                v = r['vendor']
                if v not in result:
                    result[v] = {'vendor': v, 'collect_h': 0.0,
                                 'active_collectors': set(), 'qc_pass_h': 0.0, 'qc_total_h': 0.0}
                result[v]['qc_pass_h']  += float(r['qc_pass_h'] or 0)
                result[v]['qc_total_h'] += float(r['qc_total_h'] or 0)
        conn.close()
    except Exception:
        import traceback; traceback.print_exc()

    # ── 新库 ──────────────────────────────────────────────────────
    _w_start = week_start + " 00:00:00"
    _w_end   = week_end   + " 23:59:59"
    try:
        conn = mysql_new()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT
                    ag.name AS vendor,
                    SUM(IFNULL(CAST(JSON_EXTRACT(pt.value,'$.data_info.duration') AS DECIMAL(15,4)),0)) / 3600.0 AS collect_h,
                    GROUP_CONCAT(DISTINCT hc.producer) AS producers
                FROM human_case hc
                JOIN human_case_node hcn ON hcn.human_case_id = hc.id
                LEFT JOIN human_case_tag pt ON pt.human_case_id = hc.id AND pt.type = 'produce_tags'
                LEFT JOIN auth.users au ON au.name = hc.producer
                LEFT JOIN auth.user_group aug ON aug.user_uuid = au.uuid
                LEFT JOIN auth.`groups` ag ON ag.id = aug.group_uuid AND ag.biz_type = 'produce'
                WHERE hcn.node_name = 'human_case_upload'
                  AND hcn.node_status = 3
                  AND hcn.node_updated_at BETWEEN %s AND %s
                  AND ag.name IS NOT NULL AND ag.name != ''
                GROUP BY ag.name
            """, [_w_start, _w_end])
            for r in cur.fetchall():
                producers = set((r['producers'] or '').split(',')) if r['producers'] else set()
                _merge(r['vendor'], float(r['collect_h'] or 0), producers, 0.0, 0.0)

            cur.execute("""
                SELECT
                    ag.name AS vendor,
                    SUM(CASE WHEN q.passed = 1 THEN IFNULL(CAST(JSON_EXTRACT(pt.value,'$.data_info.duration') AS DECIMAL(15,4)),0) ELSE 0 END) / 3600.0 AS qc_pass_h,
                    SUM(IFNULL(CAST(JSON_EXTRACT(pt.value,'$.data_info.duration') AS DECIMAL(15,4)),0)) / 3600.0 AS qc_total_h
                FROM human_case hc
                LEFT JOIN human_case_tag pt ON pt.human_case_id = hc.id AND pt.type = 'produce_tags'
                LEFT JOIN auth.users au ON au.name = hc.producer
                LEFT JOIN auth.user_group aug ON aug.user_uuid = au.uuid
                LEFT JOIN auth.`groups` ag ON ag.id = aug.group_uuid AND ag.biz_type = 'produce'
                JOIN (
                    SELECT hcn.human_case_id, CASE WHEN hcn.node_status = 3 THEN 1 ELSE 0 END AS passed
                    FROM human_case_node hcn
                    WHERE hcn.node_name = 'human_case_sampling' AND hcn.node_status IN (3, 4)
                      AND hcn.node_updated_at BETWEEN %s AND %s
                    UNION ALL
                    SELECT hcn.human_case_id, CASE WHEN hcn.node_status = 3 THEN 1 ELSE 0 END AS passed
                    FROM human_case_node hcn
                    WHERE hcn.node_name = 'human_case_inspect' AND hcn.node_status IN (3, 4)
                      AND hcn.node_updated_at BETWEEN %s AND %s
                      AND hcn.human_case_id NOT IN (
                          SELECT DISTINCT human_case_id FROM human_case_node WHERE node_name = 'human_case_sampling'
                      )
                ) q ON q.human_case_id = hc.id
                WHERE ag.name IS NOT NULL AND ag.name != ''
                GROUP BY ag.name
            """, [_w_start, _w_end, _w_start, _w_end])
            for r in cur.fetchall():
                v = r['vendor']
                if v not in result:
                    result[v] = {'vendor': v, 'collect_h': 0.0,
                                 'active_collectors': set(), 'qc_pass_h': 0.0, 'qc_total_h': 0.0}
                result[v]['qc_pass_h']  += float(r['qc_pass_h'] or 0)
                result[v]['qc_total_h'] += float(r['qc_total_h'] or 0)
        conn.close()
    except Exception:
        import traceback; traceback.print_exc()

    # active_collectors set → count
    for v in result.values():
        v['active_collectors'] = len(v['active_collectors'])
    return list(result.values())


def _query_vendor_daily_eff(week_start, week_end):
    """按天查每个供应商的采集人效（h/人/天），过滤当天 ≤10条的采集员，返回日均值。
    老库 + 新库合并，相同 vendor+day 合并 producers 集合，避免跨库重复计人。
    Returns: {vendor: daily_collect_eff}
    """
    from collections import defaultdict
    vendor_day = defaultdict(lambda: defaultdict(lambda: {'h': 0.0, 'producers': set()}))

    # ── 老库 ──────────────────────────────────────────────────────
    try:
        conn = mysql()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT DATE(hcn.node_updated_at) AS day,
                       hc.produced_by_group AS vendor,
                       hc.producer AS producer,
                       COUNT(*) AS cases,
                       SUM(IFNULL(hc.video_seconds, 0)) / 3600.0 AS collect_h
                FROM human_cases hc
                JOIN human_case_nodes hcn ON hcn.human_case_id = hc.id
                WHERE hcn.node_name = 'human_case_upload'
                  AND hcn.node_status = 3
                  AND DATE(hcn.node_updated_at) BETWEEN %s AND %s
                  AND hc.deleted_at IS NULL
                  AND hc.produced_by_group IS NOT NULL AND hc.produced_by_group != ''
                GROUP BY day, vendor, producer
                HAVING cases > 10
            """, [week_start, week_end])
            for r in cur.fetchall():
                day = str(r['day'])
                v = r['vendor']
                vendor_day[v][day]['h'] += float(r['collect_h'] or 0)
                vendor_day[v][day]['producers'].add(r['producer'])
        conn.close()
    except Exception:
        import traceback; traceback.print_exc()

    # ── 新库 ──────────────────────────────────────────────────────
    _w_start2 = week_start + " 00:00:00"
    _w_end2   = week_end   + " 23:59:59"
    try:
        conn = mysql_new()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT DATE(hcn.node_updated_at) AS day,
                       ag.name AS vendor,
                       hc.producer AS producer,
                       COUNT(*) AS cases,
                       SUM(IFNULL(CAST(JSON_EXTRACT(pt.value,'$.data_info.duration') AS DECIMAL(15,4)),0)) / 3600.0 AS collect_h
                FROM human_case hc
                JOIN human_case_node hcn ON hcn.human_case_id = hc.id
                LEFT JOIN human_case_tag pt ON pt.human_case_id = hc.id AND pt.type = 'produce_tags'
                LEFT JOIN auth.users au ON au.name = hc.producer
                LEFT JOIN auth.user_group aug ON aug.user_uuid = au.uuid
                LEFT JOIN auth.`groups` ag ON ag.id = aug.group_uuid AND ag.biz_type = 'produce'
                WHERE hcn.node_name = 'human_case_upload'
                  AND hcn.node_status = 3
                  AND hcn.node_updated_at BETWEEN %s AND %s
                  AND ag.name IS NOT NULL AND ag.name != ''
                GROUP BY day, ag.name, hc.producer
                HAVING cases > 10
            """, [_w_start2, _w_end2])
            for r in cur.fetchall():
                day = str(r['day'])
                v = r['vendor']
                vendor_day[v][day]['h'] += float(r['collect_h'] or 0)
                vendor_day[v][day]['producers'].add(r['producer'])
        conn.close()
    except Exception:
        import traceback; traceback.print_exc()

    # 计算日均采集人效
    result = {}
    for vendor, days in vendor_day.items():
        daily_effs = []
        for day_data in days.values():
            n = len(day_data['producers'])
            if n > 0:
                daily_effs.append(day_data['h'] / n)
        if daily_effs:
            result[vendor] = round(sum(daily_effs) / len(daily_effs), 2)
    return result


# ─── LiteLLM ─────────────────────────────────────────────────
LITELLM_URL = os.environ.get("LITELLM_URL", "https://ai.lightwheel.net:8086/v1/chat/completions")
LITELLM_KEY = os.environ["LITELLM_KEY"]

KANBAN_PASSWORD = os.environ.get("KANBAN_PASSWORD", "")
ADMIN_PASSWORD  = os.environ.get("ADMIN_PASSWORD", "")
API_KEY         = os.environ.get("API_KEY", "")

# ── 供应商账号（从 vendors.json 动态加载）────────────────────────────────────
# 数据结构：{ "聚航": { "accounts": [ {"username":"聚航_1","password":"xxx","created_at":"..."} ] } }
VENDORS_FILE = os.path.join(os.path.dirname(__file__), "vendors.json")

def _load_vendors():
    """返回 {vendor_name: {"accounts": [...]}} 结构"""
    if os.path.exists(VENDORS_FILE):
        try:
            with open(VENDORS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            # 自动迁移旧格式 {name: {password, created_at}} → 新格式
            migrated = False
            for name, info in list(data.items()):
                if "accounts" not in info:
                    data[name] = {"accounts": [{"username": name + "_1", "password": info.get("password", ""), "created_at": info.get("created_at", "")}]}
                    migrated = True
            if migrated:
                _save_vendors(data)
            return data
        except Exception:
            pass
    # 兜底：从环境变量读
    data = {}
    for k, v in os.environ.items():
        if k.startswith("VENDOR_") and v:
            name = k[7:]
            data[name] = {"accounts": [{"username": name + "_1", "password": v, "created_at": ""}]}
    return data

def _save_vendors(data):
    with open(VENDORS_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def _get_account_map():
    """返回 {username: {password, group}} 供登录校验"""
    result = {}
    for group, info in _load_vendors().items():
        for acc in info.get("accounts", []):
            result[acc["username"]] = {"password": acc["password"], "group": group}
    return result

VENDOR_PASSWORDS = {}  # 保留兼容性，不再使用

app = Flask(__name__, static_folder=".", static_url_path="")
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "lw-kanban-secret")
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"

def _warmup_cumul_cache():
    """启动时在后台预热所有供应商的累计缓存，避免首次请求超时"""
    def _run():
        try:
            vendors = list(_load_vendors().keys())
        except Exception:
            return
        for group in vendors:
            key = f"vendor_cumul:{group}"
            if _cache_get(key) is not None or _disk_cache_get(key) is not None:
                continue  # 已有缓存，跳过
            try:
                n_tc, n_tq, n_pi, n_ps, n_fc = _query_new_db_cumul(group)
                # 同时查老库（warmup 线程无 gunicorn timeout，可以慢查）
                conn = mysql()
                try:
                    with conn.cursor() as cur:
                        cur.execute("""
                            SELECT hc.producer,
                                SUM(IFNULL(hc.video_seconds, 0)) / 3600.0 AS total_collect_h,
                                MIN(DATE(hcn.node_updated_at)) AS first_collect_date
                            FROM human_cases hc
                            JOIN human_case_nodes hcn ON hcn.human_case_id = hc.id
                            WHERE hcn.node_name = 'human_case_produce_complete'
                              AND hcn.node_status = 3
                              AND hc.deleted_at IS NULL
                              AND hc.produced_by_group = %s
                            GROUP BY hc.producer
                        """, [group])
                        _vrows = cur.fetchall()
                        old_tc = {r["producer"]: float(r["total_collect_h"] or 0) for r in _vrows}
                        old_fc = {r["producer"]: str(r["first_collect_date"]) if r["first_collect_date"] else "" for r in _vrows}

                        cur.execute("""
                            SELECT hc.producer,
                                COUNT(*) AS total_qc_total,
                                SUM(CASE WHEN q.passed = 1 THEN 1 ELSE 0 END) AS total_qc_passed,
                                SUM(CASE WHEN q.passed = 1 THEN IFNULL(hc.video_seconds, 0) ELSE 0 END) / 3600.0 AS total_qc_h,
                                SUM(IFNULL(hc.video_seconds, 0)) / 3600.0 AS total_qc_total_h
                            FROM human_cases hc
                            JOIN (
                                SELECT hcn.human_case_id, CASE WHEN hcn.node_status = 3 THEN 1 ELSE 0 END AS passed
                                FROM human_case_nodes hcn
                                WHERE hcn.node_name = 'human_case_sampling' AND hcn.node_status IN (3, 4)
                                UNION ALL
                                SELECT hcn.human_case_id, CASE WHEN hcn.node_status = 3 THEN 1 ELSE 0 END AS passed
                                FROM human_case_nodes hcn
                                WHERE hcn.node_name = 'human_case_inspect' AND hcn.node_status IN (3, 4)
                                  AND NOT EXISTS (
                                      SELECT 1 FROM human_case_nodes s
                                      WHERE s.human_case_id = hcn.human_case_id AND s.node_name = 'human_case_sampling'
                                  )
                            ) q ON q.human_case_id = hc.id
                            WHERE hc.deleted_at IS NULL AND hc.produced_by_group = %s
                            GROUP BY hc.producer
                        """, [group])
                        old_tq = {r["producer"]: {
                            "total_qc_h": float(r["total_qc_h"] or 0),
                            "total_qc_passed": int(r["total_qc_passed"] or 0),
                            "total_qc_total": int(r["total_qc_total"] or 0),
                            "total_qc_total_h": float(r["total_qc_total_h"] or 0),
                        } for r in cur.fetchall()}

                        cur.execute("""
                            SELECT hc.producer,
                                SUM(IFNULL(hc.video_seconds, 0)) / 3600.0 AS pending_inspect_h
                            FROM human_cases hc
                            JOIN human_case_nodes hcn ON hcn.human_case_id = hc.id
                            WHERE hcn.node_name = 'human_case_inspect'
                              AND hcn.node_status IN (1, 2)
                              AND hc.deleted_at IS NULL AND hc.produced_by_group = %s
                            GROUP BY hc.producer
                        """, [group])
                        old_pi = {r["producer"]: float(r["pending_inspect_h"] or 0) for r in cur.fetchall()}

                        cur.execute("""
                            SELECT hc.producer,
                                SUM(IFNULL(hc.video_seconds, 0)) / 3600.0 AS pending_sampling_h
                            FROM human_cases hc
                            JOIN human_case_nodes hcn ON hcn.human_case_id = hc.id
                            WHERE hcn.node_name = 'human_case_sampling'
                              AND hcn.node_status IN (1, 2)
                              AND hc.deleted_at IS NULL AND hc.produced_by_group = %s
                            GROUP BY hc.producer
                        """, [group])
                        old_ps = {r["producer"]: float(r["pending_sampling_h"] or 0) for r in cur.fetchall()}
                finally:
                    conn.close()

                # 合并老库 + 新库，写入完整缓存
                merged_tc = _merge_float_map(old_tc, n_tc)
                merged_tq = _merge_qc_map(old_tq, n_tq)
                merged_pi = _merge_float_map(old_pi, n_pi)
                merged_ps = _merge_float_map(old_ps, n_ps)
                merged_fc = _merge_date_map(old_fc, n_fc)
                _cache_set(key, (merged_tc, merged_tq, merged_pi, merged_ps, merged_fc), 21600)
                _disk_cache_set(key, (merged_tc, merged_tq, merged_pi, merged_ps, merged_fc), 21600)
            except Exception:
                pass

    t = threading.Thread(target=_run, daemon=True)
    t.start()


_LOGIN_HTML = """<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>LW 运营看板</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,"PingFang SC","Microsoft YaHei",sans-serif;
     background:#f0f2f5;display:flex;align-items:center;justify-content:center;min-height:100vh}
.card{background:#fff;border-radius:16px;padding:40px 36px;width:360px;
      box-shadow:0 4px 24px rgba(0,0,0,.10)}
h1{font-size:20px;font-weight:700;color:#1a1a2e;margin-bottom:6px}
p{font-size:13px;color:#999;margin-bottom:28px}
label{font-size:13px;color:#555;display:block;margin-bottom:6px}
input[type=password]{width:100%;padding:10px 14px;border:1.5px solid #e2e8f0;border-radius:8px;
  font-size:14px;outline:none;transition:border .2s}
input[type=password]:focus{border-color:#6366f1}
button{width:100%;margin-top:18px;padding:11px;background:#6366f1;color:#fff;border:none;
  border-radius:8px;font-size:14px;font-weight:600;cursor:pointer;transition:background .2s}
button:hover{background:#4f46e5}
.err{margin-top:14px;font-size:13px;color:#ef4444;text-align:center}
</style>
</head>
<body>
<div class="card">
  <h1>🌐 LW 运营看板</h1>
  <p>请输入访问密码</p>
  <form method="POST" action="/login">
    <label>密码</label>
    <input type="password" name="password" autofocus autocomplete="current-password">
    <button type="submit">进入</button>
  </form>
  <div class="err">__ERROR__</div>
</div>
</body>
</html>"""


@app.before_request
def _check_auth():
    # 供应商路径走独立鉴权
    if request.path.startswith("/vendor") or request.path.startswith("/api/vendor"):
        return None
    if request.path in ("/login", "/logout"):
        return None
    if request.path.startswith("/dm") or request.path.startswith("/api/dm"):
        return None
    if request.path.startswith("/admin") or request.path.startswith("/api/admin"):
        return None
    # cron token 豁免
    _cron_token = os.environ.get("CRON_TOKEN", "")
    if _cron_token and request.args.get("token") == _cron_token:
        return None
    if not KANBAN_PASSWORD:          # 未设置密码则不拦截
        return None
    if not session.get("authed"):
        return redirect("/login")


@app.route("/login", methods=["GET", "POST"])
def login():
    error = ""
    if request.method == "POST":
        pwd = request.form.get("password", "")
        if ADMIN_PASSWORD and pwd == ADMIN_PASSWORD:
            session["authed"] = True
            session["is_admin"] = True
            return redirect("/")
        elif pwd == KANBAN_PASSWORD:
            session["authed"] = True
            session.pop("is_admin", None)
            return redirect("/")
        error = "密码错误，请重试"
    return _LOGIN_HTML.replace("__ERROR__", error)


@app.route("/api/auth/me")
def api_auth_me():
    return jsonify({"admin": bool(session.get("is_admin"))})


@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")


# ═══════════════════════════════════════════════════════════════
#  供应商登录 & 绩效页
# ═══════════════════════════════════════════════════════════════
_VENDOR_LOGIN_HTML = """<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>供应商看板</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,"PingFang SC","Microsoft YaHei",sans-serif;
     background:#0f172a;display:flex;align-items:center;justify-content:center;min-height:100vh}
.card{background:#1e293b;border-radius:16px;padding:40px 36px;width:380px;
      box-shadow:0 4px 24px rgba(0,0,0,.4)}
h1{font-size:20px;font-weight:700;color:#e2e8f0;margin-bottom:6px}
p{font-size:13px;color:#64748b;margin-bottom:28px}
label{font-size:13px;color:#94a3b8;display:block;margin-bottom:6px}
input{width:100%;padding:10px 14px;border:1.5px solid #334155;border-radius:8px;
  background:#0f172a;color:#e2e8f0;font-size:14px;outline:none;transition:border .2s;margin-bottom:16px}
input:focus{border-color:#3b82f6}
button{width:100%;margin-top:4px;padding:11px;background:#3b82f6;color:#fff;border:none;
  border-radius:8px;font-size:14px;font-weight:600;cursor:pointer;transition:background .2s}
button:hover{background:#2563eb}
.err{margin-top:14px;font-size:13px;color:#f87171;text-align:center}
</style>
</head>
<body>
<div class="card">
  <h1>供应商看板</h1>
  <p>请输入账号和密码</p>
  <form method="POST" action="/vendor/login">
    <label>账号</label>
    <input type="text" name="username" placeholder="如：聚航_1" autofocus autocomplete="username" value="__USERNAME__">
    <label>密码</label>
    <input type="password" name="password" autocomplete="current-password">
    <button type="submit">进入</button>
  </form>
  <div class="err">__ERROR__</div>
</div>
</body>
</html>"""


@app.route("/vendor")
@app.route("/vendor/")
def vendor_index():
    if session.get("vendor_group"):
        return redirect("/vendor/performance")
    return _VENDOR_LOGIN_HTML.replace("__ERROR__", "").replace("__USERNAME__", "")


@app.route("/vendor/login", methods=["POST"])
def vendor_login():
    username = request.form.get("username", "").strip()
    password = request.form.get("password", "")
    account_map = _get_account_map()
    acc = account_map.get(username)
    if acc and password == acc["password"]:
        session["vendor_group"] = acc["group"]
        return redirect("/vendor/performance")
    error = "账号或密码错误"
    return _VENDOR_LOGIN_HTML.replace("__ERROR__", error).replace("__USERNAME__", username)


@app.route("/vendor/logout")
def vendor_logout():
    session.pop("vendor_group", None)
    return redirect("/vendor")


# ── 管理员账号管理 ────────────────────────────────────────────────────────────

_ADMIN_LOGIN_HTML = """<!DOCTYPE html>
<html lang="zh-CN"><head><meta charset="UTF-8"><title>管理员登录</title>
<style>*{margin:0;padding:0;box-sizing:border-box}
body{background:#0b0b1e;color:#e2e8f0;font-family:-apple-system,sans-serif;
display:flex;align-items:center;justify-content:center;min-height:100vh}
.card{background:#13132d;border:1px solid rgba(99,102,241,.2);border-radius:12px;
padding:40px;width:360px}
h2{color:#c7d2fe;margin-bottom:24px;font-size:18px}
label{font-size:13px;color:#94a3b8;display:block;margin-bottom:6px}
input{width:100%;height:38px;padding:0 12px;border-radius:6px;
border:1px solid rgba(99,102,241,.35);background:#0b0b1e;color:#e2e8f0;
font-size:14px;outline:none;margin-bottom:16px}
button{width:100%;height:38px;border-radius:6px;border:none;cursor:pointer;
background:linear-gradient(135deg,#4f46e5,#38bdf8);color:#fff;font-size:14px}
.err{color:#f87171;font-size:13px;margin-bottom:12px}</style></head>
<body><div class="card"><h2>账号管理 · 管理员登录</h2>
<form method="POST"><div class="err">__ERROR__</div>
<label>管理员密码</label>
<input type="password" name="password" autofocus>
<button type="submit">登录</button></form></div></body></html>"""


@app.route("/admin/accounts")
def admin_accounts_page():
    if not session.get("admin_authed"):
        return redirect("/admin/login")
    return send_from_directory(".", "accounts.html")


@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        if request.form.get("password") == ADMIN_PASSWORD:
            session["admin_authed"] = True
            return redirect("/admin/accounts")
        return _ADMIN_LOGIN_HTML.replace("__ERROR__", "密码错误")
    if session.get("admin_authed"):
        return redirect("/admin/accounts")
    return _ADMIN_LOGIN_HTML.replace("__ERROR__", "")


@app.route("/admin/logout")
def admin_logout():
    session.pop("admin_authed", None)
    return redirect("/admin/login")


@app.route("/api/admin/vendors", methods=["GET"])
def api_admin_vendors_list():
    if not session.get("admin_authed"):
        return jsonify({"error": "unauthorized"}), 401
    vendors = _load_vendors()
    # 返回分组结构: [{group, accounts: [{username, password, created_at}]}]
    result = []
    for group, info in vendors.items():
        result.append({"group": group, "accounts": info.get("accounts", [])})
    return jsonify({"vendors": result})


@app.route("/api/admin/vendors/batch", methods=["POST"])
def api_admin_vendors_batch():
    """批量为某供应商创建账号
    body: {group: "聚航", count: 3}  → 自动编号
    """
    if not session.get("admin_authed"):
        return jsonify({"error": "unauthorized"}), 401
    body = request.get_json() or {}
    group = (body.get("group") or "").strip()
    count = int(body.get("count") or 0)
    if not group or count <= 0:
        return jsonify({"error": "invalid params"}), 400
    vendors = _load_vendors()
    if group not in vendors:
        vendors[group] = {"accounts": []}
    existing = vendors[group]["accounts"]
    # 找当前最大编号
    max_idx = 0
    for acc in existing:
        uname = acc.get("username", "")
        if uname.startswith(group + "_"):
            try:
                idx = int(uname[len(group) + 1:])
                max_idx = max(max_idx, idx)
            except ValueError:
                pass
    today = datetime.now().strftime("%Y-%m-%d")
    created = []
    for i in range(1, count + 1):
        username = f"{group}_{max_idx + i}"
        pwd = _gen_password()
        existing.append({"username": username, "password": pwd, "created_at": today})
        created.append({"username": username, "password": pwd})
    _save_vendors(vendors)
    return jsonify({"created": created})


@app.route("/api/admin/vendors/<group>/<username>", methods=["DELETE"])
def api_admin_vendor_delete(group, username):
    if not session.get("admin_authed"):
        return jsonify({"error": "unauthorized"}), 401
    vendors = _load_vendors()
    if group not in vendors:
        return jsonify({"error": "not found"}), 404
    accounts = vendors[group]["accounts"]
    vendors[group]["accounts"] = [a for a in accounts if a["username"] != username]
    # 如果组下无账号则删除整个组
    if not vendors[group]["accounts"]:
        del vendors[group]
    _save_vendors(vendors)
    return jsonify({"ok": True})


@app.route("/api/admin/vendors/<group>/<username>/reset", methods=["POST"])
def api_admin_vendor_reset(group, username):
    if not session.get("admin_authed"):
        return jsonify({"error": "unauthorized"}), 401
    vendors = _load_vendors()
    if group not in vendors:
        return jsonify({"error": "not found"}), 404
    body = request.get_json() or {}
    pwd = (body.get("password") or "").strip()
    if not pwd:
        return jsonify({"error": "empty password"}), 400
    for acc in vendors[group]["accounts"]:
        if acc["username"] == username:
            acc["password"] = pwd
            _save_vendors(vendors)
            return jsonify({"ok": True})
    return jsonify({"error": "not found"}), 404


def _gen_password():
    import random
    chars = "ABCDEFGHJKLMNPQRSTUVWXYZabcdefghjkmnpqrstuvwxyz23456789"
    return "".join(random.choice(chars) for _ in range(12))


@app.route("/vendor/performance")
def vendor_performance():
    if not session.get("vendor_group"):
        return redirect("/vendor")
    return send_from_directory(".", "vendor_performance.html")


@app.route("/api/vendor/collectors")
def vendor_collectors():
    """供应商专属采集详情（按天，只返回自己组）"""
    group = session.get("vendor_group")
    if not group:
        return jsonify({"error": "unauthorized"}), 401

    date_str = request.args.get("date", str(date.today()))

    # 当日数据缓存 5 分钟，累计/待质检数据缓存 6 小时（内存+磁盘）
    daily_cache_key = f"vendor_daily_v2:{group}:{date_str}"
    cumul_cache_key = f"vendor_cumul:{group}"
    cached_daily = _cache_get(daily_cache_key)
    cached_cumul = _cache_get(cumul_cache_key)
    # 内存未命中时从磁盘恢复（重启后仍有效）
    if cached_cumul is None:
        cached_cumul = _disk_cache_get(cumul_cache_key)
        if cached_cumul is not None:
            _cache_set(cumul_cache_key, cached_cumul, 21600)

    conn = mysql()
    try:
        with conn.cursor() as cur:
            if cached_daily:
                if len(cached_daily) == 5:
                    produce_rows, qc_map, collect_qc_map, collect_inspected_map, upload_h_map = cached_daily
                elif len(cached_daily) == 4:
                    produce_rows, qc_map, collect_qc_map, collect_inspected_map = cached_daily
                    upload_h_map = {}
                elif len(cached_daily) == 3:
                    produce_rows, qc_map, collect_qc_map = cached_daily
                    collect_inspected_map = {}
                    upload_h_map = {}
                else:
                    produce_rows, qc_map = cached_daily
                    collect_qc_map = {}
                    collect_inspected_map = {}
                    upload_h_map = {}
            
            cur.execute("""
                SELECT
                    hc.id AS case_id,
                    'old' AS db_source,
                    hc.producer,
                    hc.produced_by_group AS vendor,
                    MIN(hcn.node_created_at) AS t_start,
                    MAX(hcn.node_updated_at) AS t_end,
                    MAX(IFNULL(hc.video_seconds, 0)) AS vsec
                FROM human_cases hc
                JOIN human_case_nodes hcn ON hcn.human_case_id = hc.id
                WHERE hcn.node_name = 'human_case_produce_complete'
                  AND hcn.node_status = 3
                  AND DATE(hcn.node_updated_at) = %s
                  AND hc.deleted_at IS NULL
                  AND hc.produced_by_group = %s
                GROUP BY hc.id, hc.producer, hc.produced_by_group
                ORDER BY hc.producer, MIN(hcn.node_created_at)
            """, [date_str, group])
            produce_rows = cur.fetchall()

            if produce_rows:
                      # 老库有数据时才跑 QC 查询（避免对 human_case_nodes 全表扫描）
                      cur.execute("""
                          SELECT hc.producer,
                              COUNT(*) AS qc_total_cases,
                              SUM(CASE WHEN q.passed = 1 THEN 1 ELSE 0 END) AS qc_passed_cases,
                              SUM(CASE WHEN q.passed = 1 THEN IFNULL(hc.video_seconds, 0) ELSE 0 END) / 3600.0 AS qc_h,
                              SUM(IFNULL(hc.video_seconds, 0)) / 3600.0 AS qc_total_h
                          FROM human_cases hc
                          JOIN (
                              SELECT hcn.human_case_id, CASE WHEN hcn.node_status = 3 THEN 1 ELSE 0 END AS passed
                              FROM human_case_nodes hcn
                              WHERE hcn.node_name = 'human_case_sampling' AND hcn.node_status IN (3, 4)
                                AND DATE(hcn.node_updated_at) = %s
                              UNION ALL
                              SELECT hcn.human_case_id, CASE WHEN hcn.node_status = 3 THEN 1 ELSE 0 END AS passed
                              FROM human_case_nodes hcn
                              WHERE hcn.node_name = 'human_case_inspect' AND hcn.node_status IN (3, 4)
                                AND DATE(hcn.node_updated_at) = %s
                                AND NOT EXISTS (
                                    SELECT 1 FROM human_case_nodes s
                                    WHERE s.human_case_id = hcn.human_case_id AND s.node_name = 'human_case_sampling'
                                )
                          ) q ON q.human_case_id = hc.id
                          WHERE hc.deleted_at IS NULL AND hc.produced_by_group = %s
                          GROUP BY hc.producer
                      """, [date_str, date_str, group])
                      qc_map = {r["producer"]: {"qc_cases": int(r["qc_passed_cases"] or 0),
                                                 "qc_total": int(r["qc_total_cases"] or 0),
                                                 "qc_h": float(r["qc_h"] or 0),
                                                 "qc_total_h": float(r["qc_total_h"] or 0)} for r in cur.fetchall()}

                      cur.execute("""
                          SELECT hc.producer,
                              SUM(CASE
                                  WHEN EXISTS (SELECT 1 FROM human_case_nodes WHERE human_case_id = hc.id AND node_name='human_case_sampling' AND node_status=3)
                                      THEN IFNULL(hc.video_seconds, 0)
                                  WHEN EXISTS (SELECT 1 FROM human_case_nodes WHERE human_case_id = hc.id AND node_name='human_case_sampling')
                                      THEN 0
                                  WHEN EXISTS (SELECT 1 FROM human_case_nodes WHERE human_case_id = hc.id AND node_name='human_case_inspect' AND node_status=3)
                                      THEN IFNULL(hc.video_seconds, 0)
                                  ELSE 0
                              END) / 3600.0 AS collect_qc_h,
                              SUM(CASE
                                  WHEN EXISTS (SELECT 1 FROM human_case_nodes WHERE human_case_id = hc.id AND node_name='human_case_sampling' AND node_status IN (3,4))
                                      THEN IFNULL(hc.video_seconds, 0)
                                  WHEN EXISTS (SELECT 1 FROM human_case_nodes WHERE human_case_id = hc.id AND node_name='human_case_sampling')
                                      THEN 0
                                  WHEN EXISTS (SELECT 1 FROM human_case_nodes WHERE human_case_id = hc.id AND node_name='human_case_inspect' AND node_status IN (3,4))
                                      THEN IFNULL(hc.video_seconds, 0)
                                  ELSE 0
                              END) / 3600.0 AS collect_inspected_h
                          FROM human_cases hc
                          JOIN human_case_nodes hcn ON hcn.human_case_id = hc.id
                          WHERE hcn.node_name = 'human_case_produce_complete'
                            AND hcn.node_status = 3
                            AND DATE(hcn.node_updated_at) = %s
                            AND hc.deleted_at IS NULL
                            AND hc.produced_by_group = %s
                          GROUP BY hc.producer
                      """, [date_str, group])
                      _cqc_rows = cur.fetchall()
                      collect_qc_map = {r["producer"]: float(r["collect_qc_h"] or 0) for r in _cqc_rows}
                      collect_inspected_map = {r["producer"]: float(r["collect_inspected_h"] or 0) for r in _cqc_rows}
                      upload_h_map = {}
            else:
                      qc_map = {}
                      collect_qc_map = {}
                      collect_inspected_map = {}
                      upload_h_map = {}

                  # 合并新库当日数据
            new_produce, new_qc, new_collect_qc, new_collect_inspected, new_upload_h, _wrist_uuids, _nwu, _npu = _query_new_db_daily(date_str, group=group)
            produce_rows = _dedupe_produce_rows(list(produce_rows) + list(new_produce))
            qc_map = _merge_qc_map(qc_map, new_qc)
            collect_qc_map = _merge_float_map(collect_qc_map, new_collect_qc)
            collect_inspected_map = _merge_float_map(collect_inspected_map, new_collect_inspected)
            upload_h_map = _merge_float_map(upload_h_map, new_upload_h)
            _cache_set(daily_cache_key, (produce_rows, qc_map, collect_qc_map, collect_inspected_map, upload_h_map), 300)
            
            if cached_cumul:
                if len(cached_cumul) == 5:
                    total_collect_map, total_qc_map, pending_inspect_map, pending_sampling_map, first_collect_map_v = cached_cumul
                else:
                    total_collect_map, total_qc_map, pending_inspect_map, first_collect_map_v = cached_cumul if len(cached_cumul) == 4 else (*cached_cumul, {})
                    pending_sampling_map = {}
            else:
                cur.execute("""
                    SELECT hc.producer,
                        SUM(IFNULL(hc.video_seconds, 0)) / 3600.0 AS total_collect_h,
                        MIN(DATE(hcn.node_updated_at)) AS first_collect_date
                    FROM human_cases hc
                    JOIN human_case_nodes hcn ON hcn.human_case_id = hc.id
                    WHERE hcn.node_name = 'human_case_produce_complete'
                      AND hcn.node_status = 3
                      AND hc.deleted_at IS NULL
                      AND hc.produced_by_group = %s
                    GROUP BY hc.producer
                """, [group])
                _vrows = cur.fetchall()
                total_collect_map = {r["producer"]: float(r["total_collect_h"] or 0) for r in _vrows}
                first_collect_map_v = {r["producer"]: str(r["first_collect_date"]) if r["first_collect_date"] else "" for r in _vrows}

                cur.execute("""
                    SELECT hc.producer,
                        COUNT(*) AS total_qc_total,
                        SUM(CASE WHEN q.passed = 1 THEN 1 ELSE 0 END) AS total_qc_passed,
                        SUM(CASE WHEN q.passed = 1 THEN IFNULL(hc.video_seconds, 0) ELSE 0 END) / 3600.0 AS total_qc_h,
                        SUM(IFNULL(hc.video_seconds, 0)) / 3600.0 AS total_qc_total_h
                    FROM human_cases hc
                    JOIN (
                        SELECT hcn.human_case_id, CASE WHEN hcn.node_status = 3 THEN 1 ELSE 0 END AS passed
                        FROM human_case_nodes hcn
                        WHERE hcn.node_name = 'human_case_sampling' AND hcn.node_status IN (3, 4)
                        UNION ALL
                        SELECT hcn.human_case_id, CASE WHEN hcn.node_status = 3 THEN 1 ELSE 0 END AS passed
                        FROM human_case_nodes hcn
                        WHERE hcn.node_name = 'human_case_inspect' AND hcn.node_status IN (3, 4)
                          AND NOT EXISTS (
                              SELECT 1 FROM human_case_nodes s
                              WHERE s.human_case_id = hcn.human_case_id AND s.node_name = 'human_case_sampling'
                          )
                    ) q ON q.human_case_id = hc.id
                    WHERE hc.deleted_at IS NULL AND hc.produced_by_group = %s
                    GROUP BY hc.producer
                """, [group])
                total_qc_map = {r["producer"]: {
                    "total_qc_h": float(r["total_qc_h"] or 0),
                    "total_qc_passed": int(r["total_qc_passed"] or 0),
                    "total_qc_total": int(r["total_qc_total"] or 0),
                    "total_qc_total_h": float(r["total_qc_total_h"] or 0),
                } for r in cur.fetchall()}

                # 待质检：human_case_inspect status IN (1,2)，不管有无 sampling
                cur.execute("""
                    SELECT hc.producer,
                        SUM(IFNULL(hc.video_seconds, 0)) / 3600.0 AS pending_inspect_h
                    FROM human_cases hc
                    JOIN human_case_nodes hcn ON hcn.human_case_id = hc.id
                    WHERE hcn.node_name = 'human_case_inspect'
                      AND hcn.node_status IN (1, 2)
                      AND hc.deleted_at IS NULL AND hc.produced_by_group = %s
                    GROUP BY hc.producer
                """, [group])
                pending_inspect_map = {r["producer"]: float(r["pending_inspect_h"] or 0) for r in cur.fetchall()}

                # 待抽检：human_case_sampling status IN (1,2)
                cur.execute("""
                    SELECT hc.producer,
                        SUM(IFNULL(hc.video_seconds, 0)) / 3600.0 AS pending_sampling_h
                    FROM human_cases hc
                    JOIN human_case_nodes hcn ON hcn.human_case_id = hc.id
                    WHERE hcn.node_name = 'human_case_sampling'
                      AND hcn.node_status IN (1, 2)
                      AND hc.deleted_at IS NULL AND hc.produced_by_group = %s
                    GROUP BY hc.producer
                """, [group])
                pending_sampling_map = {r["producer"]: float(r["pending_sampling_h"] or 0) for r in cur.fetchall()}

                # 合并新库累计数据
                n_tc, n_tq, n_pi, n_ps, n_fc = _query_new_db_cumul(group)
                total_collect_map   = _merge_float_map(total_collect_map, n_tc)
                total_qc_map        = _merge_qc_map(total_qc_map, n_tq)
                pending_inspect_map = _merge_float_map(pending_inspect_map, n_pi)
                pending_sampling_map= _merge_float_map(pending_sampling_map, n_ps)
                first_collect_map_v = _merge_date_map(first_collect_map_v, n_fc)

                _cache_set(cumul_cache_key, (total_collect_map, total_qc_map, pending_inspect_map, pending_sampling_map, first_collect_map_v), 21600)
                _disk_cache_set(cumul_cache_key, (total_collect_map, total_qc_map, pending_inspect_map, pending_sampling_map, first_collect_map_v), 21600)

    finally:
        conn.close()

    produce_rows = dedupe_case_rows(produce_rows)

    agg = aggregate_collect_rows(produce_rows)
    sessions_by_p = agg["sessions_by_p"]
    vsec_by_p = agg["vsec_by_p"]
    cases_by_p = agg["cases_by_p"]

    online_info = calc_online_hours(sessions_by_p)
    online_h_map = online_info["online_h_map"]
    first_seen_map = online_info["first_seen_map"]
    last_seen_map = online_info["last_seen_map"]

    persons = []
    for p, segs in sessions_by_p.items():
          qc = qc_map.get(p, {})
          qc_pass = int(qc.get("qc_cases") or 0)
          qc_tot = int(qc.get("qc_total") or 0)
          qc_h = float(qc.get("qc_h") or 0)
          qc_total_h = float(qc.get("qc_total_h") or 0)
          tqc = total_qc_map.get(p, {})
          tqc_h = float(tqc.get("total_qc_h") or 0)
          tqc_total_h = float(tqc.get("total_qc_total_h") or 0)
          persons.append({
              "producer":           p,
              "first_collect_date": first_collect_map_v.get(p, ""),
              # 累计
              "total_collect_h": round(total_collect_map.get(p, 0), 1),
              "total_qc_h":      round(tqc_h, 1),
              "total_qc_rate":   round(tqc_h / tqc_total_h * 100, 1) if tqc_total_h > 0 else 0,
              # 当日
              "collect_cases":   cases_by_p[p],
              "collect_h":       round(vsec_by_p[p] / 3600, 2),
              "qc_cases":        qc_pass,
              "qc_total":        qc_tot,
              "qc_h":            round(qc_h, 2),
              "qc_total_h":      round(qc_total_h, 2),
              "qc_rate":         round(qc_h / qc_total_h * 100, 1) if qc_total_h > 0 else 0,
              "collect_qc_h":        round(collect_qc_map.get(p, 0), 2),
              "collect_inspected_h": round(collect_inspected_map.get(p, 0), 2),
              "upload_h":            round(upload_h_map.get(p, 0), 2),
              # 待质检 / 待抽检
              "pending_inspect_h":  round(pending_inspect_map.get(p, 0), 1),
              "pending_sampling_h": round(pending_sampling_map.get(p, 0), 1),
              # 在线
              "online_h":        online_h_map.get(p, 0),
              "first_seen":      first_seen_map.get(p, ""),
              "last_seen":       last_seen_map.get(p, ""),
          })
    persons.sort(key=lambda x: -x["collect_h"])

    tc = sum(p["collect_h"] for p in persons)
    tq = sum(p["qc_h"] for p in persons)
    tq_total_h = sum(p["qc_total_h"] for p in persons)
    return jsonify({
        "date": date_str,
        "group": group,
        "total_persons": len(persons),
        "total_collect_h": round(tc, 2),
        "total_qc_h": round(tq, 2),
        "overall_qc_rate": round(tq / tq_total_h * 100, 1) if tq_total_h > 0 else 0,
        "persons": persons,
    })


@app.route("/api/vendor/sparklines")
def vendor_sparklines():
    """供应商采集员近 N 天每日质检通过时长（折线图数据）"""
    group = session.get("vendor_group")
    if not group:
        return jsonify({"error": "unauthorized"}), 401

    end_date_str = request.args.get("date", str(date.today()))
    days = int(request.args.get("days", 14))

    cache_key = f"vendor_sparklines:{group}:{end_date_str}:{days}"
    cached = _cache_get(cache_key)
    if cached:
        return jsonify(cached)

    from datetime import timedelta
    end_date   = date.fromisoformat(end_date_str)
    start_date = end_date - timedelta(days=days - 1)
    date_list  = [(start_date + timedelta(days=i)).isoformat() for i in range(days)]

    conn = mysql()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT hc.producer,
                    DATE(q.node_updated_at) AS day,
                    SUM(CASE WHEN q.passed = 1 THEN IFNULL(hc.video_seconds, 0) ELSE 0 END) / 3600.0 AS qc_h
                FROM human_cases hc
                JOIN (
                    SELECT hcn.human_case_id,
                           CASE WHEN hcn.node_status = 3 THEN 1 ELSE 0 END AS passed,
                           hcn.node_updated_at
                    FROM human_case_nodes hcn
                    WHERE hcn.node_name = 'human_case_sampling'
                      AND hcn.node_status IN (3, 4)
                      AND DATE(hcn.node_updated_at) BETWEEN %s AND %s
                    UNION ALL
                    SELECT hcn.human_case_id,
                           CASE WHEN hcn.node_status = 3 THEN 1 ELSE 0 END AS passed,
                           hcn.node_updated_at
                    FROM human_case_nodes hcn
                    WHERE hcn.node_name = 'human_case_inspect'
                      AND hcn.node_status IN (3, 4)
                      AND DATE(hcn.node_updated_at) BETWEEN %s AND %s
                      AND hcn.human_case_id NOT IN (
                          SELECT DISTINCT human_case_id FROM human_case_nodes
                          WHERE node_name = 'human_case_sampling'
                      )
                ) q ON q.human_case_id = hc.id
                WHERE hc.deleted_at IS NULL AND hc.produced_by_group = %s
                GROUP BY hc.producer, day
                ORDER BY hc.producer, day
            """, [start_date, end_date_str, start_date, end_date_str, group])
            rows = cur.fetchall()
    finally:
        conn.close()

    # 按人整理，补全缺失日期为 0
    from collections import defaultdict
    day_map = defaultdict(dict)  # producer -> {day: qc_h}
    for r in rows:
        day_map[r["producer"]][str(r["day"])] = round(float(r["qc_h"] or 0), 2)

    result = {
        "dates": date_list,
        "producers": {
            producer: [day_map[producer].get(d, 0) for d in date_list]
            for producer in day_map
        }
    }
    _cache_set(cache_key, result, 3600)
    return jsonify(result)


@app.route("/api/vendor/collector-stats")
def vendor_collector_stats():
    """供应商专属：只返回自己组的数据"""
    group = session.get("vendor_group")
    if not group:
        return jsonify({"error": "unauthorized"}), 401

    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")

    # 复用主接口逻辑，强制 group 为 session 中的值
    from flask import current_app
    with current_app.test_request_context(
        f"/api/collector-stats?start_date={start_date}&end_date={end_date}&group={group}"
    ):
        # 直接调用底层查询函数
        pass

    # 直接内联查询（避免 request context 问题）
    cache_key = f"vendor_stats:{group}:{start_date}:{end_date}"
    cached = _cache_get(cache_key)
    if cached:
        return jsonify(cached)

    conn = mysql()
    try:
        with conn.cursor() as cur:
            sql_collect = """
                SELECT hc.producer, hc.produced_by_group AS grp,
                       COUNT(DISTINCT hc.id) AS total_cases
                FROM human_cases hc
                JOIN human_case_nodes hcn ON hcn.human_case_id = hc.id
                WHERE hcn.node_name = 'human_case_produce_complete'
                  AND hcn.node_status = 3
                  AND DATE(hcn.node_updated_at) BETWEEN %s AND %s
                  AND hc.deleted_at IS NULL
                  AND hc.produced_by_group = %s
                GROUP BY hc.producer, hc.produced_by_group
            """
            cur.execute(sql_collect, [start_date, end_date, group])
            collect_map = {r["producer"]: r for r in cur.fetchall()}

            sql_qc = """
                SELECT hc.producer,
                    COUNT(*) AS qc_total,
                    SUM(CASE WHEN q.passed = 1 THEN 1 ELSE 0 END) AS qc_passed,
                    SUM(CASE WHEN q.passed = 1 THEN IFNULL(hc.video_seconds, 0) ELSE 0 END) / 3600.0 AS qc_hours
                FROM human_cases hc
                JOIN (
                    SELECT hcn.human_case_id, CASE WHEN hcn.node_status = 3 THEN 1 ELSE 0 END AS passed
                    FROM human_case_nodes hcn
                    WHERE hcn.node_name = 'human_case_sampling' AND hcn.node_status IN (3, 4)
                      AND DATE(hcn.node_updated_at) BETWEEN %s AND %s
                    UNION ALL
                    SELECT hcn.human_case_id, CASE WHEN hcn.node_status = 3 THEN 1 ELSE 0 END AS passed
                    FROM human_case_nodes hcn
                    WHERE hcn.node_name = 'human_case_inspect' AND hcn.node_status IN (3, 4)
                      AND DATE(hcn.node_updated_at) BETWEEN %s AND %s
                      AND hcn.human_case_id NOT IN (
                          SELECT DISTINCT human_case_id FROM human_case_nodes WHERE node_name = 'human_case_sampling'
                      )
                ) q ON q.human_case_id = hc.id
                WHERE hc.deleted_at IS NULL AND hc.produced_by_group = %s
                GROUP BY hc.producer
            """
            cur.execute(sql_qc, [start_date, end_date, start_date, end_date, group])
            qc_map = {r["producer"]: r for r in cur.fetchall()}

            sql_sampling = """
                SELECT hc.producer,
                    COUNT(*) AS sampling_total,
                    SUM(CASE WHEN hcn.node_status = 3 THEN 1 ELSE 0 END) AS sampling_passed,
                    SUM(CASE WHEN hcn.node_status = 3 THEN IFNULL(hc.video_seconds, 0) ELSE 0 END) / 3600.0 AS sampling_hours
                FROM human_cases hc
                JOIN human_case_nodes hcn ON hcn.human_case_id = hc.id
                WHERE hcn.node_name = 'human_case_sampling' AND hcn.node_status IN (3, 4)
                  AND DATE(hcn.node_updated_at) BETWEEN %s AND %s
                  AND hc.deleted_at IS NULL AND hc.produced_by_group = %s
                GROUP BY hc.producer
            """
            cur.execute(sql_sampling, [start_date, end_date, group])
            sampling_map = {r["producer"]: r for r in cur.fetchall()}

            sql_errors = """
                SELECT hc.producer, et.name_cn AS error, COUNT(*) AS cnt
                FROM human_cases hc
                JOIN human_inspect_error_type et ON et.id = hc.inspect_error_type_id
                JOIN human_case_nodes hcn ON hcn.human_case_id = hc.id
                WHERE hcn.node_name = 'human_case_produce_complete' AND hcn.node_status = 3
                  AND DATE(hcn.node_updated_at) BETWEEN %s AND %s
                  AND hc.deleted_at IS NULL AND et.name_cn != '无'
                  AND hc.produced_by_group = %s
                GROUP BY hc.producer, et.name_cn ORDER BY hc.producer, cnt DESC
            """
            cur.execute(sql_errors, [start_date, end_date, group])
            errors_raw = cur.fetchall()
    finally:
        conn.close()

    errors_map = defaultdict(list)
    for r in errors_raw:
        p = r["producer"]
        if len(errors_map[p]) < 5:
            errors_map[p].append({"error": r["error"], "cnt": int(r["cnt"])})

    collectors = []
    all_producers = set(collect_map.keys()) | set(qc_map.keys()) | set(sampling_map.keys())
    for producer in all_producers:
        c = collect_map.get(producer, {})
        q = qc_map.get(producer, {})
        s = sampling_map.get(producer, {})
        qc_passed = int(q.get("qc_passed") or 0)
        qc_total = int(q.get("qc_total") or 0)
        sampling_passed = int(s.get("sampling_passed") or 0)
        sampling_total = int(s.get("sampling_total") or 0)
        collectors.append({
            "name": producer,
            "total_cases": int(c.get("total_cases") or 0),
            "qc_pass_rate": round(qc_passed / qc_total * 100, 1) if qc_total > 0 else 0,
            "sampling_pass_rate": round(sampling_passed / sampling_total * 100, 1) if sampling_total > 0 else 0,
            "total_qc_hours": round(float(q.get("qc_hours") or 0), 1),
            "total_sampling_hours": round(float(s.get("sampling_hours") or 0), 1),
            "top5_errors": errors_map.get(producer, []),
        })
    collectors.sort(key=lambda x: -x["total_cases"])
    result = {"group": group, "collectors": collectors}
    _cache_set(cache_key, result, 300)
    return jsonify(result)


# ═══════════════════════════════════════════════════════════════
#  静态文件
# ═══════════════════════════════════════════════════════════════
@app.route("/")
def index():
    return send_from_directory(".", "index.html")


# ═══════════════════════════════════════════════════════════════
#  项目搜索（Clickhouse）
# ═══════════════════════════════════════════════════════════════
@app.route("/api/projects")
def projects():
    q = request.args.get("q", "").strip()
    if not q:
        return jsonify([])
    cache_key = f"projects:{q.lower()}"
    cached = _cache_get(cache_key)
    if cached is not None:
        return jsonify(cached)
    c = ck()
    rows = c.execute(
        "SELECT toString(uuid), name FROM project "
        "WHERE name ILIKE %(q)s ORDER BY name LIMIT 60",
        {"q": f"%{q}%"}
    )
    seen = set()
    result = []
    for r in rows:
        if r[0] not in seen:
            seen.add(r[0])
            result.append({"id": r[0], "name": r[1]})
    _cache_set(cache_key, result, 600)  # 10分钟
    return jsonify(result)


# ═══════════════════════════════════════════════════════════════
#  存量查询（MySQL，与 delivery tracker 计算逻辑完全一致）
# ═══════════════════════════════════════════════════════════════

def _ph(ids):
    return ",".join(["%s"] * len(ids))

def _query_node(cur, project_ids, node_name, node_status):
    """通用单节点查询：video_seconds"""
    ph = _ph(project_ids)
    cur.execute(f"""
        SELECT hc.project_id,
               COUNT(*) AS cnt,
               SUM(IFNULL(hc.video_seconds, 0)) / 3600.0 AS hours
        FROM human_cases hc
        WHERE hc.project_id IN ({ph})
          AND EXISTS (
              SELECT 1 FROM human_case_nodes hcn
              WHERE hcn.human_case_id = hc.id
                AND hcn.node_name = %s AND hcn.node_status = %s
          )
        GROUP BY hc.project_id
    """, project_ids + [node_name, node_status])
    return {r["project_id"]: float(r["hours"] or 0) for r in cur.fetchall()}


def _query_labeling_inprogress(cur, project_ids):
    """标注中（去重）：semantics OR pose，case 级去重"""
    ph = _ph(project_ids)
    cur.execute(f"""
        SELECT hc.project_id,
               COUNT(*) AS cnt,
               SUM(IFNULL(hc.video_seconds, 0)) / 3600.0 AS hours
        FROM human_cases hc
        WHERE hc.project_id IN ({ph})
          AND EXISTS (
              SELECT 1 FROM human_case_nodes hcn
              WHERE hcn.human_case_id = hc.id
                AND hcn.node_name IN ('semantics_labeling', 'pose_labeling')
                AND hcn.node_status = 1
          )
        GROUP BY hc.project_id
    """, project_ids)
    return {r["project_id"]: float(r["hours"] or 0) for r in cur.fetchall()}


def _query_packaged(cur, project_ids):
    """打包完成：complete_job 最新记录 status=3，取 delivery_video_seconds"""
    ph = _ph(project_ids)
    cur.execute(f"""
        SELECT hc.project_id,
               COUNT(*) AS cnt,
               SUM(IFNULL(hc.delivery_video_seconds, 0)) / 3600.0 AS hours
        FROM human_cases hc
        WHERE hc.project_id IN ({ph})
          AND hc.id IN (
              SELECT hcn.human_case_id
              FROM human_case_nodes hcn
              INNER JOIN (
                  SELECT human_case_id, MAX(id) AS max_id
                  FROM human_case_nodes
                  WHERE project_id IN ({ph}) AND node_name = 'complete_job'
                  GROUP BY human_case_id
              ) latest ON hcn.id = latest.max_id
              WHERE hcn.node_status = 3
          )
        GROUP BY hc.project_id
    """, project_ids * 2)
    return {r["project_id"]: float(r["hours"] or 0) for r in cur.fetchall()}


def _query_qc_pending(cur, project_ids):
    """待质检：human_case_inspect status IN (1,2)，取每个 case 最新一条"""
    ph = _ph(project_ids)
    cur.execute(f"""
        SELECT hc.project_id,
               SUM(IFNULL(hc.video_seconds, 0)) / 3600.0 AS hours
        FROM human_cases hc
        INNER JOIN human_case_nodes hcn ON hc.id = hcn.human_case_id
        WHERE hc.project_id IN ({ph})
          AND hcn.node_name = 'human_case_inspect'
          AND hcn.node_status IN (1, 2)
          AND hcn.id IN (
              SELECT MAX(id)
              FROM human_case_nodes
              WHERE project_id IN ({ph}) AND node_name = 'human_case_inspect'
              GROUP BY human_case_id
          )
        GROUP BY hc.project_id
    """, project_ids * 2)
    return {r["project_id"]: float(r["hours"] or 0) for r in cur.fetchall()}


@app.route("/api/stock")
def stock():
    ids_raw = request.args.get("projects", "").strip()
    force   = request.args.get("force", "").strip()
    if not ids_raw:
        return jsonify({})
    project_ids = safe_uuids([p.strip() for p in ids_raw.split(",") if p.strip()])
    if not project_ids:
        return jsonify({})

    cache_key = "stock:" + ",".join(sorted(project_ids))
    if not force:
        cached = _cache_get(cache_key)
        if cached is not None:
            return jsonify(cached)

    # 从 Clickhouse 拿项目名
    in_clause = ",".join(f"'{pid}'" for pid in project_ids)
    c = ck()
    name_rows = c.execute(
        f"SELECT toString(uuid), name FROM project WHERE toString(uuid) IN ({in_clause})"
    )
    names = {r[0]: r[1] for r in name_rows}

    # 从 MySQL 查各阶段存量（与 delivery tracker 完全一致）
    conn = mysql()
    try:
        with conn.cursor() as cur:
            collected_h  = _query_node(cur, project_ids, "human_case_produce",  3)
            qc_pending_h = _query_qc_pending(cur, project_ids)
            qc_passed_h  = _query_node(cur, project_ids, "human_case_inspect",  3)
            sem_ing_h    = _query_node(cur, project_ids, "semantics_labeling",  1)
            pose_ing_h   = _query_node(cur, project_ids, "pose_labeling",       1)
            lab_ing_h    = _query_labeling_inprogress(cur, project_ids)
            labeled_h    = _query_node(cur, project_ids, "labeling_complete",   3)
            packaged_h   = _query_packaged(cur, project_ids)
    finally:
        conn.close()

    result = {}
    for pid in project_ids:
        result[pid] = {
            "name":         names.get(pid, pid),
            "collected_h":  collected_h.get(pid,  0.0),
            "qc_pending_h": qc_pending_h.get(pid, 0.0),
            "qc_passed_h":  qc_passed_h.get(pid,  0.0),
            "sem_ing_h":    sem_ing_h.get(pid,    0.0),
            "pose_ing_h":   pose_ing_h.get(pid,   0.0),
            "lab_ing_h":    lab_ing_h.get(pid,    0.0),
            "labeled_h":    labeled_h.get(pid,    0.0),
            "packaged_h":   packaged_h.get(pid,   0.0),
            "fetched_at":   __import__("datetime").datetime.now().strftime("%Y-%m-%d %H:%M"),
        }
    _cache_set(cache_key, result, 300)  # 5分钟
    return jsonify(result)


# ═══════════════════════════════════════════════════════════════
#  历史产出（按天聚合）
# ═══════════════════════════════════════════════════════════════

def _query_daily(cur, project_ids, node_name, node_status, start, end):
    """按日聚合某节点某状态的 video_seconds（小时）"""
    ph = _ph(project_ids)
    cur.execute(f"""
        SELECT DATE(hcn.updated_at) AS day,
               SUM(IFNULL(hc.video_seconds, 0)) / 3600.0 AS hours
        FROM human_cases hc
        JOIN human_case_nodes hcn ON hc.id = hcn.human_case_id
        WHERE hc.project_id IN ({ph})
          AND hcn.node_name = %s
          AND hcn.node_status = %s
          AND DATE(hcn.updated_at) BETWEEN %s AND %s
        GROUP BY DATE(hcn.updated_at)
        ORDER BY day
    """, project_ids + [node_name, node_status, start, end])
    return {str(r["day"]): float(r["hours"] or 0) for r in cur.fetchall()}


def _query_daily_mango_dedup(cur, project_ids, node_name, node_status, start, end):
    """Mango专用：按日聚合，同一采集员对同一task只计算一次"""
    ph = _ph(project_ids)
    cur.execute(f"""
        SELECT DATE(first_updated) AS day,
               SUM(video_seconds) / 3600.0 AS hours
        FROM (
            SELECT hc.collector_id, hc.task_id,
                   MIN(hc.video_seconds) AS video_seconds,
                   MIN(hcn.updated_at) AS first_updated
            FROM human_cases hc
            JOIN human_case_nodes hcn ON hc.id = hcn.human_case_id
            WHERE hc.project_id IN ({ph})
              AND hcn.node_name = %s
              AND hcn.node_status = %s
              AND DATE(hcn.updated_at) BETWEEN %s AND %s
            GROUP BY hc.collector_id, hc.task_id
        ) AS dedup
        GROUP BY DATE(first_updated)
        ORDER BY day
    """, project_ids + [node_name, node_status, start, end])
    return {str(r["day"]): float(r["hours"] or 0) for r in cur.fetchall()}


@app.route("/api/history")
def history():
    ids_raw = request.args.get("projects", "").strip()
    start   = request.args.get("start",    "").strip()
    end     = request.args.get("end",      "").strip()
    if not ids_raw or not start or not end:
        return jsonify([])
    project_ids = safe_uuids([p.strip() for p in ids_raw.split(",") if p.strip()])
    if not project_ids:
        return jsonify([])

    conn = mysql()
    try:
        with conn.cursor() as cur:
            collected = _query_daily(cur, project_ids, "human_case_produce",  3, start, end)
            qc_passed = _query_daily(cur, project_ids, "human_case_inspect",  3, start, end)
            labeled   = _query_daily(cur, project_ids, "labeling_complete",   3, start, end)
    finally:
        conn.close()

    # 填满完整日期区间（无数据的天返回 0）
    s, e = date.fromisoformat(start), date.fromisoformat(end)
    result, d = [], s
    while d <= e:
        ds = str(d)
        result.append({
            "date":      ds,
            "collected": collected.get(ds, 0),
            "qc_passed": qc_passed.get(ds, 0),
            "labeled":   labeled.get(ds, 0),
        })
        d += timedelta(days=1)
    return jsonify(result)


# ═══════════════════════════════════════════════════════════════
#  Claude 代理（流式）
# ═══════════════════════════════════════════════════════════════
@app.route("/api/claude", methods=["POST"])
def claude_proxy():
    body = request.get_data()
    data = json.loads(body)
    stream = data.get("stream", False)

    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE

    req = urllib.request.Request(
        LITELLM_URL, data=body, method="POST",
        headers={"Content-Type": "application/json",
                 "Authorization": f"Bearer {LITELLM_KEY}"},
    )

    if stream:
        def generate():
            with urllib.request.urlopen(req, context=ctx) as resp:
                while True:
                    chunk = resp.read(512)
                    if not chunk:
                        break
                    yield chunk
        return Response(generate(), content_type="text/event-stream",
                        headers={"X-Accel-Buffering": "no"})
    else:
        with urllib.request.urlopen(req, context=ctx) as resp:
            return Response(resp.read(), content_type="application/json")


# ═══════════════════════════════════════════════════════════════
#  供应商列表
# ═══════════════════════════════════════════════════════════════
@app.route("/api/vendor-week-compare")
def vendor_week_compare():
    _require_auth()
    # 接受 week=YYYY-Www（周选择器格式）或 date=YYYY-MM-DD（兼容旧调用）
    week_str = request.args.get("week", "")
    today = date.today()
    if week_str:
        try:
            year, w = week_str.split("-W")
            this_week_start = date.fromisocalendar(int(year), int(w), 1)
        except Exception:
            return jsonify({"error": "invalid week"}), 400
    else:
        date_str = request.args.get("date", str(today))
        try:
            d = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            return jsonify({"error": "invalid date"}), 400
        this_week_start = d - timedelta(days=d.weekday())

    this_week_sunday = this_week_start + timedelta(days=6)
    this_week_end    = min(this_week_sunday, today)   # 本周截至今天
    last_week_end    = this_week_start - timedelta(days=1)   # 上周日
    last_week_start  = last_week_end - timedelta(days=6)     # 上周一

    this_days = (this_week_end - this_week_start).days + 1
    last_days = 7

    cache_key = f"vendor_week_compare:{this_week_start}"
    cached = _cache_get(cache_key)
    if cached:
        return jsonify(cached)

    this_stats = _query_vendor_week_stats(str(this_week_start), str(this_week_end))
    last_stats = _query_vendor_week_stats(str(last_week_start), str(last_week_end))
    this_daily_eff = _query_vendor_daily_eff(str(this_week_start), str(this_week_end))
    last_daily_eff = _query_vendor_daily_eff(str(last_week_start), str(last_week_end))

    last_map = {s['vendor']: s for s in last_stats}
    all_vendors = sorted(
        set(s['vendor'] for s in this_stats) | set(last_map.keys())
    )

    vendors = []
    for v in all_vendors:
        tw = next((s for s in this_stats if s['vendor'] == v),
                  {'collect_h': 0.0, 'active_collectors': 0, 'qc_pass_h': 0.0, 'qc_total_h': 0.0})
        lw = last_map.get(v,
                  {'collect_h': 0.0, 'active_collectors': 0, 'qc_pass_h': 0.0, 'qc_total_h': 0.0})
        vendors.append({
            'name': v,
            'this_week_h':            round(tw['collect_h'], 1),
            'last_week_h':            round(lw['collect_h'], 1),
            'this_week_daily_h':      round(tw['collect_h'] / this_days, 1),
            'last_week_daily_h':      round(lw['collect_h'] / last_days, 1),
            'this_week_active':       tw['active_collectors'],
            'last_week_active':       lw['active_collectors'],
            'this_week_qc_rate':      round(tw['qc_pass_h'] / tw['qc_total_h'], 3) if tw['qc_total_h'] > 0 else None,
            'last_week_qc_rate':      round(lw['qc_pass_h'] / lw['qc_total_h'], 3) if lw['qc_total_h'] > 0 else None,
            'tw_daily_collect_eff':   this_daily_eff.get(v),
            'lw_daily_collect_eff':   last_daily_eff.get(v),
        })
    vendors.sort(key=lambda x: x['this_week_h'], reverse=True)

    ot_h  = sum(s['collect_h'] for s in this_stats)
    ol_h  = sum(s['collect_h'] for s in last_stats)
    ot_ac = sum(s['active_collectors'] for s in this_stats)
    ol_ac = sum(s['active_collectors'] for s in last_stats)
    ot_qp = sum(s['qc_pass_h'] for s in this_stats)
    ot_qt = sum(s['qc_total_h'] for s in this_stats)
    ol_qp = sum(s['qc_pass_h'] for s in last_stats)
    ol_qt = sum(s['qc_total_h'] for s in last_stats)

    payload = {
        'this_week': {'start': str(this_week_start), 'end': str(this_week_end), 'days': this_days},
        'last_week': {'start': str(last_week_start), 'end': str(last_week_end), 'days': last_days},
        'overall': {
            'this_week_h':       round(ot_h, 1),
            'last_week_h':       round(ol_h, 1),
            'this_week_daily_h': round(ot_h / this_days, 1),
            'last_week_daily_h': round(ol_h / last_days, 1),
            'this_week_active':  ot_ac,
            'last_week_active':  ol_ac,
            'this_week_qc_rate': round(ot_qp / ot_qt, 3) if ot_qt > 0 else None,
            'last_week_qc_rate': round(ol_qp / ol_qt, 3) if ol_qt > 0 else None,
        },
        'vendors': vendors,
    }
    _cache_set(cache_key, payload, 4 * 3600)
    return jsonify(payload)


@app.route("/api/vendors")
def get_vendors():
    _require_auth()
    cached = _cache_get("vendors")
    if cached is not None:
        return jsonify(cached)
    conn = mysql()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT DISTINCT produced_by_group AS name
                FROM human_cases
                WHERE produced_by_group IS NOT NULL AND produced_by_group != ''
                ORDER BY produced_by_group
            """)
            vendors = [row["name"] for row in cur.fetchall()]
        _cache_set("vendors", vendors, 3600)  # 1小时
        return jsonify(vendors)
    finally:
        conn.close()


# ═══════════════════════════════════════════════════════════════
#  采集员统计
# ═══════════════════════════════════════════════════════════════
@app.route("/api/collectors/sparklines")
def collectors_sparklines():
    """管理员：所有采集员近 N 天每日质检通过时长"""
    if not session.get("authed"):
        return jsonify({"error": "unauthorized"}), 401

    end_date_str = request.args.get("date", str(date.today()))
    days = int(request.args.get("days", 14))

    cache_key = f"collectors_sparklines:{end_date_str}:{days}"
    cached = _cache_get(cache_key)
    if cached:
        return jsonify(cached)

    from datetime import timedelta
    end_date   = date.fromisoformat(end_date_str)
    start_date = end_date - timedelta(days=days - 1)
    date_list  = [(start_date + timedelta(days=i)).isoformat() for i in range(days)]

    conn = mysql()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT hc.producer,
                    DATE(q.node_updated_at) AS day,
                    SUM(CASE WHEN q.passed = 1 THEN IFNULL(hc.video_seconds, 0) ELSE 0 END) / 3600.0 AS qc_h
                FROM human_cases hc
                JOIN (
                    SELECT hcn.human_case_id,
                           CASE WHEN hcn.node_status = 3 THEN 1 ELSE 0 END AS passed,
                           hcn.node_updated_at
                    FROM human_case_nodes hcn
                    WHERE hcn.node_name = 'human_case_sampling'
                      AND hcn.node_status IN (3, 4)
                      AND DATE(hcn.node_updated_at) BETWEEN %s AND %s
                    UNION ALL
                    SELECT hcn.human_case_id,
                           CASE WHEN hcn.node_status = 3 THEN 1 ELSE 0 END AS passed,
                           hcn.node_updated_at
                    FROM human_case_nodes hcn
                    WHERE hcn.node_name = 'human_case_inspect'
                      AND hcn.node_status IN (3, 4)
                      AND DATE(hcn.node_updated_at) BETWEEN %s AND %s
                      AND hcn.human_case_id NOT IN (
                          SELECT DISTINCT human_case_id FROM human_case_nodes
                          WHERE node_name = 'human_case_sampling'
                      )
                ) q ON q.human_case_id = hc.id
                WHERE hc.deleted_at IS NULL
                GROUP BY hc.producer, day
                ORDER BY hc.producer, day
            """, [start_date, end_date_str, start_date, end_date_str])
            rows = cur.fetchall()
    finally:
        conn.close()

    from collections import defaultdict
    day_map = defaultdict(dict)
    for r in rows:
        day_map[r["producer"]][str(r["day"])] = round(float(r["qc_h"] or 0), 2)

    result = {
        "dates": date_list,
        "producers": {
            producer: [day_map[producer].get(d, 0) for d in date_list]
            for producer in day_map
        }
    }
    _cache_set(cache_key, result, 3600)
    return jsonify(result)


@app.route("/api/collectors")
def collectors():
    date_str = request.args.get("date", str(date.today()))
    refresh = request.args.get("refresh") == "1"
    if not refresh:
        cached = _collectors_file_get(date_str)
        if cached:
            return jsonify(cached)
    conn = mysql()
    collect_qc_map = {}
    collect_inspected_map = {}
    upload_h_map = {}
    old_wrist_ids = set()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT
                    hc.id AS case_id,
                    'old' AS db_source,
                    hc.producer,
                    hc.produced_by_group AS vendor,
                    MIN(hcn.node_created_at) AS t_start,
                    MAX(hcn.node_updated_at) AS t_end,
                    MAX(IFNULL(hc.video_seconds, 0)) AS vsec,
                    hc.project_id AS project_id
                FROM human_cases hc
                JOIN human_case_nodes hcn ON hcn.human_case_id = hc.id
                WHERE hcn.node_name = 'human_case_produce_complete'
                  AND hcn.node_status = 3
                  AND DATE(hcn.node_updated_at) = %s
                  AND hc.deleted_at IS NULL
                GROUP BY hc.id, hc.producer, hc.produced_by_group, hc.project_id
                ORDER BY hc.producer, MIN(hcn.node_created_at)
            """, [date_str])
            produce_rows = cur.fetchall()

            # 腕部相机项目 ID 集合（旧库）
            cur.execute("""
                SELECT ptp.project_id FROM project_tag_project ptp
                JOIN project_tag pt ON pt.id = ptp.tag_id
                WHERE pt.tag_name = '腕部Wrist'
            """)
            old_wrist_ids = set(r["project_id"] for r in cur.fetchall())

            # 质检通过/总量（新逻辑）
            # - 有 sampling 节点 → 只看 sampling，status=3 才通过
            # - 无 sampling 节点 → 看 human_case_inspect，status=3 通过
            # - 特殊：有 sampling（任意状态）时，inspect=3 不算通过
            cur.execute("""
                SELECT
                    hc.producer,
                    COUNT(*) AS qc_total_cases,
                    SUM(CASE WHEN q.passed = 1 THEN 1 ELSE 0 END) AS qc_passed_cases,
                    SUM(CASE WHEN q.passed = 1 THEN IFNULL(hc.video_seconds, 0) ELSE 0 END) / 3600.0 AS qc_h,
                    SUM(IFNULL(hc.video_seconds, 0)) / 3600.0 AS qc_total_h
                FROM human_cases hc
                JOIN (
                    -- 有 sampling 且已决定（3/4）的 case，今天更新的
                    SELECT hcn.human_case_id,
                           CASE WHEN hcn.node_status = 3 THEN 1 ELSE 0 END AS passed
                    FROM human_case_nodes hcn
                    WHERE hcn.node_name = 'human_case_sampling'
                      AND hcn.node_status IN (3, 4)
                      AND DATE(hcn.node_updated_at) = %s
                    UNION ALL
                    -- 无 sampling 节点的 case，inspect 今天决定的
                    SELECT hcn.human_case_id,
                           CASE WHEN hcn.node_status = 3 THEN 1 ELSE 0 END AS passed
                    FROM human_case_nodes hcn
                    WHERE hcn.node_name = 'human_case_inspect'
                      AND hcn.node_status IN (3, 4)
                      AND DATE(hcn.node_updated_at) = %s
                      AND NOT EXISTS (
                          SELECT 1 FROM human_case_nodes s
                          WHERE s.human_case_id = hcn.human_case_id AND s.node_name = 'human_case_sampling'
                      )
                ) q ON q.human_case_id = hc.id
                WHERE hc.deleted_at IS NULL
                GROUP BY hc.producer
            """, [date_str, date_str])
            qc_map = {}
            for r in cur.fetchall():
                qc_map[r["producer"]] = {
                    "qc_cases":   int(r["qc_passed_cases"] or 0),
                    "qc_total":   int(r["qc_total_cases"] or 0),
                    "qc_h":       float(r["qc_h"] or 0),
                    "qc_total_h": float(r["qc_total_h"] or 0),
                }

            # ── 采集人效：今日采集中已质检通过 / 已完成质检时长（含 wrist/pt 拆分）──────────
            collect_qc_map = {}
            collect_inspected_map = {}
            wrist_collect_qc_map = {}
            wrist_collect_inspected_map = {}
            pt_collect_qc_map = {}
            pt_collect_inspected_map = {}
            if produce_rows:
                cur.execute("""
                    SELECT hc.producer, hc.project_id,
                        SUM(CASE
                            WHEN EXISTS (SELECT 1 FROM human_case_nodes WHERE human_case_id=hc.id AND node_name='human_case_sampling' AND node_status=3) THEN IFNULL(hc.video_seconds,0)
                            WHEN EXISTS (SELECT 1 FROM human_case_nodes WHERE human_case_id=hc.id AND node_name='human_case_sampling') THEN 0
                            WHEN EXISTS (SELECT 1 FROM human_case_nodes WHERE human_case_id=hc.id AND node_name='human_case_inspect' AND node_status=3) THEN IFNULL(hc.video_seconds,0)
                            ELSE 0
                        END) / 3600.0 AS collect_qc_h,
                        SUM(CASE
                            WHEN EXISTS (SELECT 1 FROM human_case_nodes WHERE human_case_id=hc.id AND node_name='human_case_sampling' AND node_status IN (3,4)) THEN IFNULL(hc.video_seconds,0)
                            WHEN EXISTS (SELECT 1 FROM human_case_nodes WHERE human_case_id=hc.id AND node_name='human_case_sampling') THEN 0
                            WHEN EXISTS (SELECT 1 FROM human_case_nodes WHERE human_case_id=hc.id AND node_name='human_case_inspect' AND node_status IN (3,4)) THEN IFNULL(hc.video_seconds,0)
                            ELSE 0
                        END) / 3600.0 AS collect_inspected_h
                    FROM human_cases hc
                    JOIN human_case_nodes hcn ON hcn.human_case_id=hc.id
                    WHERE hcn.node_name='human_case_produce_complete' AND hcn.node_status=3
                      AND DATE(hcn.node_updated_at)=%s AND hc.deleted_at IS NULL
                    GROUP BY hc.producer, hc.project_id
                """, [date_str])
                for r in cur.fetchall():
                    p = r["producer"]; pid = r["project_id"] or ""
                    qc_h   = float(r["collect_qc_h"] or 0)
                    insp_h = float(r["collect_inspected_h"] or 0)
                    collect_qc_map[p]        = collect_qc_map.get(p, 0) + qc_h
                    collect_inspected_map[p] = collect_inspected_map.get(p, 0) + insp_h
                    if pid in old_wrist_ids:
                        wrist_collect_qc_map[p]        = wrist_collect_qc_map.get(p, 0) + qc_h
                        wrist_collect_inspected_map[p] = wrist_collect_inspected_map.get(p, 0) + insp_h
                    else:
                        pt_collect_qc_map[p]        = pt_collect_qc_map.get(p, 0) + qc_h
                        pt_collect_inspected_map[p] = pt_collect_inspected_map.get(p, 0) + insp_h

            # ── 原始人效：human_case_upload 节点当日 video_seconds ──
            upload_h_map = {}
            wrist_upload_map = {}
            pt_upload_map = {}
            if produce_rows:
                cur.execute("""
                    SELECT hc.producer, hc.project_id,
                        SUM(IFNULL(hc.video_seconds, 0)) / 3600.0 AS upload_h
                    FROM human_cases hc
                    JOIN human_case_nodes hcn ON hcn.human_case_id = hc.id
                    WHERE hcn.node_name = 'human_case_upload'
                      AND hcn.node_status = 3
                      AND DATE(hcn.node_updated_at) = %s
                      AND hc.deleted_at IS NULL
                    GROUP BY hc.producer, hc.project_id
                """, [date_str])
                for r in cur.fetchall():
                    p = r["producer"]; h = float(r["upload_h"] or 0)
                    upload_h_map[p] = upload_h_map.get(p, 0) + h
                    pid = r["project_id"] or ""
                    if pid in old_wrist_ids:
                        wrist_upload_map[p] = wrist_upload_map.get(p, 0) + h
                    else:
                        pt_upload_map[p] = pt_upload_map.get(p, 0) + h

            # ── 累计数据（后台异步刷新，不阻塞请求）──────────────────
            _cumul_cache_key = "coll_cumul:all"
            _cumul = _cache_get(_cumul_cache_key)
            if _cumul:
                total_collect_map, first_collect_map, total_qc_map = _cumul
                pending_qc_map = _cache_get("coll_pending:all") or {}
            else:
                total_collect_map, first_collect_map, total_qc_map, pending_qc_map = {}, {}, {}, {}
                def _refresh_cumul_all():
                    try:
                        _conn = mysql()
                        with _conn.cursor() as _cur:
                            _cur.execute("""
                                SELECT hc.producer,
                                    SUM(IFNULL(hc.video_seconds, 0)) / 3600.0 AS total_collect_h,
                                    MIN(DATE(hcn.node_updated_at)) AS first_collect_date
                                FROM human_cases hc
                                JOIN human_case_nodes hcn ON hcn.human_case_id = hc.id
                                WHERE hcn.node_name = 'human_case_produce_complete'
                                  AND hcn.node_status = 3
                                  AND hc.deleted_at IS NULL
                                GROUP BY hc.producer
                            """)
                            _rows = _cur.fetchall()
                            _tc = {r["producer"]: float(r["total_collect_h"] or 0) for r in _rows}
                            _fc = {r["producer"]: str(r["first_collect_date"]) if r["first_collect_date"] else "" for r in _rows}
                            _cur.execute("""
                                SELECT hc.producer,
                                    COUNT(*) AS total_qc_total,
                                    SUM(CASE WHEN q.passed = 1 THEN 1 ELSE 0 END) AS total_qc_passed,
                                    SUM(CASE WHEN q.passed = 1 THEN IFNULL(hc.video_seconds, 0) ELSE 0 END) / 3600.0 AS total_qc_h,
                                    SUM(IFNULL(hc.video_seconds, 0)) / 3600.0 AS total_qc_total_h
                                FROM human_cases hc
                                JOIN (
                                    SELECT hcn.human_case_id, CASE WHEN hcn.node_status = 3 THEN 1 ELSE 0 END AS passed
                                    FROM human_case_nodes hcn
                                    WHERE hcn.node_name = 'human_case_sampling' AND hcn.node_status IN (3, 4)
                                    UNION ALL
                                    SELECT hcn.human_case_id, CASE WHEN hcn.node_status = 3 THEN 1 ELSE 0 END AS passed
                                    FROM human_case_nodes hcn
                                    LEFT JOIN (
                                        SELECT DISTINCT human_case_id FROM human_case_nodes
                                        WHERE node_name = 'human_case_sampling'
                                    ) s ON s.human_case_id = hcn.human_case_id
                                    WHERE hcn.node_name = 'human_case_inspect' AND hcn.node_status IN (3, 4)
                                      AND s.human_case_id IS NULL
                                ) q ON q.human_case_id = hc.id
                                WHERE hc.deleted_at IS NULL
                                GROUP BY hc.producer
                            """)
                            _tq = {r["producer"]: {
                                "total_qc_h":       float(r["total_qc_h"] or 0),
                                "total_qc_passed":  int(r["total_qc_passed"] or 0),
                                "total_qc_total":   int(r["total_qc_total"] or 0),
                                "total_qc_total_h": float(r["total_qc_total_h"] or 0),
                            } for r in _cur.fetchall()}
                        _conn.close()
                        _cache_set(_cumul_cache_key, (_tc, _fc, _tq), 21600)
                    except Exception as _e:
                        app.logger.error(f"cumul_all bg refresh failed: {_e}")
                import threading as _th
                _th.Thread(target=_refresh_cumul_all, daemon=True).start()
                def _refresh_pending_all():
                    try:
                        _conn = mysql()
                        with _conn.cursor() as _cur:
                            _cur.execute("""
                                SELECT hc.producer,
                                    SUM(IFNULL(hc.video_seconds, 0)) / 3600.0 AS pending_qc_h
                                FROM human_cases hc
                                JOIN (
                                    SELECT hcn.human_case_id FROM human_case_nodes hcn
                                    WHERE hcn.node_name = 'human_case_sampling' AND hcn.node_status IN (1, 2)
                                    UNION ALL
                                    SELECT hcn.human_case_id FROM human_case_nodes hcn
                                    LEFT JOIN (
                                        SELECT DISTINCT human_case_id FROM human_case_nodes
                                        WHERE node_name = 'human_case_sampling'
                                    ) s ON s.human_case_id = hcn.human_case_id
                                    WHERE hcn.node_name = 'human_case_inspect' AND hcn.node_status IN (1, 2)
                                      AND s.human_case_id IS NULL
                                ) pending ON pending.human_case_id = hc.id
                                WHERE hc.deleted_at IS NULL
                                GROUP BY hc.producer
                            """)
                            _pm = {r["producer"]: float(r["pending_qc_h"] or 0) for r in _cur.fetchall()}
                        _conn.close()
                        _cache_set("coll_pending:all", _pm, 3600)
                    except Exception as _e:
                        app.logger.error(f"pending_all bg refresh failed: {_e}")
                _th.Thread(target=_refresh_pending_all, daemon=True).start()

            # 本周质检通过时长（周一到查询日期，限定当日活跃采集员）
            _d = date.fromisoformat(date_str)
            week_start = str(_d - timedelta(days=_d.weekday()))
            active_producers = list(set(r["producer"] for r in produce_rows))
            week_qc_map = {}
            if active_producers:
                _ph = ",".join(["%s"] * len(active_producers))
                cur.execute(f"""
                    SELECT hc.producer,
                        SUM(CASE WHEN q.passed = 1 THEN IFNULL(hc.video_seconds, 0) ELSE 0 END) / 3600.0 AS week_qc_h
                    FROM human_cases hc
                    JOIN (
                        SELECT hcn.human_case_id, CASE WHEN hcn.node_status = 3 THEN 1 ELSE 0 END AS passed
                        FROM human_case_nodes hcn
                        WHERE hcn.node_name = 'human_case_sampling' AND hcn.node_status IN (3, 4)
                          AND DATE(hcn.node_updated_at) BETWEEN %s AND %s
                        UNION ALL
                        SELECT hcn.human_case_id, CASE WHEN hcn.node_status = 3 THEN 1 ELSE 0 END AS passed
                        FROM human_case_nodes hcn
                        LEFT JOIN human_case_nodes samp
                            ON samp.human_case_id = hcn.human_case_id AND samp.node_name = 'human_case_sampling'
                        WHERE hcn.node_name = 'human_case_inspect' AND hcn.node_status IN (3, 4)
                          AND DATE(hcn.node_updated_at) BETWEEN %s AND %s
                          AND samp.human_case_id IS NULL
                    ) q ON q.human_case_id = hc.id
                    WHERE hc.deleted_at IS NULL AND hc.producer IN ({_ph})
                    GROUP BY hc.producer
                """, [week_start, date_str, week_start, date_str] + active_producers)
                week_qc_map = {r["producer"]: float(r["week_qc_h"] or 0) for r in cur.fetchall()}

    finally:
        conn.close()

    # ── 合并新库数据 ──────────────────────────────────────────
    new_produce, new_qc_daily, new_collect_qc, new_collect_inspected, new_upload_h, new_wrist_uuids, new_wrist_upload, new_pt_upload = _query_new_db_daily(date_str)
    wrist_ids = old_wrist_ids | new_wrist_uuids
    produce_rows = _dedupe_produce_rows(list(produce_rows) + list(new_produce))
    qc_map = _merge_qc_map(qc_map, new_qc_daily)
    collect_qc_map        = _merge_float_map(collect_qc_map, new_collect_qc)
    collect_inspected_map = _merge_float_map(collect_inspected_map, new_collect_inspected)
    upload_h_map          = _merge_float_map(upload_h_map, new_upload_h)
    # 合并腕部/PT upload 拆分
    for p, h in new_wrist_upload.items():
        wrist_upload_map[p] = wrist_upload_map.get(p, 0) + h
    for p, h in new_pt_upload.items():
        pt_upload_map[p] = pt_upload_map.get(p, 0) + h

    # 新库累计数据：从缓存取，后台异步刷新（新库无索引，同步调用会卡死 worker）
    _new_cumul_key = "new_db_cumul:all"
    _new_cumul = _cache_get(_new_cumul_key) or _disk_cache_get(_new_cumul_key)
    if _new_cumul:
        n_tc, n_tqm, n_fc = _new_cumul
        total_collect_map = _merge_float_map(total_collect_map, n_tc)
        total_qc_map      = _merge_qc_map(total_qc_map, n_tqm)
        first_collect_map = _merge_date_map(first_collect_map, n_fc)
    else:
        def _bg_new_cumul():
            try:
                n_tc2, n_tqm2, _pi, _ps, n_fc2 = _query_new_db_cumul()
                _cache_set(_new_cumul_key, (n_tc2, n_tqm2, n_fc2), 21600)
                _disk_cache_set(_new_cumul_key, (n_tc2, n_tqm2, n_fc2), 21600)
            except Exception:
                import traceback; traceback.print_exc()
        import threading as _th2
        _th2.Thread(target=_bg_new_cumul, daemon=True).start()

    # 新库本周质检通过时长：从缓存取，后台异步刷新
    all_active = list(set(r["producer"] for r in produce_rows))
    _new_wqc_key = f"new_db_week_qc:{week_start}:{date_str}"
    _cached_new_wqc = _cache_get(_new_wqc_key)
    if _cached_new_wqc is not None:
        week_qc_map = _merge_float_map(week_qc_map, _cached_new_wqc)
    else:
        _all_active_snap = list(all_active)
        def _bg_new_week_qc():
            try:
                nwq = _query_new_db_week_qc(week_start, date_str, _all_active_snap)
                _cache_set(_new_wqc_key, nwq, 1800)
            except Exception:
                import traceback; traceback.print_exc()
        import threading as _th3
        _th3.Thread(target=_bg_new_week_qc, daemon=True).start()

    produce_rows = dedupe_case_rows(produce_rows)

    agg = aggregate_collect_rows(produce_rows)
    sessions_by_p = agg["sessions_by_p"]
    vendor_by_p = agg["vendor_by_p"]
    vsec_by_p = agg["vsec_by_p"]
    cases_by_p = agg["cases_by_p"]

    wrist_vsec_by_p = defaultdict(float)
    pt_vsec_by_p = defaultdict(float)
    for row in produce_rows:
          p = row.get("producer")
          if not p:
              continue
          vsec = float(row.get("vsec") or 0)
          pid = row.get("project_id") or ""
          if pid in wrist_ids:
              wrist_vsec_by_p[p] += vsec
          else:
              pt_vsec_by_p[p] += vsec

    online_info = calc_online_hours(sessions_by_p)
    online_h_map = online_info["online_h_map"]
    first_seen_map = online_info["first_seen_map"]
    last_seen_map = online_info["last_seen_map"]

    persons = []
    for p, segs in sessions_by_p.items():
          c_h = vsec_by_p[p] / 3600
          qc = qc_map.get(p, {})
          q_h = float(qc.get("qc_h") or 0)
          qc_pass = int(qc.get("qc_cases") or 0)
          qc_tot = int(qc.get("qc_total") or 0)
          tqc = total_qc_map.get(p, {})
          tqc_passed = int(tqc.get("total_qc_passed") or 0)
          tqc_total = int(tqc.get("total_qc_total") or 0)
          _tot_h = round(total_collect_map.get(p, 0), 1)
          persons.append({
              "producer":           p,
              "vendor":             vendor_by_p[p],
              "first_collect_date": first_collect_map.get(p, ""),
              "is_new":             _tot_h < 10,
              # 累计
              "total_collect_h":  _tot_h,
              "total_qc_h":       round(tqc.get("total_qc_h") or 0, 1),
              "total_qc_rate":    round(tqc_passed / tqc_total * 100, 1) if tqc_total > 0 else 0,
              # 当日
              "collect_cases":    cases_by_p[p],
              "collect_h":        round(c_h, 2),
              "qc_cases":         qc_pass,
              "qc_total":         qc_tot,
              "qc_h":             round(q_h, 2),
              "qc_rate":          round(qc_pass / qc_tot * 100, 1) if qc_tot > 0 else 0,
              "collect_qc_h":             round(collect_qc_map.get(p, 0), 2),
              "collect_inspected_h":      round(collect_inspected_map.get(p, 0), 2),
              "wrist_collect_qc_h":       round(wrist_collect_qc_map.get(p, 0), 2),
              "wrist_collect_inspected_h":round(wrist_collect_inspected_map.get(p, 0), 2),
              "pt_collect_qc_h":          round(pt_collect_qc_map.get(p, 0), 2),
              "pt_collect_inspected_h":   round(pt_collect_inspected_map.get(p, 0), 2),
              "upload_h":            round(upload_h_map.get(p, 0), 2),
              "wrist_upload_h":      round(wrist_upload_map.get(p, 0), 2),
              "pt_upload_h":         round(pt_upload_map.get(p, 0), 2),
              "wrist_h":             round(wrist_vsec_by_p[p] / 3600, 2),
              "pt_h":                round(pt_vsec_by_p[p] / 3600, 2),
              # 本周
              "week_qc_h":        round(week_qc_map.get(p, 0), 1),
              # 待质检
              "pending_qc_h":     round(pending_qc_map.get(p, 0), 1),
              # 在线
              "online_h":         online_h_map.get(p, 0),
              "first_seen":       first_seen_map.get(p, ""),
              "last_seen":        last_seen_map.get(p, ""),
          })

    vendors_map = defaultdict(list)
    for p in persons:
        vendors_map[p["vendor"]].append(p)

    vendors = []
    for vname, vp in vendors_map.items():
        seniors = [x for x in vp if not x["is_new"]]
        tc      = sum(x["collect_h"] for x in vp)
        tq      = sum(x["qc_h"] for x in vp)
        tqpass  = sum(x["qc_cases"] for x in vp)
        tqtotal = sum(x["qc_total"] for x in vp)
        n       = len(vp)
        stc     = sum(x["collect_h"] for x in seniors)
        sn      = len(seniors)
        vendors.append({
            "vendor":            vname,
            "count":             n,
            "senior_count":      sn,
            "junior_count":      n - sn,
            "avg_collect_h":     round(stc / sn, 2) if sn > 0 else 0,
            "total_collect_h":   round(tc, 2),
            "total_qc_h":        round(tq, 2),
            "qc_rate":           round(tqpass / tqtotal * 100, 1) if tqtotal > 0 else 0,
            "total_wrist_h":     round(sum(x["wrist_h"] for x in vp), 2),
            "total_pt_h":        round(sum(x["pt_h"] for x in vp), 2),
            "avg_wrist_h":       round(sum(x["wrist_h"] for x in vp) / n, 2) if n > 0 else 0,
            "avg_pt_h":          round(sum(x["pt_h"] for x in vp) / n, 2) if n > 0 else 0,
            "persons":           sorted(vp, key=lambda x: -x["collect_h"]),
        })
    vendors.sort(key=lambda x: -x["total_collect_h"])

    seniors_all = [p for p in persons if not p["is_new"]]
    tc_all      = sum(p["collect_h"] for p in persons)
    stc_all     = sum(p["collect_h"] for p in seniors_all)
    tq_all      = sum(p["qc_h"] for p in persons)
    tqpass_all  = sum(p["qc_cases"] for p in persons)
    tqtotal_all = sum(p["qc_total"] for p in persons)
    result = {
        "date":              date_str,
        "total_persons":     len(persons),
        "senior_persons":    len(seniors_all),
        "junior_persons":    len(persons) - len(seniors_all),
        "total_vendors":     len(vendors),
        "total_collect_h":   round(tc_all, 2),
        "senior_avg_collect_h": round(stc_all / len(seniors_all), 2) if seniors_all else 0,
        "total_qc_h":        round(tq_all, 2),
        "overall_qc_rate":   round(tqpass_all / tqtotal_all * 100, 1) if tqtotal_all > 0 else 0,
        "vendors":           vendors,
    }
    _collectors_file_set(date_str, result)
    result["updated_at"] = time.strftime("%H:%M")
    return jsonify(result)


@app.route("/api/collectors/wrist-camera")
def collectors_wrist_camera():
    """腕部相机人效：采集完成时间=目标日期的 case，统计其质检通过情况（不限质检日期）"""
    if not session.get("authed"):
        return jsonify({"error": "unauthorized"}), 401

    date_str = request.args.get("date", str(date.today()))
    refresh = request.args.get("refresh") == "1"
    cache_key = f"wrist_cam:{date_str}"
    cached = _cache_get(cache_key)
    if cached and not refresh:
        return jsonify(cached)

    rows = []

    # ── 旧库：通过 project_tag 找腕部相机项目 ──────────────────────
    conn_old = mysql()
    try:
        with conn_old.cursor() as cur:
            cur.execute("""
                SELECT p.id FROM projects p
                JOIN project_tag_project ptp ON ptp.project_id = p.id
                JOIN project_tag pt ON pt.id = ptp.tag_id
                WHERE pt.tag_name = '腕部Wrist'
            """)
            old_ids = [r["id"] for r in cur.fetchall()]
            if old_ids:
                fmt_old = ",".join(["%s"] * len(old_ids))
                cur.execute(f"""
                    SELECT hc.producer AS producer,
                           hc.produced_by_group AS producer_group,
                           COUNT(*) AS collect_cases,
                           SUM(IFNULL(hc.video_seconds, 0)) / 3600.0 AS collect_h,
                           SUM(CASE
                               WHEN EXISTS(SELECT 1 FROM human_case_nodes n WHERE n.human_case_id=hc.id AND n.node_name='human_case_sampling' AND n.node_status=3) THEN 1
                               WHEN EXISTS(SELECT 1 FROM human_case_nodes n WHERE n.human_case_id=hc.id AND n.node_name='human_case_sampling') THEN 0
                               WHEN EXISTS(SELECT 1 FROM human_case_nodes n WHERE n.human_case_id=hc.id AND n.node_name='human_case_inspect'  AND n.node_status=3) THEN 1
                               ELSE 0 END) AS qc_passed,
                           SUM(CASE WHEN EXISTS(SELECT 1 FROM human_case_nodes n WHERE n.human_case_id=hc.id AND n.node_name='human_case_sampling' AND n.node_status IN(1,2)) THEN 1 ELSE 0 END) AS pending_sampling,
                           SUM(CASE WHEN EXISTS(SELECT 1 FROM human_case_nodes n WHERE n.human_case_id=hc.id AND n.node_name='human_case_inspect' AND n.node_status IN(1,2)) THEN 1 ELSE 0 END) AS pending_inspect,
                           SUM(CASE
                               WHEN EXISTS(SELECT 1 FROM human_case_nodes n WHERE n.human_case_id=hc.id AND n.node_name='human_case_sampling' AND n.node_status=3)
                                   THEN IFNULL(hc.video_seconds, 0)
                               WHEN EXISTS(SELECT 1 FROM human_case_nodes n WHERE n.human_case_id=hc.id AND n.node_name='human_case_inspect'  AND n.node_status=3
                                           AND NOT EXISTS(SELECT 1 FROM human_case_nodes n2 WHERE n2.human_case_id=hc.id AND n2.node_name='human_case_sampling'))
                                   THEN IFNULL(hc.video_seconds, 0)
                               ELSE 0 END) / 3600.0 AS qc_h
                    FROM human_cases hc
                    JOIN human_case_nodes hcn ON hcn.human_case_id = hc.id
                    WHERE hcn.node_name = 'human_case_produce_complete'
                      AND hcn.node_status = 3
                      AND DATE(hcn.node_updated_at) = %s
                      AND hc.project_id IN ({fmt_old})
                      AND hc.deleted_at IS NULL
                    GROUP BY hc.producer, hc.produced_by_group
                """, [date_str] + old_ids)
                rows += cur.fetchall()
    finally:
        conn_old.close()

    # ── 新库：通过 project_tag 找腕部相机项目 ──────────────────────
    conn_new = mysql_new()
    try:
        with conn_new.cursor() as cur:
            cur.execute("""
                SELECT p.uuid FROM project_tag_project ptp
                JOIN project_tag pt ON pt.id = ptp.tag_id
                JOIN project p ON p.id = ptp.project_id
                WHERE pt.tag_name = '腕部Wrist' AND ptp.deleted_at IS NULL
            """)
            new_uuids = [r["uuid"] for r in cur.fetchall()]
            if new_uuids:
                fmt_new = ",".join(["%s"] * len(new_uuids))
                cur.execute(f"""
                    SELECT
                        hc.producer AS producer,
                        ag.name AS producer_group,
                        COUNT(*) AS collect_cases,
                        SUM(IFNULL(CAST(JSON_EXTRACT(pt2.value,'$.data_info.duration') AS DECIMAL(15,4)),0)) / 3600.0 AS collect_h,
                        SUM(CASE
                            WHEN COALESCE(ns.sampling_status3,0)=1 THEN 1
                            WHEN COALESCE(ns.sampling_any,0)=1 THEN 0
                            WHEN COALESCE(ns.inspect_status3,0)=1 THEN 1
                            ELSE 0 END) AS qc_passed,
                        SUM(CASE WHEN COALESCE(ns.sampling_pending,0)=1 THEN 1 ELSE 0 END) AS pending_sampling,
                        SUM(CASE WHEN COALESCE(ns.inspect_pending,0)=1 THEN 1 ELSE 0 END) AS pending_inspect,
                        SUM(CASE
                            WHEN COALESCE(ns.sampling_status3,0)=1 THEN IFNULL(CAST(JSON_EXTRACT(pt2.value,'$.data_info.duration') AS DECIMAL(15,4)),0)
                            WHEN COALESCE(ns.sampling_any,0)=1 THEN 0
                            WHEN COALESCE(ns.inspect_status3,0)=1 THEN IFNULL(CAST(JSON_EXTRACT(pt2.value,'$.data_info.duration') AS DECIMAL(15,4)),0)
                            ELSE 0 END) / 3600.0 AS qc_h
                    FROM human_case hc
                    JOIN human_case_node hcn ON hcn.human_case_id = hc.id
                    LEFT JOIN human_case_tag pt2 ON pt2.human_case_id = hc.id AND pt2.type = 'produce_tags'
                    LEFT JOIN auth.users au ON au.name = hc.producer
                    LEFT JOIN auth.user_group aug ON aug.user_uuid = au.uuid
                    LEFT JOIN auth.`groups` ag ON ag.id = aug.group_uuid AND ag.biz_type = 'produce'
                    LEFT JOIN (
                        SELECT hcn2.human_case_id,
                            MAX(CASE WHEN hcn2.node_name='human_case_sampling' AND hcn2.node_status=3 THEN 1 ELSE 0 END) AS sampling_status3,
                            MAX(CASE WHEN hcn2.node_name='human_case_sampling' THEN 1 ELSE 0 END) AS sampling_any,
                            MAX(CASE WHEN hcn2.node_name='human_case_inspect' AND hcn2.node_status=3 THEN 1 ELSE 0 END) AS inspect_status3,
                            MAX(CASE WHEN hcn2.node_name='human_case_sampling' AND hcn2.node_status IN(1,2) THEN 1 ELSE 0 END) AS sampling_pending,
                            MAX(CASE WHEN hcn2.node_name='human_case_inspect' AND hcn2.node_status IN(1,2) THEN 1 ELSE 0 END) AS inspect_pending
                        FROM human_case_node hcn2
                        WHERE hcn2.node_name IN ('human_case_sampling','human_case_inspect')
                        GROUP BY hcn2.human_case_id
                    ) ns ON ns.human_case_id = hc.id
                    WHERE hcn.node_name = 'human_case_produce_complete'
                      AND hcn.node_status = 3
                      AND DATE(hcn.node_updated_at) = %s
                      AND hc.project_uuid IN ({fmt_new})
                    GROUP BY hc.producer, ag.name
                """, [date_str] + new_uuids)
                rows += cur.fetchall()
    finally:
        conn_new.close()

    result = _build_wrist_like_result(rows, date_str, min_cases=10)
    _cache_set(cache_key, result, 300)
    return jsonify(result)


@app.route("/api/collectors/wrist-camera/range")
def collectors_wrist_camera_range():
    """腕部相机人效近 7 天：按 vendor→producer→date 返回每天质检通过时长"""
    if not session.get("authed"):
        return jsonify({"error": "unauthorized"}), 401

    end   = date.today()
    start = end - timedelta(days=6)
    start_str, end_str = str(start), str(end)
    dates = [str(start + timedelta(days=i)) for i in range(7)]

    # {vendor: {producer: {date: qc_h}}}
    result_map = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))

    # ── 旧库 ──────────────────────────────────────────────────────
    conn_old = mysql()
    try:
        with conn_old.cursor() as cur:
            cur.execute("""SELECT p.id FROM projects p
                JOIN project_tag_project ptp ON ptp.project_id=p.id
                JOIN project_tag pt ON pt.id=ptp.tag_id
                WHERE pt.tag_name='腕部Wrist'""")
            old_ids = [r["id"] for r in cur.fetchall()]
            if old_ids:
                fmt = ",".join(["%s"] * len(old_ids))
                cur.execute(f"""
                    SELECT hc.producer, hc.produced_by_group AS vendor,
                           DATE(hcn.node_updated_at) AS collect_date,
                           SUM(CASE
                               WHEN EXISTS(SELECT 1 FROM human_case_nodes n WHERE n.human_case_id=hc.id AND n.node_name='human_case_sampling' AND n.node_status=3) THEN IFNULL(hc.video_seconds,0)
                               WHEN EXISTS(SELECT 1 FROM human_case_nodes n WHERE n.human_case_id=hc.id AND n.node_name='human_case_sampling') THEN 0
                               WHEN EXISTS(SELECT 1 FROM human_case_nodes n WHERE n.human_case_id=hc.id AND n.node_name='human_case_inspect' AND n.node_status=3) THEN IFNULL(hc.video_seconds,0)
                               ELSE 0 END) / 3600.0 AS qc_h
                    FROM human_cases hc
                    JOIN human_case_nodes hcn ON hcn.human_case_id=hc.id
                    WHERE hcn.node_name='human_case_produce_complete' AND hcn.node_status=3
                      AND DATE(hcn.node_updated_at) BETWEEN %s AND %s
                      AND hc.project_id IN ({fmt}) AND hc.deleted_at IS NULL
                    GROUP BY hc.producer, hc.produced_by_group, collect_date
                """, [start_str, end_str] + old_ids)
                for r in cur.fetchall():
                    vendor = r["vendor"] or "未知"
                    result_map[vendor][r["producer"]][str(r["collect_date"])] += float(r["qc_h"] or 0)
    finally:
        conn_old.close()

    # ── 新库 ──────────────────────────────────────────────────────
    conn_new = mysql_new()
    try:
        with conn_new.cursor() as cur:
            cur.execute("""SELECT ptp.project_id FROM project_tag_project ptp
                JOIN project_tag pt ON pt.id=ptp.tag_id
                WHERE pt.tag_name='腕部Wrist' AND ptp.deleted_at IS NULL""")
            new_uuids = [r["project_id"] for r in cur.fetchall()]
            if new_uuids:
                fmt = ",".join(["%s"] * len(new_uuids))
                cur.execute(f"""
                    SELECT hc.producer AS producer,
                           hc.produced_by_group AS vendor,
                           DATE(hcn.node_updated_at) AS collect_date,
                           SUM(CASE
                               WHEN COALESCE(ns.sampling_status3,0)=1 THEN IFNULL(hc.video_seconds,0)
                               WHEN COALESCE(ns.sampling_any,0)=1 THEN 0
                               WHEN COALESCE(ns.inspect_status3,0)=1 THEN IFNULL(hc.video_seconds,0)
                               ELSE 0 END) / 3600.0 AS qc_h
                    FROM human_cases hc
                    JOIN human_case_nodes hcn ON hcn.human_case_id=hc.id
                    LEFT JOIN (
                        SELECT hcn2.human_case_id,
                            MAX(CASE WHEN hcn2.node_name='human_case_sampling' AND hcn2.node_status=3 THEN 1 ELSE 0 END) AS sampling_status3,
                            MAX(CASE WHEN hcn2.node_name='human_case_sampling' THEN 1 ELSE 0 END) AS sampling_any,
                            MAX(CASE WHEN hcn2.node_name='human_case_inspect' AND hcn2.node_status=3 THEN 1 ELSE 0 END) AS inspect_status3
                        FROM human_case_nodes hcn2
                        WHERE hcn2.node_name IN ('human_case_sampling','human_case_inspect')
                        GROUP BY hcn2.human_case_id
                    ) ns ON ns.human_case_id = hc.id
                    WHERE hcn.node_name='human_case_produce_complete' AND hcn.node_status=3
                      AND DATE(hcn.node_updated_at) BETWEEN %s AND %s
                      AND hc.project_id IN ({fmt}) AND hc.deleted_at IS NULL
                    GROUP BY producer, vendor, collect_date
                """, [start_str, end_str] + new_uuids)
                for r in cur.fetchall():
                    vendor = r["vendor"] or "未知"
                    result_map[vendor][r["producer"]][str(r["collect_date"])] += float(r["qc_h"] or 0)
    finally:
        conn_new.close()

    # 整理成前端需要的结构
    vendors = {}
    for vendor, persons in result_map.items():
        vendors[vendor] = {
            producer: {d: round(day_map.get(d, 0), 2) for d in dates}
            for producer, day_map in persons.items()
        }

    return jsonify({"dates": dates, "vendors": vendors})


# ── P+T 人效（非腕部相机项目）────────────────────────────────────────────────

def _build_wrist_like_result(rows, date_str, min_cases=10):
    """通用汇总：按供应商分组。collect_cases <= min_cases 的人仍展示但不计入聚合指标。"""
    vendor_map = defaultdict(lambda: {"persons": [], "collect_cases": 0, "collect_h": 0.0,
                                      "qc_passed": 0, "pending_inspect": 0,
                                      "pending_sampling": 0, "qc_h": 0.0,
                                      "person_count": 0})
    total_persons = 0
    for r in rows:
        collect_cases    = int(r["collect_cases"] or 0)
        grp              = r.get("producer_group") or "未知"
        qc_passed        = int(r["qc_passed"] or 0)
        pending_inspect  = int(r["pending_inspect"] or 0)
        pending_sampling = int(r["pending_sampling"] or 0)
        qc_h             = float(r["qc_h"] or 0)
        collect_h        = float(r["collect_h"] or 0)
        is_filtered      = collect_cases <= min_cases  # 采集数过少，不计入人均
        vendor_map[grp]["persons"].append({
            "producer":         r["producer"],
            "collect_cases":    collect_cases,
            "collect_h":        round(collect_h, 2),
            "qc_passed":        qc_passed,
            "pending_inspect":  pending_inspect,
            "pending_sampling": pending_sampling,
            "qc_h":             round(qc_h, 2),
            "is_filtered":      is_filtered,
        })
        if not is_filtered:
            vendor_map[grp]["collect_cases"]    += collect_cases
            vendor_map[grp]["collect_h"]        += collect_h
            vendor_map[grp]["qc_passed"]        += qc_passed
            vendor_map[grp]["pending_inspect"]  += pending_inspect
            vendor_map[grp]["pending_sampling"] += pending_sampling
            vendor_map[grp]["qc_h"]             += qc_h
            vendor_map[grp]["person_count"]     += 1
            total_persons += 1

    vendors = []
    for grp, v in vendor_map.items():
        v["persons"].sort(key=lambda x: -x["qc_passed"])
        vendors.append({
            "group":            grp,
            "person_count":     v["person_count"],
            "collect_cases":    v["collect_cases"],
            "collect_h":        round(v["collect_h"], 2),
            "qc_passed":        v["qc_passed"],
            "pending_inspect":  v["pending_inspect"],
            "pending_sampling": v["pending_sampling"],
            "qc_h":             round(v["qc_h"], 2),
            "persons":          v["persons"],
        })
    vendors.sort(key=lambda x: -x["qc_passed"])
    return {
        "date":                date_str,
        "total_persons":       total_persons,
        "total_collect_cases": sum(v["collect_cases"] for v in vendors),
        "total_qc_cases":      sum(v["qc_passed"] for v in vendors),
        "total_qc_h":          round(sum(v["qc_h"] for v in vendors), 2),
        "vendors":             vendors,
    }


@app.route("/api/collectors/pt-camera")
def collectors_pt_camera():
    """P+T 人效：腕部相机以外的所有项目，当天采集 >10 条才计入。"""
    if not session.get("authed"):
        return jsonify({"error": "unauthorized"}), 401

    date_str  = request.args.get("date", str(date.today()))
    refresh   = request.args.get("refresh") == "1"
    cache_key = f"pt_cam:{date_str}"
    cached    = _cache_get(cache_key)
    if cached and not refresh:
        return jsonify(cached)

    rows = []

    # ── 旧库：排除腕部Wrist项目 ──────────────────────────────────────────
    conn_old = mysql()
    try:
        with conn_old.cursor() as cur:
            cur.execute("""
                SELECT id FROM projects p
                WHERE p.id IN (
                    SELECT ptp.project_id FROM project_tag_project ptp
                    JOIN project_tag pt ON pt.id = ptp.tag_id
                    WHERE pt.tag_name = '腕部Wrist'
                )
            """)
            wrist_old_ids = [r["id"] for r in cur.fetchall()]
            excl = f"AND hc.project_id NOT IN ({','.join(['%s']*len(wrist_old_ids))})" if wrist_old_ids else ""
            cur.execute(f"""
                SELECT hc.producer AS producer,
                       hc.produced_by_group AS producer_group,
                       COUNT(*) AS collect_cases,
                       SUM(IFNULL(hc.video_seconds, 0)) / 3600.0 AS collect_h,
                       SUM(CASE
                           WHEN EXISTS(SELECT 1 FROM human_case_nodes n WHERE n.human_case_id=hc.id AND n.node_name='human_case_sampling' AND n.node_status=3) THEN 1
                           WHEN EXISTS(SELECT 1 FROM human_case_nodes n WHERE n.human_case_id=hc.id AND n.node_name='human_case_sampling') THEN 0
                           WHEN EXISTS(SELECT 1 FROM human_case_nodes n WHERE n.human_case_id=hc.id AND n.node_name='human_case_inspect'  AND n.node_status=3) THEN 1
                           ELSE 0 END) AS qc_passed,
                       SUM(CASE WHEN EXISTS(SELECT 1 FROM human_case_nodes n WHERE n.human_case_id=hc.id AND n.node_name='human_case_sampling' AND n.node_status IN(1,2)) THEN 1 ELSE 0 END) AS pending_sampling,
                       SUM(CASE WHEN EXISTS(SELECT 1 FROM human_case_nodes n WHERE n.human_case_id=hc.id AND n.node_name='human_case_inspect' AND n.node_status IN(1,2)) THEN 1 ELSE 0 END) AS pending_inspect,
                       SUM(CASE
                           WHEN EXISTS(SELECT 1 FROM human_case_nodes n WHERE n.human_case_id=hc.id AND n.node_name='human_case_sampling' AND n.node_status=3)
                               THEN IFNULL(hc.video_seconds, 0)
                           WHEN EXISTS(SELECT 1 FROM human_case_nodes n WHERE n.human_case_id=hc.id AND n.node_name='human_case_inspect' AND n.node_status=3
                                       AND NOT EXISTS(SELECT 1 FROM human_case_nodes n2 WHERE n2.human_case_id=hc.id AND n2.node_name='human_case_sampling'))
                               THEN IFNULL(hc.video_seconds, 0)
                           ELSE 0 END) / 3600.0 AS qc_h
                FROM human_cases hc
                JOIN human_case_nodes hcn ON hcn.human_case_id = hc.id
                WHERE hcn.node_name = 'human_case_produce_complete'
                  AND hcn.node_status = 3
                  AND DATE(hcn.node_updated_at) = %s
                  AND hc.deleted_at IS NULL
                  {excl}
                GROUP BY hc.producer, hc.produced_by_group
            """, [date_str] + wrist_old_ids)
            rows += cur.fetchall()
    finally:
        conn_old.close()

    # ── 新库：排除腕部Wrist项目 ──────────────────────────────────────────
    conn_new = mysql_new()
    try:
        with conn_new.cursor() as cur:
            cur.execute("""
                SELECT p.uuid FROM project_tag_project ptp
                JOIN project_tag pt ON pt.id = ptp.tag_id
                JOIN project p ON p.id = ptp.project_id
                WHERE pt.tag_name = '腕部Wrist' AND ptp.deleted_at IS NULL
            """)
            wrist_new_uuids = [r["uuid"] for r in cur.fetchall()]
            excl_new = f"AND hc.project_uuid NOT IN ({','.join(['%s']*len(wrist_new_uuids))})" if wrist_new_uuids else ""
            cur.execute(f"""
                SELECT
                    hc.producer AS producer,
                    ag.name AS producer_group,
                    COUNT(*) AS collect_cases,
                    SUM(IFNULL(CAST(JSON_EXTRACT(pt2.value,'$.data_info.duration') AS DECIMAL(15,4)),0)) / 3600.0 AS collect_h,
                    SUM(CASE
                        WHEN COALESCE(ns.sampling_status3,0)=1 THEN 1
                        WHEN COALESCE(ns.sampling_any,0)=1 THEN 0
                        WHEN COALESCE(ns.inspect_status3,0)=1 THEN 1
                        ELSE 0 END) AS qc_passed,
                    SUM(CASE WHEN COALESCE(ns.sampling_pending,0)=1 THEN 1 ELSE 0 END) AS pending_sampling,
                    SUM(CASE WHEN COALESCE(ns.inspect_pending,0)=1 THEN 1 ELSE 0 END) AS pending_inspect,
                    SUM(CASE
                        WHEN COALESCE(ns.sampling_status3,0)=1 THEN IFNULL(CAST(JSON_EXTRACT(pt2.value,'$.data_info.duration') AS DECIMAL(15,4)),0)
                        WHEN COALESCE(ns.sampling_any,0)=1 THEN 0
                        WHEN COALESCE(ns.inspect_status3,0)=1 THEN IFNULL(CAST(JSON_EXTRACT(pt2.value,'$.data_info.duration') AS DECIMAL(15,4)),0)
                        ELSE 0 END) / 3600.0 AS qc_h
                FROM human_case hc
                JOIN human_case_node hcn ON hcn.human_case_id = hc.id
                LEFT JOIN human_case_tag pt2 ON pt2.human_case_id = hc.id AND pt2.type = 'produce_tags'
                LEFT JOIN auth.users au ON au.name = hc.producer
                LEFT JOIN auth.user_group aug ON aug.user_uuid = au.uuid
                LEFT JOIN auth.`groups` ag ON ag.id = aug.group_uuid AND ag.biz_type = 'produce'
                LEFT JOIN (
                    SELECT hcn2.human_case_id,
                        MAX(CASE WHEN hcn2.node_name='human_case_sampling' AND hcn2.node_status=3 THEN 1 ELSE 0 END) AS sampling_status3,
                        MAX(CASE WHEN hcn2.node_name='human_case_sampling' THEN 1 ELSE 0 END) AS sampling_any,
                        MAX(CASE WHEN hcn2.node_name='human_case_inspect' AND hcn2.node_status=3 THEN 1 ELSE 0 END) AS inspect_status3,
                        MAX(CASE WHEN hcn2.node_name='human_case_sampling' AND hcn2.node_status IN(1,2) THEN 1 ELSE 0 END) AS sampling_pending,
                        MAX(CASE WHEN hcn2.node_name='human_case_inspect' AND hcn2.node_status IN(1,2) THEN 1 ELSE 0 END) AS inspect_pending
                    FROM human_case_node hcn2
                    WHERE hcn2.node_name IN ('human_case_sampling','human_case_inspect')
                    GROUP BY hcn2.human_case_id
                ) ns ON ns.human_case_id = hc.id
                WHERE hcn.node_name = 'human_case_produce_complete'
                  AND hcn.node_status = 3
                  AND DATE(hcn.node_updated_at) = %s
                  {excl_new}
                GROUP BY hc.producer, ag.name
            """, [date_str] + wrist_new_uuids)
            rows += cur.fetchall()
    finally:
        conn_new.close()

    result = _build_wrist_like_result(rows, date_str, min_cases=10)
    _cache_set(cache_key, result, 300)
    return jsonify(result)


# ── 关注采集员 ────────────────────────────────────────────────────────────────

WATCHLIST_FILE = os.path.join(os.path.dirname(__file__), "watchlist.json")


def _load_watchlist():
    if os.path.exists(WATCHLIST_FILE):
        try:
            with open(WATCHLIST_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {"wrist": [], "pt": []}


def _save_watchlist(data):
    with open(WATCHLIST_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)


@app.route("/api/watch/list")
def watch_list():
    if not session.get("authed"):
        return jsonify({"error": "unauthorized"}), 401
    tab = request.args.get("tab", "wrist")
    wl  = _load_watchlist()
    return jsonify({"tab": tab, "producers": wl.get(tab, [])})


@app.route("/api/watch/toggle", methods=["POST"])
def watch_toggle():
    if not session.get("authed"):
        return jsonify({"error": "unauthorized"}), 401
    body     = request.get_json() or {}
    producer = (body.get("producer") or "").strip()
    tab      = body.get("tab", "wrist")
    if not producer or tab not in ("wrist", "pt"):
        return jsonify({"error": "invalid"}), 400
    wl  = _load_watchlist()
    lst = wl.setdefault(tab, [])
    if producer in lst:
        lst.remove(producer)
        watching = False
    else:
        lst.append(producer)
        watching = True
    _save_watchlist(wl)
    return jsonify({"tab": tab, "producer": producer, "watching": watching,
                    "producers": lst})


@app.route("/api/vendor/watch/list")
def vendor_watch_list():
    grp = session.get("vendor_group")
    if not grp:
        return jsonify({"error": "unauthorized"}), 401
    wl = _load_watchlist()
    key = f"vendor_{grp}"
    return jsonify({"producers": wl.get(key, [])})


@app.route("/api/vendor/watch/toggle", methods=["POST"])
def vendor_watch_toggle():
    grp = session.get("vendor_group")
    if not grp:
        return jsonify({"error": "unauthorized"}), 401
    body = request.get_json() or {}
    producer = (body.get("producer") or "").strip()
    if not producer:
        return jsonify({"error": "invalid"}), 400
    wl = _load_watchlist()
    key = f"vendor_{grp}"
    lst = wl.setdefault(key, [])
    if producer in lst:
        lst.remove(producer)
        watching = False
    else:
        lst.append(producer)
        watching = True
    _save_watchlist(wl)
    return jsonify({"producer": producer, "watching": watching, "producers": lst})


@app.route("/api/collectors/by-project")
def collectors_by_project():
    date_str = request.args.get("date", str(date.today()))
    refresh = request.args.get("refresh") == "1"
    cache_key = f"coll_by_proj:{date_str}"
    cached = _cache_get(cache_key)
    if cached and not refresh:
        return jsonify(cached)

    conn = mysql()
    try:
        with conn.cursor() as cur:
            # Step 1: 当日采集 - 按 project_id 聚合（不 JOIN projects，避免扫描开销）
            cur.execute("""
                SELECT hc.project_id,
                       COUNT(DISTINCT hc.id) AS collect_cases,
                       SUM(IFNULL(hc.video_seconds, 0)) / 3600.0 AS collect_h,
                       COUNT(DISTINCT hc.producer) AS collector_count
                FROM human_cases hc
                JOIN human_case_nodes hcn ON hcn.human_case_id = hc.id
                WHERE hcn.node_name = 'human_case_produce_complete'
                  AND hcn.node_status = 3
                  AND DATE(hcn.node_updated_at) = %s
                  AND hc.deleted_at IS NULL
                GROUP BY hc.project_id
            """, [date_str])
            collect_rows = {r["project_id"]: r for r in cur.fetchall()}

            if not collect_rows:
                return jsonify({"date": date_str, "projects": []})

            # Step 2: 查项目名称（主键 IN 查询，极快）
            pids = list(collect_rows.keys())
            ph = ",".join(["%s"] * len(pids))
            cur.execute(f"SELECT id, name FROM projects WHERE id IN ({ph})", pids)
            name_map = {r["id"]: r["name"] for r in cur.fetchall()}

            # Step 3: 当日质检 - 限定 project_id IN 活跃项目，大幅减少扫描范围
            cur.execute(f"""
                SELECT hc.project_id,
                       COUNT(*) AS qc_total,
                       SUM(CASE WHEN q.passed = 1 THEN 1 ELSE 0 END) AS qc_passed,
                       SUM(CASE WHEN q.passed = 1 THEN IFNULL(hc.video_seconds, 0) ELSE 0 END) / 3600.0 AS qc_h
                FROM human_cases hc
                JOIN (
                    SELECT hcn.human_case_id, CASE WHEN hcn.node_status = 3 THEN 1 ELSE 0 END AS passed
                    FROM human_case_nodes hcn
                    WHERE hcn.node_name = 'human_case_sampling'
                      AND hcn.node_status IN (3, 4)
                      AND DATE(hcn.node_updated_at) = %s
                    UNION ALL
                    SELECT hcn.human_case_id, CASE WHEN hcn.node_status = 3 THEN 1 ELSE 0 END AS passed
                    FROM human_case_nodes hcn
                    WHERE hcn.node_name = 'human_case_inspect'
                      AND hcn.node_status IN (3, 4)
                      AND DATE(hcn.node_updated_at) = %s
                      AND hcn.human_case_id NOT IN (
                          SELECT DISTINCT human_case_id FROM human_case_nodes WHERE node_name = 'human_case_sampling'
                      )
                ) q ON q.human_case_id = hc.id
                WHERE hc.deleted_at IS NULL
                  AND hc.project_id IN ({ph})
                GROUP BY hc.project_id
            """, [date_str, date_str] + pids)
            qc_rows = {r["project_id"]: r for r in cur.fetchall()}
    finally:
        conn.close()

    # ── 新库：合并今日采集项目 ────────────────────────────────────
    conn_new = mysql_new()
    try:
        with conn_new.cursor() as cur2:
            cur2.execute("""
                SELECT hc.project_id AS pid,
                       COUNT(DISTINCT hc.id) AS collect_cases,
                       SUM(IFNULL(hc.video_seconds, 0)) / 3600.0 AS collect_h,
                       COUNT(DISTINCT hc.producer) AS collector_count
                FROM human_cases hc
                JOIN human_case_nodes hcn ON hcn.human_case_id = hc.id
                WHERE hcn.node_name = 'human_case_produce_complete'
                  AND hcn.node_status = 3
                  AND DATE(hcn.node_updated_at) = %s
                  AND hc.deleted_at IS NULL
                GROUP BY hc.project_id
            """, [date_str])
            new_collect_rows = {r["pid"]: r for r in cur2.fetchall()}

            new_qc_rows = {}
            if new_collect_rows:
                new_pids = list(new_collect_rows.keys())
                nph = ",".join(["%s"] * len(new_pids))
                cur2.execute(f"""
                    SELECT hc.project_id AS pid,
                           COUNT(*) AS qc_total,
                           SUM(CASE WHEN q.passed=1 THEN 1 ELSE 0 END) AS qc_passed,
                           SUM(CASE WHEN q.passed=1 THEN IFNULL(hc.video_seconds,0) ELSE 0 END)/3600.0 AS qc_h
                    FROM human_cases hc
                    JOIN (
                        SELECT hcn.human_case_id, CASE WHEN hcn.node_status=3 THEN 1 ELSE 0 END AS passed
                        FROM human_case_nodes hcn
                        WHERE hcn.node_name='human_case_sampling' AND hcn.node_status IN(3,4)
                          AND DATE(hcn.node_updated_at)=%s
                        UNION ALL
                        SELECT hcn.human_case_id, CASE WHEN hcn.node_status=3 THEN 1 ELSE 0 END AS passed
                        FROM human_case_nodes hcn
                        WHERE hcn.node_name='human_case_inspect' AND hcn.node_status IN(3,4)
                          AND DATE(hcn.node_updated_at)=%s
                          AND NOT EXISTS(SELECT 1 FROM human_case_nodes s WHERE s.human_case_id=hcn.human_case_id AND s.node_name='human_case_sampling')
                    ) q ON q.human_case_id=hc.id
                    WHERE hc.deleted_at IS NULL AND hc.project_id IN ({nph})
                    GROUP BY hc.project_id
                """, [date_str, date_str] + new_pids)
                new_qc_rows = {r["pid"]: r for r in cur2.fetchall()}

            # 新库项目名称
            if new_collect_rows:
                new_pids2 = list(new_collect_rows.keys())
                nph2 = ",".join(["%s"] * len(new_pids2))
                cur2.execute(f"SELECT id, name FROM projects WHERE id IN ({nph2}) AND deleted_at IS NULL", new_pids2)
                for r in cur2.fetchall():
                    if r["id"] in new_collect_rows:
                        new_collect_rows[r["id"]]["project_name"] = r["name"]
    finally:
        conn_new.close()

    projects_out = []
    for pid, c in collect_rows.items():
        q = qc_rows.get(pid, {})
        qc_passed = int(q.get("qc_passed") or 0)
        qc_total  = int(q.get("qc_total") or 0)
        projects_out.append({
            "project_id":      pid,
            "project_name":    name_map.get(pid, pid),
            "collect_cases":   int(c["collect_cases"]),
            "collect_h":       round(float(c["collect_h"] or 0), 2),
            "qc_h":            round(float(q.get("qc_h") or 0), 2),
            "qc_passed":       qc_passed,
            "qc_total":        qc_total,
            "qc_rate":         round(qc_passed / qc_total * 100, 1) if qc_total > 0 else 0,
            "collector_count": int(c["collector_count"]),
        })
    for pid, c in new_collect_rows.items():
        q = new_qc_rows.get(pid, {})
        qc_passed = int(q.get("qc_passed") or 0)
        qc_total  = int(q.get("qc_total") or 0)
        projects_out.append({
            "project_id":      pid,
            "project_name":    c.get("project_name") or pid,
            "collect_cases":   int(c["collect_cases"]),
            "collect_h":       round(float(c["collect_h"] or 0), 2),
            "qc_h":            round(float(q.get("qc_h") or 0), 2),
            "qc_passed":       qc_passed,
            "qc_total":        qc_total,
            "qc_rate":         round(qc_passed / qc_total * 100, 1) if qc_total > 0 else 0,
            "collector_count": int(c["collector_count"]),
        })
    projects_out.sort(key=lambda x: -x["collect_h"])

    result = {"date": date_str, "projects": projects_out}
    _cache_set(cache_key, result, 300)
    return jsonify(result)


@app.route("/api/collector-stats")
def collector_stats():
    """采集员绩效统计（日期范围 + 用户组筛选）"""
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")
    group = request.args.get("group", "")

    # 缓存 key
    cache_key = f"collector_stats:{start_date}:{end_date}:{group}"
    cached = _cache_get(cache_key)
    if cached:
        return jsonify(cached)

    conn = mysql()
    try:
        with conn.cursor() as cur:
            # 1. 采集完成数据
            sql_collect = """
                SELECT hc.producer, hc.produced_by_group AS grp,
                       COUNT(DISTINCT hc.id) AS total_cases
                FROM human_cases hc
                JOIN human_case_nodes hcn ON hcn.human_case_id = hc.id
                WHERE hcn.node_name = 'human_case_produce_complete'
                  AND hcn.node_status = 3
                  AND DATE(hcn.node_updated_at) BETWEEN %s AND %s
                  AND hc.deleted_at IS NULL
            """
            params_collect = [start_date, end_date]
            if group:
                sql_collect += " AND hc.produced_by_group = %s"
                params_collect.append(group)
            sql_collect += " GROUP BY hc.producer, hc.produced_by_group"

            cur.execute(sql_collect, params_collect)
            collect_map = {r["producer"]: r for r in cur.fetchall()}

            # 2. 质检通过数据（sampling 优先）
            sql_qc = """
                SELECT
                    hc.producer,
                    COUNT(*) AS qc_total,
                    SUM(CASE WHEN q.passed = 1 THEN 1 ELSE 0 END) AS qc_passed,
                    SUM(CASE WHEN q.passed = 1 THEN IFNULL(hc.video_seconds, 0) ELSE 0 END) / 3600.0 AS qc_hours
                FROM human_cases hc
                JOIN (
                    SELECT hcn.human_case_id,
                           CASE WHEN hcn.node_status = 3 THEN 1 ELSE 0 END AS passed
                    FROM human_case_nodes hcn
                    WHERE hcn.node_name = 'human_case_sampling'
                      AND hcn.node_status IN (3, 4)
                      AND DATE(hcn.node_updated_at) BETWEEN %s AND %s
                    UNION ALL
                    SELECT hcn.human_case_id,
                           CASE WHEN hcn.node_status = 3 THEN 1 ELSE 0 END AS passed
                    FROM human_case_nodes hcn
                    WHERE hcn.node_name = 'human_case_inspect'
                      AND hcn.node_status IN (3, 4)
                      AND DATE(hcn.node_updated_at) BETWEEN %s AND %s
                      AND hcn.human_case_id NOT IN (
                          SELECT DISTINCT human_case_id FROM human_case_nodes WHERE node_name = 'human_case_sampling'
                      )
                ) q ON q.human_case_id = hc.id
                WHERE hc.deleted_at IS NULL
            """
            params_qc = [start_date, end_date, start_date, end_date]
            if group:
                sql_qc += " AND hc.produced_by_group = %s"
                params_qc.append(group)
            sql_qc += " GROUP BY hc.producer"

            cur.execute(sql_qc, params_qc)
            qc_map = {r["producer"]: r for r in cur.fetchall()}

            # 3. 抽检通过数据
            sql_sampling = """
                SELECT
                    hc.producer,
                    COUNT(*) AS sampling_total,
                    SUM(CASE WHEN hcn.node_status = 3 THEN 1 ELSE 0 END) AS sampling_passed,
                    SUM(CASE WHEN hcn.node_status = 3 THEN IFNULL(hc.video_seconds, 0) ELSE 0 END) / 3600.0 AS sampling_hours
                FROM human_cases hc
                JOIN human_case_nodes hcn ON hcn.human_case_id = hc.id
                WHERE hcn.node_name = 'human_case_sampling'
                  AND hcn.node_status IN (3, 4)
                  AND DATE(hcn.node_updated_at) BETWEEN %s AND %s
                  AND hc.deleted_at IS NULL
            """
            params_sampling = [start_date, end_date]
            if group:
                sql_sampling += " AND hc.produced_by_group = %s"
                params_sampling.append(group)
            sql_sampling += " GROUP BY hc.producer"

            cur.execute(sql_sampling, params_sampling)
            sampling_map = {r["producer"]: r for r in cur.fetchall()}

            # 4. Top 错误类型（排除"无"，按出现次数降序，取前5）
            sql_errors = """
                SELECT hc.producer, et.name_cn AS error, COUNT(*) AS cnt
                FROM human_cases hc
                JOIN human_inspect_error_type et ON et.id = hc.inspect_error_type_id
                JOIN human_case_nodes hcn ON hcn.human_case_id = hc.id
                WHERE hcn.node_name = 'human_case_produce_complete'
                  AND hcn.node_status = 3
                  AND DATE(hcn.node_updated_at) BETWEEN %s AND %s
                  AND hc.deleted_at IS NULL
                  AND et.name_cn != '无'
            """
            params_errors = [start_date, end_date]
            if group:
                sql_errors += " AND hc.produced_by_group = %s"
                params_errors.append(group)
            sql_errors += " GROUP BY hc.producer, et.name_cn ORDER BY hc.producer, cnt DESC"

            cur.execute(sql_errors, params_errors)
            errors_raw = cur.fetchall()

    finally:
        conn.close()

    # 整理 top errors：每个 producer 取前5
    from collections import defaultdict
    errors_map = defaultdict(list)
    for r in errors_raw:
        p = r["producer"]
        if len(errors_map[p]) < 5:
            errors_map[p].append({"error": r["error"], "cnt": int(r["cnt"])})

    # 合并数据
    collectors = []
    all_producers = set(collect_map.keys()) | set(qc_map.keys()) | set(sampling_map.keys())

    for producer in all_producers:
        c = collect_map.get(producer, {})
        q = qc_map.get(producer, {})
        s = sampling_map.get(producer, {})

        qc_passed = int(q.get("qc_passed") or 0)
        qc_total = int(q.get("qc_total") or 0)
        sampling_passed = int(s.get("sampling_passed") or 0)
        sampling_total = int(s.get("sampling_total") or 0)

        collectors.append({
            "name": producer,
            "group": c.get("grp") or "未知",
            "total_cases": int(c.get("total_cases") or 0),
            "qc_pass_rate": round(qc_passed / qc_total * 100, 1) if qc_total > 0 else 0,
            "sampling_pass_rate": round(sampling_passed / sampling_total * 100, 1) if sampling_total > 0 else 0,
            "total_qc_hours": round(float(q.get("qc_hours") or 0), 1),
            "total_sampling_hours": round(float(s.get("sampling_hours") or 0), 1),
            "avg_daily_hours": 0,
            "top5_errors": errors_map.get(producer, []),
        })

    collectors.sort(key=lambda x: -x["total_cases"])
    return jsonify({"collectors": collectors})


# ═══════════════════════════════════════════════════════════════
#  全盘数据（production_list.xlsx + all_data.py + analysis_all.py）
# ═══════════════════════════════════════════════════════════════

_SCHED_DIR = os.path.dirname(os.path.abspath(__file__))
_REJECTED_PATH = os.path.join(_SCHED_DIR, "rejected_projects.json")

def _load_rejected():
    """返回用户已手动跳过的项目 ID 集合"""
    if os.path.exists(_REJECTED_PATH):
        try:
            with open(_REJECTED_PATH, "r", encoding="utf-8") as f:
                return set(json.load(f))
        except Exception:
            pass
    return set()

def _save_rejected(id_set):
    with open(_REJECTED_PATH, "w", encoding="utf-8") as f:
        json.dump(list(id_set), f, ensure_ascii=False)

# production_list.xlsx 路径（优先同目录，次选桌面运营日报）
_PROD_LIST_PATHS = [
    os.path.join(_SCHED_DIR, "production_list.xlsx"),
    os.path.expanduser("~/Desktop/运营日报/production_list.xlsx"),
]

def _prod_xlsx_path():
    for p in _PROD_LIST_PATHS:
        if os.path.exists(p):
            return p
    return None

def _load_production_list():
    """返回 [{id, name, form, scheme, region, label_ver}]"""
    p = _prod_xlsx_path()
    if not p:
        return []
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    headers = [str(h).strip() if h else "" for h in rows[0]]
    idx = {h: i for i, h in enumerate(headers)}
    projects = []
    for row in rows[1:]:
        i_id = idx.get("项目id", 1)
        pid = str(row[i_id]).strip() if row[i_id] else ""
        if not pid or not UUID_RE.match(pid):
            continue
        name = str(row[idx.get("项目", 0)] or "").strip()
        projects.append({
            "id":        pid,
            "name":      name,
            "form":      str(row[idx.get("采集形式", 2)] or "").strip(),
            "scheme":    str(row[idx.get("采集方案", 3)] or "").strip(),
            "region":    str(row[idx.get("地域", 4)] or "").strip(),
            "label_ver": str(row[idx.get("标注版本", 5)] or "").strip(),
        })
    return projects


@app.route("/api/overview/new-projects")
def overview_new_projects():
    """返回在 MySQL 里有数据但不在 production_list.xlsx 且未被拒绝的项目"""
    existing_ids = {p["id"] for p in _load_production_list()}
    rejected_ids = _load_rejected()

    conn = mysql()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT project_id, COUNT(*) AS case_count
                FROM human_cases
                WHERE deleted_at IS NULL AND project_id IS NOT NULL
                GROUP BY project_id
                HAVING COUNT(*) >= 5
            """)
            rows = cur.fetchall()
    finally:
        conn.close()

    new_ids = [r["project_id"] for r in rows
               if r["project_id"] and UUID_RE.match(str(r["project_id"]))
               and r["project_id"] not in existing_ids
               and r["project_id"] not in rejected_ids]
    if not new_ids:
        return jsonify([])

    # 从 Clickhouse 取项目名
    in_clause = ",".join(f"'{pid}'" for pid in new_ids[:300])
    try:
        c = ck()
        name_rows = c.execute(
            f"SELECT toString(uuid), name FROM project WHERE toString(uuid) IN ({in_clause})"
        )
        names = {r[0]: r[1] for r in name_rows}
    except Exception:
        names = {}

    case_map = {r["project_id"]: r["case_count"] for r in rows}
    result = [{"id": pid, "name": names.get(pid, ""), "case_count": case_map.get(pid, 0)}
              for pid in new_ids[:300]]
    result.sort(key=lambda x: -x["case_count"])
    return jsonify(result)


@app.route("/api/overview/add-projects", methods=["POST"])
def overview_add_projects():
    """将新项目追加到 production_list.xlsx"""
    projects = request.json  # [{id, name, form, scheme, region, label_ver}]
    if not projects:
        return jsonify({"added": 0})
    xlsx_path = _prod_xlsx_path()
    if not xlsx_path:
        return jsonify({"error": "production_list.xlsx 未找到"}), 404
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    for p in projects:
        ws.append([
            p.get("name", ""),
            p.get("id", ""),
            p.get("form", ""),
            p.get("scheme", ""),
            p.get("region", ""),
            p.get("label_ver", ""),
        ])
    wb.save(xlsx_path)
    return jsonify({"added": len(projects)})


@app.route("/api/overview/reject-projects", methods=["POST"])
def overview_reject_projects():
    """将项目 ID 加入永久拒绝列表（下次 new-projects 不再显示）"""
    ids = request.json  # list of id strings
    if not ids:
        return jsonify({"rejected": 0})
    rejected = _load_rejected()
    before = len(rejected)
    rejected.update(str(i) for i in ids if i)
    _save_rejected(rejected)
    return jsonify({"rejected": len(rejected) - before})


@app.route("/api/overview/run")
def overview_run():
    """流式执行 all_data.py → analysis_all.py，SSE 输出进度"""
    all_data_script   = os.path.join(_SCHED_DIR, "all_data.py")
    analysis_script   = os.path.join(_SCHED_DIR, "analysis_all.py")

    def generate():
        for script in [all_data_script, analysis_script]:
            name = os.path.basename(script)
            if not os.path.exists(script):
                yield f"data: ERROR: 找不到脚本 {name}\n\n"
                yield "data: DONE:error\n\n"
                return
            yield f"data: ▶ 运行 {name} ...\n\n"
            proc = subprocess.Popen(
                ["python3", script],
                cwd=_SCHED_DIR,
                stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                text=True, bufsize=1,
            )
            for line in proc.stdout:
                line = line.rstrip()
                if line:
                    yield f"data: {line}\n\n"
            proc.wait()
            if proc.returncode != 0:
                yield f"data: ERROR: {name} 退出码 {proc.returncode}\n\n"
                yield "data: DONE:error\n\n"
                return
            yield f"data: ✓ {name} 完成\n\n"
        yield "data: DONE:ok\n\n"

    return Response(generate(), content_type="text/event-stream",
                    headers={"X-Accel-Buffering": "no", "Cache-Control": "no-cache"})


@app.route("/api/overview/report")
def overview_report():
    """返回最新的 analysis_*.html 报告"""
    html_files = glob.glob(os.path.join(_SCHED_DIR, "analysis_*.html"))
    if not html_files:
        return "暂无报告，请先拉取数据", 404
    latest = max(html_files, key=os.path.getmtime)
    return send_from_directory(_SCHED_DIR, os.path.basename(latest), mimetype="text/html")


@app.route("/api/overview/status")
def overview_status():
    """返回最新报告文件名和时间"""
    html_files = glob.glob(os.path.join(_SCHED_DIR, "analysis_*.html"))
    if not html_files:
        return jsonify({"has_report": False})
    latest = max(html_files, key=os.path.getmtime)
    return jsonify({
        "has_report": True,
        "filename": os.path.basename(latest),
        "mtime": os.path.getmtime(latest),
    })


# ═══════════════════════════════════════════════════════════════
#  全盘每日采集完成量趋势
# ═══════════════════════════════════════════════════════════════
def _compute_trend():
    """实际执行采集趋势查询（90天），结果写入文件缓存"""
    days = 90
    end   = date.today()
    start = end - timedelta(days=days - 1)
    projects = _load_production_list()
    if not projects:
        return []
    project_ids = [p["id"] for p in projects]
    conn = mysql()
    try:
        with conn.cursor() as cur:
            ph = _ph(project_ids)
            cur.execute(f"""
                SELECT DATE(hcn.node_updated_at) AS day,
                       COUNT(*) AS cases,
                       SUM(IFNULL(hc.video_seconds, 0)) / 3600.0 AS hours
                FROM human_cases hc
                JOIN human_case_nodes hcn ON hcn.human_case_id = hc.id
                WHERE hc.project_id IN ({ph})
                  AND hcn.node_name = 'human_case_produce_complete'
                  AND hcn.node_status = 3
                  AND DATE(hcn.node_updated_at) BETWEEN %s AND %s
                  AND hc.deleted_at IS NULL
                GROUP BY DATE(hcn.node_updated_at)
                ORDER BY day
            """, project_ids + [str(start), str(end)])
            collect_rows = cur.fetchall()
            cur.execute(f"""
                SELECT DATE(q.node_updated_at) AS day,
                       COUNT(*) AS qc_cases,
                       SUM(IFNULL(hc.video_seconds, 0)) / 3600.0 AS qc_hours
                FROM human_cases hc
                JOIN (
                    SELECT hcn.human_case_id, hcn.node_updated_at
                    FROM human_case_nodes hcn
                    WHERE hcn.node_name = 'human_case_sampling'
                      AND hcn.node_status = 3
                      AND DATE(hcn.node_updated_at) BETWEEN %s AND %s
                    UNION ALL
                    SELECT hcn.human_case_id, hcn.node_updated_at
                    FROM human_case_nodes hcn
                    WHERE hcn.node_name = 'human_case_inspect'
                      AND hcn.node_status = 3
                      AND DATE(hcn.node_updated_at) BETWEEN %s AND %s
                      AND hcn.human_case_id NOT IN (
                          SELECT DISTINCT human_case_id FROM human_case_nodes
                          WHERE node_name = 'human_case_sampling'
                      )
                ) q ON q.human_case_id = hc.id
                WHERE hc.project_id IN ({ph})
                  AND hc.deleted_at IS NULL
                GROUP BY day
                ORDER BY day
            """, [str(start), str(end), str(start), str(end)] + project_ids)
            qc_rows = cur.fetchall()
    finally:
        conn.close()
    collect_map = {
        str(r["day"]): {"cases": int(r["cases"]), "hours": round(float(r["hours"] or 0), 2)}
        for r in collect_rows
    }
    qc_map = {
        str(r["day"]): {"qc_cases": int(r["qc_cases"]), "qc_hours": round(float(r["qc_hours"] or 0), 2)}
        for r in qc_rows
    }
    result = []
    d = start
    while d <= end:
        ds = str(d)
        info = collect_map.get(ds, {"cases": 0, "hours": 0.0})
        qc   = qc_map.get(ds, {"qc_cases": 0, "qc_hours": 0.0})
        result.append({"date": ds, "cases": info["cases"], "hours": info["hours"],
                        "qc_cases": qc["qc_cases"], "qc_hours": qc["qc_hours"]})
        d += timedelta(days=1)
    _trend_cache_set(result)
    return result


@app.route("/api/overview/daily-collect")
def overview_daily_collect():
    """全盘每日采集完成量（基于 production_list.xlsx 的项目）
    ?force=1 强制刷新；默认返回文件缓存"""
    force = request.args.get("force") == "1"
    if not force:
        cached = _trend_cache_get()
        if cached and cached.get("data"):
            return jsonify({"data": cached["data"], "updated_at": cached.get("updated_at", "")})
    result = _compute_trend()
    updated_at = time.strftime("%Y-%m-%d %H:%M")
    return jsonify({"data": result, "updated_at": updated_at})


# ═══════════════════════════════════════════════════════════════
#  API Key 鉴权（供 OpenClaw 等外部工具调用）
# ═══════════════════════════════════════════════════════════════
def _check_api_key():
    """验证 Authorization: Bearer <key>，返回 True 表示通过"""
    if not API_KEY:
        return True
    auth = request.headers.get("Authorization", "")
    return auth == f"Bearer {API_KEY}"


# ═══════════════════════════════════════════════════════════════
#  排期数据持久化
# ═══════════════════════════════════════════════════════════════
_SCHEDULES_PATH = os.path.join(_SCHED_DIR, "schedules.json")

def _load_schedules():
    if os.path.exists(_SCHEDULES_PATH):
        try:
            with open(_SCHEDULES_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {"schedules": []}

def _save_schedules(data):
    with open(_SCHEDULES_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def _get_schedule(sid):
    data = _load_schedules()
    for s in data["schedules"]:
        if s["id"] == sid:
            return s, data
    return None, data


# ═══════════════════════════════════════════════════════════════
#  排期 CRUD（浏览器 session 鉴权）
# ═══════════════════════════════════════════════════════════════
@app.route("/api/schedules", methods=["GET"])
def list_schedules():
    data = _load_schedules()
    # 只返回摘要，不返回全量 days 数据
    result = []
    for s in data["schedules"]:
        cur_v = next((v for v in s["versions"] if v["v"] == s["current_version"]), None)
        result.append({
            "id":              s["id"],
            "name":            s["name"],
            "project_ids":     s.get("project_ids", []),
            "project_names":   s.get("project_names", []),
            "current_version": s["current_version"],
            "version_count":   len(s["versions"]),
            "target_h":        cur_v["target_h"] if cur_v else 0,
            "start_date":      cur_v["start_date"] if cur_v else "",
            "saved_at":        cur_v["saved_at"] if cur_v else "",
        })
    return jsonify(result)


@app.route("/api/schedules", methods=["POST"])
def create_schedule():
    body = request.json or {}
    sid  = str(_uuid.uuid4())
    now  = __import__("datetime").datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
    schedule = {
        "id":              sid,
        "name":            body.get("name", "新排期"),
        "project_ids":     body.get("project_ids", []),
        "project_names":   body.get("project_names", []),
        "current_version": 1,
        "versions": [{
            "v":          1,
            "saved_at":   now,
            "target_h":   body.get("target_h", 0),
            "start_date": body.get("start_date", ""),
            "params":     body.get("params", {}),
            "days":       body.get("days", []),
        }],
    }
    data = _load_schedules()
    data["schedules"].append(schedule)
    _save_schedules(data)
    return jsonify(schedule), 201


@app.route("/api/schedules/<sid>", methods=["GET"])
def get_schedule(sid):
    s, _ = _get_schedule(sid)
    if not s:
        return jsonify({"error": "not found"}), 404
    return jsonify(s)


@app.route("/api/schedules/<sid>", methods=["PUT"])
def update_schedule(sid):
    """直接覆盖当前版本（单版本模式）"""
    s, data = _get_schedule(sid)
    if not s:
        return jsonify({"error": "not found"}), 404
    body = request.json or {}
    now  = __import__("datetime").datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
    version = {
        "v":          1,
        "saved_at":   now,
        "target_h":   body.get("target_h", 0),
        "start_date": body.get("start_date", ""),
        "params":     body.get("params", {}),
        "days":       body.get("days", []),
    }
    s["versions"] = [version]
    s["current_version"] = 1
    if "name"          in body: s["name"]          = body["name"]
    if "project_ids"   in body: s["project_ids"]   = body["project_ids"]
    if "project_names" in body: s["project_names"] = body["project_names"]
    _save_schedules(data)
    return jsonify(s)


@app.route("/api/schedules/<sid>", methods=["DELETE"])
def delete_schedule(sid):
    s, data = _get_schedule(sid)
    if not s:
        return jsonify({"error": "not found"}), 404
    data["schedules"] = [x for x in data["schedules"] if x["id"] != sid]
    _save_schedules(data)
    return jsonify({"deleted": sid})


# ═══════════════════════════════════════════════════════════════
#  排期 × 实际对比（OpenClaw API Key 鉴权 + 浏览器均可访问）
# ═══════════════════════════════════════════════════════════════
def _require_auth():
    """浏览器 session 或 API Key 任一通过即可"""
    if _check_api_key():
        return None
    if not KANBAN_PASSWORD or session.get("authed"):
        return None
    return jsonify({"error": "unauthorized"}), 401


@app.route("/api/schedules/<sid>/actual")
def schedule_actual(sid):
    err = _require_auth()
    if err: return err

    s, _ = _get_schedule(sid)
    if not s:
        return jsonify({"error": "not found"}), 404

    project_ids = safe_uuids(s.get("project_ids", []))
    if not project_ids:
        return jsonify({"error": "no project_ids bound"}), 400

    # 取当前版本参数
    cur_v = next((v for v in s["versions"] if v["v"] == s["current_version"]), None)
    if not cur_v:
        return jsonify({"error": "no version"}), 400

    start = cur_v["start_date"]
    end   = str(date.today())

    conn = mysql()
    try:
        with conn.cursor() as cur:
            collected = _query_daily(cur, project_ids, "human_case_produce_complete", 3, start, end)
            qc_passed = _query_daily(cur, project_ids, "human_case_inspect",          3, start, end)
            labeled   = _query_daily(cur, project_ids, "labeling_complete",           3, start, end)
    finally:
        conn.close()

    # 填满日期区间
    s_d, e_d = date.fromisoformat(start), date.fromisoformat(end)
    rows, d = [], s_d
    while d <= e_d:
        ds = str(d)
        rows.append({
            "date":      ds,
            "collected": round(collected.get(ds, 0), 2),
            "qc_passed": round(qc_passed.get(ds, 0), 2),
            "labeled":   round(labeled.get(ds, 0), 2),
        })
        d += timedelta(days=1)
    return jsonify(rows)


@app.route("/api/schedules/<sid>/compare")
def schedule_compare(sid):
    """计划 vs 实际对比，含偏差和预测"""
    err = _require_auth()
    if err: return err

    s, _ = _get_schedule(sid)
    if not s:
        return jsonify({"error": "not found"}), 404

    project_ids = safe_uuids(s.get("project_ids", []))
    # 允许前端通过 query param 传入（未保存项目关联时的临时使用）
    ids_raw = request.args.get("project_ids", "").strip()
    if ids_raw:
        project_ids = safe_uuids([p.strip() for p in ids_raw.split(",") if p.strip()])
    cur_v = next((v for v in s["versions"] if v["v"] == s["current_version"]), None)
    if not cur_v:
        return jsonify({"error": "no version"}), 400

    params      = cur_v.get("params", {})
    qc_rate     = float(params.get("qc_rate", 85)) / 100
    pack_rate   = float(params.get("pack_rate", 90)) / 100
    target_h    = float(cur_v["target_h"])

    # 计划 days 索引
    plan_by_date = {d["date"]: d for d in cur_v.get("days", [])}

    # 实际查询范围：取排期第一天 ~ 今天
    today = str(date.today())
    if plan_by_date:
        query_start = min(plan_by_date.keys())
    else:
        query_start = (date.today() - timedelta(days=90)).isoformat()

    # 实际产出
    actual_collected, actual_labeled, actual_packed = {}, {}, {}
    if project_ids:
        # 判断是否是Mango项目（检查project_names）
        project_names = s.get("project_names", [])
        is_mango = any("mango" in str(name).lower() for name in project_names)

        conn = mysql()
        try:
            with conn.cursor() as cur:
                if is_mango:
                    actual_collected = _query_daily_mango_dedup(cur, project_ids, "human_case_produce_complete", 3, query_start, today)
                    actual_labeled   = _query_daily_mango_dedup(cur, project_ids, "labeling_complete",           3, query_start, today)
                    actual_packed    = _query_daily_mango_dedup(cur, project_ids, "complete_job",                3, query_start, today)
                else:
                    actual_collected = _query_daily(cur, project_ids, "human_case_produce_complete", 3, query_start, today)
                    actual_labeled   = _query_daily(cur, project_ids, "labeling_complete",           3, query_start, today)
                    actual_packed    = _query_daily(cur, project_ids, "complete_job",                3, query_start, today)
        finally:
            conn.close()

    # 逐天对比（采集量 = 采集完成 × qc_rate，打包完成 = 标注完成 × pack_rate）
    all_dates = sorted(set(list(plan_by_date.keys()) + list(actual_collected.keys()) + list(actual_labeled.keys()) + list(actual_packed.keys())))
    cum_act_collect  = 0.0
    cum_act_label    = 0.0
    rows = []
    for i, ds in enumerate(all_dates):
        pd_ = plan_by_date.get(ds, {})
        plan_l = float(pd_.get("plan_label_h", 0))
        # 实际采集完成 × qc_rate
        act_c  = actual_collected.get(ds, 0.0) * qc_rate
        act_l  = actual_labeled.get(ds, 0.0)
        # 打包完成 = 标注完成 × pack_rate
        act_p  = act_l * pack_rate
        cum_act_collect  += act_c
        cum_act_label    += act_l
        rows.append({
            "date":             ds,
            "actual_collect":   round(act_c,  2),
            "cum_act_collect":  round(cum_act_collect,  2),
            "plan_label":       round(plan_l, 2),
            "actual_label":     round(act_l,  2),
            "actual_packed":    round(act_p,  2),
        })

    # 累计打包完成
    cum_act_packed = sum(actual_packed.get(ds, 0.0) for ds in all_dates if ds <= today)

    # 过去各天的计划累计（截止今天）
    cum_plan_collect_to_today = sum(
        float(plan_by_date.get(ds, {}).get("plan_collect_h", 0)) * qc_rate
        for ds in all_dates if ds <= today
    )
    cum_plan_label_to_today = sum(
        float(plan_by_date.get(ds, {}).get("plan_label_h", 0))
        for ds in all_dates if ds <= today
    )

    # 后续排期可交付（明天开始，计划标注×打包留存率）
    future_deliverable = sum(
        float(plan_by_date.get(ds, {}).get("plan_label_h", 0)) * pack_rate
        for ds in all_dates if ds > today
    )

    return jsonify({
        "schedule_id":              sid,
        "schedule_name":            s["name"],
        "target_h":                 target_h,
        "current_version":          s["current_version"],
        "pack_rate":                pack_rate,
        "cum_act_collect":          round(cum_act_collect,          2),
        "cum_act_label":            round(cum_act_label,            2),
        "cum_act_packed":           round(cum_act_packed,           2),
        "cum_plan_collect_to_today":round(cum_plan_collect_to_today,2),
        "cum_plan_label_to_today":  round(cum_plan_label_to_today,  2),
        "stock_deliverable":        round(cum_act_packed,           2),
        "future_deliverable":       round(future_deliverable,       2),
        "collect_gap":              round(cum_plan_collect_to_today - cum_act_collect, 2),
        "label_gap":                round(cum_plan_label_to_today   - cum_act_label,   2),
        "rows":                     rows,
    })


# ─── Delivery Tracker 整合 ───────────────────────────────────────────────────

DELIVERY_DIR       = "/root/delivery-tracker"
DELIVERY_SCRIPT    = f"{DELIVERY_DIR}/scripts/query.py"
DELIVERY_CONFIG    = f"{DELIVERY_DIR}/projects.json"
DELIVERY_CACHE_DIR = "/root/schedule/delivery_cache"   # 文件缓存目录，所有 worker 共享
DELIVERY_CACHE_TTL = 7200   # 2小时
DELIVERY_LOCK_FILE = "/tmp/delivery_refreshing.lock"
os.makedirs(DELIVERY_CACHE_DIR, exist_ok=True)


def _delivery_cache_path(name):
    safe = name.replace("/", "_").replace(" ", "_")
    return os.path.join(DELIVERY_CACHE_DIR, f"{safe}.json")


def _delivery_file_get(name, ttl=None):
    """从文件缓存读取，超 TTL 返回 (None, None)"""
    path = _delivery_cache_path(name)
    effective_ttl = ttl if ttl is not None else DELIVERY_CACHE_TTL
    try:
        with open(path, encoding="utf-8") as f:
            obj = json.load(f)
        if time.time() - obj.get("ts", 0) > effective_ttl:
            return None, None
        return obj.get("data"), obj.get("updated_at")
    except Exception:
        return None, None


def _delivery_file_is_fresh(name, ttl):
    """检查缓存是否在 ttl 秒内仍有效"""
    path = _delivery_cache_path(name)
    try:
        with open(path, encoding="utf-8") as f:
            obj = json.load(f)
        return time.time() - obj.get("ts", 0) <= ttl
    except Exception:
        return False


def _delivery_file_set(name, data):
    path = _delivery_cache_path(name)
    obj = {"data": data, "updated_at": time.strftime("%Y-%m-%d %H:%M:%S"), "ts": time.time()}
    with open(path, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False)


def _delivery_is_refreshing():
    try:
        if not os.path.exists(DELIVERY_LOCK_FILE):
            return False
        # 锁文件超过 30 分钟认为已过期
        return time.time() - os.path.getmtime(DELIVERY_LOCK_FILE) < 1800
    except Exception:
        return False


# ─── 采集趋势 文件缓存（每天 9:00 自动刷新，按需手动刷新） ────
TREND_CACHE_FILE = "/root/schedule/trend_cache.json"

def _trend_cache_get():
    try:
        with open(TREND_CACHE_FILE, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None

def _trend_cache_set(data):
    obj = {"data": data, "updated_at": time.strftime("%Y-%m-%d %H:%M"), "ts": time.time()}
    with open(TREND_CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False)


# ─── 采集详情 文件缓存（每小时自动刷新） ──────────────────────
COLLECTORS_CACHE_DIR = "/root/schedule/collectors_cache"
COLLECTORS_CACHE_TTL = 3600
os.makedirs(COLLECTORS_CACHE_DIR, exist_ok=True)

def _collectors_cache_path(date_str):
    return os.path.join(COLLECTORS_CACHE_DIR, f"{date_str}.json")

def _collectors_file_get(date_str):
    try:
        with open(_collectors_cache_path(date_str), encoding="utf-8") as f:
            obj = json.load(f)
        if time.time() - obj.get("ts", 0) > COLLECTORS_CACHE_TTL:
            return None
        return obj
    except Exception:
        return None

def _collectors_file_set(date_str, data):
    obj = {**data, "updated_at": time.strftime("%H:%M"), "ts": time.time()}
    with open(_collectors_cache_path(date_str), "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False)


def _load_delivery_projects():
    """读取 projects.json，返回所有 active 项目列表"""
    try:
        with open(DELIVERY_CONFIG, encoding="utf-8") as f:
            config = json.load(f)
        return [p for p in config.get("projects", []) if p.get("status") == "active"]
    except Exception:
        return []


def _run_single_delivery(proj_name):
    """运行 query.py --project <name> --json --no-save，返回解析后的 dict"""
    result = subprocess.run(
        ["python3", DELIVERY_SCRIPT, "--project", proj_name, "--json", "--no-save"],
        capture_output=True, text=True, timeout=7200, close_fds=True,
    )
    if result.returncode != 0:
        raise RuntimeError(result.stderr[:500])
    return json.loads(result.stdout)


def _delivery_refresh_all(force=False):
    """后台刷新所有 active 项目，写入文件缓存。
    force=False 时跳过缓存仍有效的项目（按各项目 refresh_interval）。"""
    try:
        open(DELIVERY_LOCK_FILE, "w").close()
        projects = _load_delivery_projects()
        for proj in projects:
            name = proj["name"]
            ttl = int(proj.get("refresh_interval", DELIVERY_CACHE_TTL))
            if not force and _delivery_file_is_fresh(name, ttl):
                print(f"[delivery] {name} 缓存有效，跳过", flush=True)
                continue
            try:
                data = _run_single_delivery(name)
                _delivery_file_set(name, data)
                print(f"[delivery] {name} 刷新完成", flush=True)
            except Exception as e:
                print(f"[delivery] {name} 查询失败: {e}", flush=True)
    finally:
        try:
            os.remove(DELIVERY_LOCK_FILE)
        except Exception:
            pass


def _delivery_refresh_loop():
    """守护线程：启动延迟 60 秒后开始，每小时检查一次，按各项目 refresh_interval 决定是否刷新"""
    time.sleep(60)
    while True:
        _delivery_refresh_all(force=False)
        time.sleep(3600)


@app.route("/api/delivery/projects")
def api_delivery_projects():
    if not session.get("authed"):
        return jsonify({"error": "未登录"}), 401
    projects = _load_delivery_projects()
    return jsonify({"projects": projects})


@app.route("/api/delivery/data")
def api_delivery_data():
    if not session.get("authed"):
        return jsonify({"error": "未登录"}), 401
    projects = _load_delivery_projects()
    result = {}
    for proj in projects:
        name = proj["name"]
        ttl = int(proj.get("refresh_interval", DELIVERY_CACHE_TTL))
        cached, updated_at = _delivery_file_get(name, ttl=ttl)
        result[name] = {
            "config": proj,
            "data": cached,
            "updated_at": updated_at,
        }
    return jsonify({"projects": result, "refreshing": _delivery_is_refreshing()})


@app.route("/api/delivery/data/<name>")
def api_delivery_data_single(name):
    if not session.get("authed"):
        return jsonify({"error": "未登录"}), 401
    cached, updated_at = _delivery_file_get(name)
    return jsonify({"data": cached, "updated_at": updated_at, "refreshing": _delivery_is_refreshing()})


@app.route("/api/delivery/refresh", methods=["POST"])
def api_delivery_refresh():
    if not session.get("authed"):
        return jsonify({"error": "未登录"}), 401
    if _delivery_is_refreshing():
        return jsonify({"status": "already_running"})
    t = threading.Thread(target=_delivery_refresh_all, args=(True,), daemon=True)
    t.start()
    return jsonify({"status": "started"})


# 启动时只让 PID 最小的 worker 做首次刷新，避免多 worker 重复执行
_t = threading.Thread(target=_delivery_refresh_loop, daemon=True)
_t.start()


# ─── 采集趋势：每天 9:00 自动刷新 ────────────────────────────
def _trend_bg_loop():
    last_day = None
    time.sleep(30)  # 等待服务启动
    while True:
        now = datetime.now()
        today = now.date()
        if now.hour >= 9 and last_day != today:
            try:
                _compute_trend()
                last_day = today
            except Exception:
                pass
        time.sleep(60)

threading.Thread(target=_trend_bg_loop, daemon=True).start()


# ═══════════════════════════════════════════════════════════════
#  DM 看板
# ═══════════════════════════════════════════════════════════════

_DM_ENV_ORDER = ["家居", "商超", "酒店", "物流"]
_DM_ENV_ALIASES = {
    "home": "家居",
    "supermarket": "商超", "retail": "商超",
    "hospitality": "酒店", "hotel": "酒店", "hospital": "酒店",
    "logistics": "物流", "distribution_center": "物流",
}

def _dm_parse_env(env_type_name, environment_num, env_num):
    for raw in (env_type_name, environment_num, env_num):
        v = str(raw or "").strip().strip("\"'")
        if not v or v in ("None", "null", "nan"):
            continue
        key = v
        m = re.match(r"^([a-z_]+?)_x_", v) or re.match(r"^([a-z_]+?)_\d", v)
        if m:
            key = m.group(1)
        mapped = _DM_ENV_ALIASES.get(key)
        if mapped:
            return mapped
        # 直接是中文环境名
        if key in _DM_ENV_ORDER:
            return key
    return ""

def _dm_env_sort(name):
    if name in _DM_ENV_ORDER:
        return (0, _DM_ENV_ORDER.index(name))
    return (1, name)

def _dm_build_summary(rows):
    grouped = defaultdict(lambda: {"projects": set(), "scenes": set(), "tasks": set()})
    for r in rows:
        env = _dm_parse_env(r.get("env_type_name"), r.get("environment_num"), r.get("env_num"))
        if not env:
            continue
        task = (r.get("task_name") or "").strip()
        if not task:
            continue
        grouped[env]["projects"].add(r.get("project_name") or "")
        if r.get("scene_num"):
            grouped[env]["scenes"].add(r["scene_num"])
        grouped[env]["tasks"].add(task)
    result = []
    for env in sorted(grouped.keys(), key=_dm_env_sort):
        g = grouped[env]
        result.append({
            "env": env,
            "count": len(g["tasks"]),
            "projects": sorted(g["projects"]),
            "scenes": sorted(g["scenes"]),
            "tasks": sorted(g["tasks"]),
        })
    return result

@app.route("/dm")
def dm_page():
    return send_from_directory(".", "dm.html")

@app.route("/api/dm/summary")
def dm_summary():
    cached = _cache_get("dm_summary")
    if cached:
        return jsonify(cached)

    rows = []

    # 旧库
    try:
        conn = mysql()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT p.name AS project_name,
                       hc.task_name,
                       JSON_UNQUOTE(JSON_EXTRACT(hc.metadata, '$.env_type_name'))   AS env_type_name,
                       JSON_UNQUOTE(JSON_EXTRACT(hc.metadata, '$.environment_num')) AS environment_num,
                       JSON_UNQUOTE(JSON_EXTRACT(hc.metadata, '$.env_num'))         AS env_num,
                       JSON_UNQUOTE(JSON_EXTRACT(hc.metadata, '$.scene_num'))       AS scene_num
                FROM human_cases hc
                JOIN projects p ON p.id = hc.project_id
                WHERE hc.project_id IN (
                    SELECT DISTINCT p2.id FROM project_tag_project ptp
                    JOIN project_tag t ON t.id = ptp.tag_id
                    JOIN projects p2 ON p2.id = ptp.project_id
                    WHERE ptp.deleted_at IS NULL AND p2.is_deleted = 0
                      AND LOWER(t.tag_name) = 'dm'
                )
                AND hc.id IN (
                    SELECT hcn.human_case_id FROM human_case_nodes hcn
                    INNER JOIN (
                        SELECT human_case_id, MAX(id) AS max_id
                        FROM human_case_nodes
                        WHERE node_name = 'complete_job' GROUP BY human_case_id
                    ) latest ON latest.max_id = hcn.id
                    WHERE hcn.node_status = 3
                )
            """)
            rows += list(cur.fetchall())
        conn.close()
    except Exception as e:
        app.logger.warning(f"DM old db error: {e}")

    # 新库
    try:
        conn2 = mysql_new()
        with conn2.cursor() as cur:
            cur.execute("""
                SELECT p.name AS project_name,
                       COALESCE(
                           NULLIF(JSON_UNQUOTE(JSON_EXTRACT(ctx.value, '$.task_name')), ''),
                           NULLIF(slog.task_name, ''),
                           CONCAT('task_id:', CAST(hc.task_id AS CHAR))
                       ) AS task_name,
                       JSON_UNQUOTE(JSON_EXTRACT(ctx.value, '$.env_type_name'))   AS env_type_name,
                       COALESCE(JSON_UNQUOTE(JSON_EXTRACT(ctx.value, '$.environment_num')),
                                JSON_UNQUOTE(JSON_EXTRACT(ctx.value, '$.env_num')), '') AS environment_num,
                       COALESCE(JSON_UNQUOTE(JSON_EXTRACT(ctx.value, '$.env_num')),
                                JSON_UNQUOTE(JSON_EXTRACT(ctx.value, '$.environment_num')), '') AS env_num,
                       COALESCE(JSON_UNQUOTE(JSON_EXTRACT(ctx.value, '$.scene_num')), '') AS scene_num
                FROM human_case hc
                JOIN project p ON p.uuid = hc.project_uuid
                LEFT JOIN (
                    SELECT t1.human_case_id, t1.value FROM human_case_tag t1
                    INNER JOIN (SELECT human_case_id, MAX(id) AS max_id FROM human_case_tag
                                WHERE type='context_tags' GROUP BY human_case_id) tc ON tc.max_id = t1.id
                ) ctx ON ctx.human_case_id = hc.id
                LEFT JOIN (
                    SELECT l1.human_case_uuid, l1.task_name FROM human_case_sample_log l1
                    INNER JOIN (SELECT human_case_uuid, MAX(id) AS max_id FROM human_case_sample_log
                                GROUP BY human_case_uuid) ls ON ls.max_id = l1.id
                ) slog ON slog.human_case_uuid = hc.uuid
                WHERE p.deleted_at IS NULL AND hc.deleted_at IS NULL
                  AND p.id IN (
                      SELECT DISTINCT p2.id FROM project p2
                      JOIN project_tag_project ptp ON ptp.project_id = p2.id AND ptp.deleted_at IS NULL
                      JOIN project_tag t ON t.id = ptp.tag_id
                      WHERE p2.deleted_at IS NULL AND LOWER(t.tag_name) = 'dm'
                  )
                  AND hc.id IN (
                      SELECT x.human_case_id FROM (
                          SELECT human_case_id, node_name, node_status,
                                 ROW_NUMBER() OVER (PARTITION BY human_case_id, node_name ORDER BY id DESC) AS rn
                          FROM human_case_node
                          WHERE node_name IN ('complete_job', 'delivery_packaging')
                      ) x WHERE x.rn = 1 AND x.node_status = 3
                  )
            """)
            rows += list(cur.fetchall())
        conn2.close()
    except Exception as e:
        app.logger.warning(f"DM new db error: {e}")

    summary = _dm_build_summary(rows)
    total = sum(g["count"] for g in summary)
    result = {"summary": summary, "total": total, "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M")}
    _cache_set("dm_summary", result, 1800)  # 30 分钟缓存
    return jsonify(result)


# ─── 人效统计 ─────────────────────────────────────────────────────────────────
_EFFICIENCY_CACHE_PATH = os.path.join(os.path.dirname(__file__), "efficiency_cache.json")

def _load_efficiency_cache():
    if os.path.exists(_EFFICIENCY_CACHE_PATH):
        try:
            with open(_EFFICIENCY_CACHE_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def _save_efficiency_cache(data):
    with open(_EFFICIENCY_CACHE_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def _compute_efficiency(date_str):
    """计算指定日期各供应商腕部/PT 人效，返回 dict"""
    # ── 1. 获取腕部项目 ID 集合（旧库 + 新库）──────────────────
    old_wrist_ids = set()
    new_wrist_ids = set()
    try:
        conn = mysql()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT ptp.project_id FROM project_tag_project ptp
                JOIN project_tag pt ON pt.id = ptp.tag_id
                WHERE pt.tag_name = '腕部Wrist' AND ptp.deleted_at IS NULL
            """)
            old_wrist_ids = set(r["project_id"] for r in cur.fetchall())
        conn.close()
    except Exception:
        pass
    try:
        conn2 = mysql_new()
        with conn2.cursor() as cur:
            cur.execute("""
                SELECT p.uuid FROM project_tag_project ptp
                JOIN project_tag pt ON pt.id = ptp.tag_id
                JOIN project p ON p.id = ptp.project_id
                WHERE pt.tag_name = '腕部Wrist' AND ptp.deleted_at IS NULL
            """)
            new_wrist_ids = set(r["uuid"] for r in cur.fetchall())
        conn2.close()
    except Exception:
        pass
    wrist_ids = old_wrist_ids | new_wrist_ids
    _new_wrist_uuids_cache = new_wrist_ids  # 供后续新库 case 分类使用

    # ── 2. 查询旧库当日采集（含 project_id）──────────────────────
    # 每个采集员按 is_wrist 分组：upload_h, collect_qc_h, collect_inspected_h
    person_data = defaultdict(lambda: {
        "vendor": "未知",
        "total_collect_h": 0.0,   # 累计采集（判断老人用）
        "wrist_upload_h": 0.0, "wrist_collect_qc_h": 0.0, "wrist_collect_inspected_h": 0.0,
        "pt_upload_h": 0.0,    "pt_collect_qc_h": 0.0,    "pt_collect_inspected_h": 0.0,
    })

    # 旧库
    try:
        conn = mysql()
        with conn.cursor() as cur:
            # 当日 upload 完成，带 project_id
            cur.execute("""
                SELECT hc.producer,
                       hc.produced_by_group AS vendor,
                       hc.project_id,
                       IFNULL(hc.video_seconds, 0) / 3600.0 AS vsec_h
                FROM human_cases hc
                JOIN human_case_nodes hcn ON hcn.human_case_id = hc.id
                WHERE hcn.node_name = 'human_case_upload'
                  AND hcn.node_status = 3
                  AND DATE(hcn.node_updated_at) = %s
                  AND hc.deleted_at IS NULL
            """, [date_str])
            for r in cur.fetchall():
                p = r["producer"]; pid = r["project_id"] or ""
                person_data[p]["vendor"] = r["vendor"] or "未知"
                if pid in wrist_ids:
                    person_data[p]["wrist_upload_h"] += float(r["vsec_h"])
                else:
                    person_data[p]["pt_upload_h"] += float(r["vsec_h"])

            # 累计采集（判断老人）—— 复用 collectors 缓存，避免全表扫描
            _cumul = None
            for _try_key in [f"cumulative:{date_str}", "cumulative:"]:
                _cumul = _cache_get(_try_key) or _disk_cache_get(_try_key)
                if _cumul: break
            if _cumul and isinstance(_cumul, (list, tuple)) and len(_cumul) >= 1:
                _tc_map = _cumul[0]
                for _p, _h in _tc_map.items():
                    person_data[_p]["total_collect_h"] = max(person_data[_p]["total_collect_h"], float(_h))
            # 若缓存没有，所有人都当老人处理（分母不变，数据仍有意义）

            # 采集抽检人效（旧库：video_seconds，按 project_id 分腕部/PT）
            cur.execute("""
                SELECT hc.producer, hc.project_id,
                    SUM(CASE
                        WHEN COALESCE(ns.sampling_status3,0)=1 THEN IFNULL(hc.video_seconds,0)
                        WHEN COALESCE(ns.sampling_any,0)=1     THEN 0
                        WHEN COALESCE(ns.inspect_status3,0)=1  THEN IFNULL(hc.video_seconds,0)
                        ELSE 0 END) / 3600.0 AS qc_h,
                    SUM(CASE
                        WHEN COALESCE(ns.sampling_status34,0)=1 THEN IFNULL(hc.video_seconds,0)
                        WHEN COALESCE(ns.sampling_any,0)=1      THEN 0
                        WHEN COALESCE(ns.inspect_status34,0)=1  THEN IFNULL(hc.video_seconds,0)
                        ELSE 0 END) / 3600.0 AS inspected_h
                FROM human_cases hc
                JOIN human_case_nodes hcn ON hcn.human_case_id = hc.id
                LEFT JOIN (
                    SELECT hcn2.human_case_id,
                        MAX(CASE WHEN hcn2.node_name='human_case_sampling' AND hcn2.node_status=3    THEN 1 ELSE 0 END) AS sampling_status3,
                        MAX(CASE WHEN hcn2.node_name='human_case_sampling'                           THEN 1 ELSE 0 END) AS sampling_any,
                        MAX(CASE WHEN hcn2.node_name='human_case_inspect'  AND hcn2.node_status=3   THEN 1 ELSE 0 END) AS inspect_status3,
                        MAX(CASE WHEN hcn2.node_name='human_case_sampling' AND hcn2.node_status IN(3,4) THEN 1 ELSE 0 END) AS sampling_status34,
                        MAX(CASE WHEN hcn2.node_name='human_case_inspect'  AND hcn2.node_status IN(3,4) THEN 1 ELSE 0 END) AS inspect_status34
                    FROM human_case_nodes hcn2
                    WHERE hcn2.node_name IN ('human_case_sampling','human_case_inspect')
                    GROUP BY hcn2.human_case_id
                ) ns ON ns.human_case_id = hc.id
                WHERE hcn.node_name = 'human_case_upload'
                  AND hcn.node_status = 3
                  AND DATE(hcn.node_updated_at) = %s
                  AND hc.deleted_at IS NULL
                GROUP BY hc.producer, hc.project_id
            """, [date_str])
            for r in cur.fetchall():
                p = r["producer"]; pid = r["project_id"] or ""
                if pid in wrist_ids:
                    person_data[p]["wrist_collect_qc_h"]         += float(r["qc_h"] or 0)
                    person_data[p]["wrist_collect_inspected_h"]  += float(r["inspected_h"] or 0)
                else:
                    person_data[p]["pt_collect_qc_h"]            += float(r["qc_h"] or 0)
                    person_data[p]["pt_collect_inspected_h"]      += float(r["inspected_h"] or 0)
        conn.close()
    except Exception:
        import traceback; traceback.print_exc()

    # 新库
    _d_start = date_str + " 00:00:00"
    _d_end   = date_str + " 23:59:59"
    try:
        conn2 = mysql_new()
        with conn2.cursor() as cur:
            # 当日采集（新库正确结构）
            cur.execute("""
                SELECT
                    hc.producer AS producer,
                    ag.name AS vendor,
                    hc.project_uuid AS project_id,
                    IFNULL(CAST(JSON_EXTRACT(pt.value,'$.data_info.duration') AS DECIMAL(15,4)),0) / 3600.0 AS vsec_h
                FROM human_case hc
                JOIN human_case_node hcn ON hcn.human_case_id = hc.id
                LEFT JOIN human_case_tag pt ON pt.human_case_id = hc.id AND pt.type = 'produce_tags'
                LEFT JOIN auth.users au ON au.name = hc.producer
                LEFT JOIN auth.user_group aug ON aug.user_uuid = au.uuid
                LEFT JOIN auth.`groups` ag ON ag.id = aug.group_uuid AND ag.biz_type = 'produce'
                WHERE hcn.node_name='human_case_upload'
                  AND hcn.node_status=3
                  AND hcn.node_updated_at BETWEEN %s AND %s
            """, [_d_start, _d_end])
            for r in cur.fetchall():
                p = r["producer"]; pid = r["project_id"] or ""
                if not p: continue
                person_data[p]["vendor"] = r["vendor"] or person_data[p]["vendor"]
                if pid in _new_wrist_uuids_cache:
                    person_data[p]["wrist_upload_h"] += float(r["vsec_h"] or 0)
                else:
                    person_data[p]["pt_upload_h"] += float(r["vsec_h"] or 0)

            # 新库累计 upload 时长（用于判断老人 total_collect_h >= 10h）
            cur.execute("""
                SELECT hc.producer,
                       SUM(IFNULL(CAST(JSON_EXTRACT(pt.value,'$.data_info.duration') AS DECIMAL(15,4)),0))/3600.0 AS total_h
                FROM human_case hc
                JOIN human_case_node hcn ON hcn.human_case_id = hc.id
                LEFT JOIN human_case_tag pt ON pt.human_case_id = hc.id AND pt.type = 'produce_tags'
                WHERE hcn.node_name = 'human_case_upload'
                  AND hcn.node_status = 3
                GROUP BY hc.producer
            """)
            for r in cur.fetchall():
                p = r["producer"]
                if not p: continue
                person_data[p]["total_collect_h"] = max(
                    person_data[p]["total_collect_h"],
                    float(r["total_h"] or 0)
                )

            # 采集抽检（新库，按 project_uuid 分腕部/PT）
            cur.execute("""
                SELECT
                    hc.producer AS producer,
                    hc.project_uuid AS project_id,
                    SUM(CASE
                        WHEN COALESCE(ns.sampling_status3,0)=1 THEN IFNULL(CAST(JSON_EXTRACT(pt.value,'$.data_info.duration') AS DECIMAL(15,4)),0)
                        WHEN COALESCE(ns.sampling_any,0)=1 THEN 0
                        WHEN COALESCE(ns.inspect_status3,0)=1 THEN IFNULL(CAST(JSON_EXTRACT(pt.value,'$.data_info.duration') AS DECIMAL(15,4)),0)
                        ELSE 0 END)/3600.0 AS qc_h,
                    SUM(CASE
                        WHEN COALESCE(ns.sampling_status34,0)=1 THEN IFNULL(CAST(JSON_EXTRACT(pt.value,'$.data_info.duration') AS DECIMAL(15,4)),0)
                        WHEN COALESCE(ns.sampling_any,0)=1 THEN 0
                        WHEN COALESCE(ns.inspect_status34,0)=1 THEN IFNULL(CAST(JSON_EXTRACT(pt.value,'$.data_info.duration') AS DECIMAL(15,4)),0)
                        ELSE 0 END)/3600.0 AS inspected_h
                FROM human_case hc
                JOIN human_case_node hcn ON hcn.human_case_id = hc.id
                LEFT JOIN human_case_tag pt ON pt.human_case_id = hc.id AND pt.type = 'produce_tags'
                LEFT JOIN (
                    SELECT hcn_qc.human_case_id,
                        MAX(CASE WHEN hcn_qc.node_name='human_case_sampling' AND hcn_qc.node_status=3    THEN 1 ELSE 0 END) AS sampling_status3,
                        MAX(CASE WHEN hcn_qc.node_name='human_case_sampling'                             THEN 1 ELSE 0 END) AS sampling_any,
                        MAX(CASE WHEN hcn_qc.node_name='human_case_inspect'  AND hcn_qc.node_status=3   THEN 1 ELSE 0 END) AS inspect_status3,
                        MAX(CASE WHEN hcn_qc.node_name='human_case_sampling' AND hcn_qc.node_status IN(3,4) THEN 1 ELSE 0 END) AS sampling_status34,
                        MAX(CASE WHEN hcn_qc.node_name='human_case_inspect'  AND hcn_qc.node_status IN(3,4) THEN 1 ELSE 0 END) AS inspect_status34
                    FROM human_case_node hcn_qc
                    JOIN human_case_node hcn_pc ON hcn_pc.human_case_id = hcn_qc.human_case_id
                        AND hcn_pc.node_name = 'human_case_upload'
                        AND hcn_pc.node_status = 3
                        AND hcn_pc.node_updated_at BETWEEN %s AND %s
                    WHERE hcn_qc.node_name IN ('human_case_sampling','human_case_inspect')
                    GROUP BY hcn_qc.human_case_id
                ) ns ON ns.human_case_id = hc.id
                WHERE hcn.node_name='human_case_upload'
                  AND hcn.node_status=3
                  AND hcn.node_updated_at BETWEEN %s AND %s
                GROUP BY hc.producer, hc.project_uuid
            """, [_d_start, _d_end, _d_start, _d_end])
            for r in cur.fetchall():
                p = r["producer"]; pid = r["project_id"] or ""
                if not p: continue
                if pid in _new_wrist_uuids_cache:
                    person_data[p]["wrist_collect_qc_h"]        += float(r["qc_h"] or 0)
                    person_data[p]["wrist_collect_inspected_h"] += float(r["inspected_h"] or 0)
                else:
                    person_data[p]["pt_collect_qc_h"]           += float(r["qc_h"] or 0)
                    person_data[p]["pt_collect_inspected_h"]     += float(r["inspected_h"] or 0)
        conn2.close()
    except Exception:
        import traceback; traceback.print_exc()

    # ── 3. 按供应商汇总（只用老人，total_collect_h >= 10h）────────
    vendors_map = defaultdict(list)
    for p, d in person_data.items():
        if d["total_collect_h"] >= 10:
            vendors_map[d["vendor"]].append(d)

    def _type_avg(lst, upload_key, qc_key, inspected_key):
        """分母 = 有该类型采集数据的老人数，分子 = 这些人的数据总和"""
        active = [x for x in lst if x[upload_key] > 0]
        n = len(active)
        if n == 0:
            return {"upload_avg_h": 0, "collect_qc_avg_h": 0, "predicted_avg_h": 0, "active_count": 0}
        total_upload   = sum(x[upload_key] for x in active)
        total_qc       = sum(x[qc_key] for x in active)
        total_inspected = sum(x[inspected_key] for x in active)
        upload_avg  = round(total_upload / n, 2)
        qc_avg      = round(total_qc / n, 2)
        predicted   = round(total_upload * (total_qc / total_inspected) / n, 2) if total_inspected > 0 else 0
        return {"upload_avg_h": upload_avg, "collect_qc_avg_h": qc_avg, "predicted_avg_h": predicted, "active_count": n}

    result_vendors = []
    for vname, seniors in sorted(vendors_map.items()):
        n = len(seniors)
        wrist_stats = _type_avg(seniors, "wrist_upload_h", "wrist_collect_qc_h", "wrist_collect_inspected_h")
        pt_stats    = _type_avg(seniors, "pt_upload_h",    "pt_collect_qc_h",    "pt_collect_inspected_h")
        # 原始人均：有任意采集数据的老人，分子=全部时长之和
        active_any = [x for x in seniors if x["wrist_upload_h"] + x["pt_upload_h"] > 0]
        upload_avg_total = round(sum(x["wrist_upload_h"] + x["pt_upload_h"] for x in active_any) / len(active_any), 2) if active_any else 0
        result_vendors.append({
            "vendor":       vname,
            "senior_count": n,
            "upload_avg_h": upload_avg_total,
            "wrist": wrist_stats,
            "pt":    pt_stats,
        })

    return {"date": date_str, "vendors": result_vendors}


@app.route("/api/efficiency/daily")
def efficiency_daily():
    cron_token = os.environ.get("CRON_TOKEN", "")
    req_token  = request.args.get("token", "")
    if not session.get("logged_in") and not (cron_token and req_token == cron_token):
        return jsonify({"error": "unauthorized"}), 401

    date_str = request.args.get("date", str(date.today()))
    save = request.args.get("save", "0") == "1"

    # 先查磁盘缓存
    disk = _load_efficiency_cache()
    if date_str in disk and not save:
        return jsonify(disk[date_str])

    result = _compute_efficiency(date_str)

    if save or True:  # 每次计算都存盘，方便历史查询
        disk[date_str] = result
        _save_efficiency_cache(disk)

    return jsonify(result)


# gunicorn 导入模块时触发预热（每个 worker 独立运行，磁盘缓存防重复计算）
_warmup_cumul_cache()

if __name__ == "__main__":
    print("排期看板服务已启动 → http://localhost:8000")
    app.run(host="0.0.0.0", port=8000, debug=False, threaded=True)
